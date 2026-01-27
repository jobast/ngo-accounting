"""
CREATES NGO Accounting System
Application de comptabilité pour ONG - Conforme SYSCOHADA
"""

from flask import Flask, render_template, request, redirect, url_for, flash, jsonify, Response, send_file, send_from_directory
from flask_sqlalchemy import SQLAlchemy
from flask_login import LoginManager, UserMixin, login_user, logout_user, login_required, current_user
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename
from datetime import datetime, date, timedelta
from decimal import Decimal
from functools import wraps
import os
import json
import shutil
import glob as glob_module
from io import BytesIO
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email import encoders

# SECURITY: Rate limiting
try:
    from flask_limiter import Limiter
    from flask_limiter.util import get_remote_address
    RATE_LIMITING_ENABLED = True
except ImportError:
    RATE_LIMITING_ENABLED = False

app = Flask(__name__)

# SECURITY: Secret key configuration
_secret_key = os.environ.get('SECRET_KEY')
if not _secret_key:
    import warnings
    warnings.warn("SECRET_KEY not set! Using insecure default. Set SECRET_KEY environment variable in production.")
    _secret_key = 'dev-only-insecure-key-do-not-use-in-production'

app.config['SECRET_KEY'] = _secret_key
app.config['SQLALCHEMY_DATABASE_URI'] = os.environ.get('DATABASE_URL', 'sqlite:///ngo_accounting.db')
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['UPLOAD_FOLDER'] = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'uploads')
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB max upload

# Create upload folder if not exists
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)

# =============================================================================
# ORGANIZATION INFO
# =============================================================================
ORG_INFO = {
    'nom': 'GIE CREATES',
    'nom_complet': 'Centre de Recherche-Action sur les Transformations Ecologiques et Sociales',
    'adresse': 'Quartier Ngane, Ngaparou, derrière Sportand',
    'ville': 'Département de Mbour',
    'pays': 'Sénégal',
    'site_web': 'www.creates.ngo',
    'logo': 'static/img/logo.png'
}

# Make ORG_INFO available in all templates
@app.context_processor
def inject_org_info():
    return {'org': ORG_INFO}

db = SQLAlchemy(app)

# SECURITY: Enable SQLite foreign key enforcement
from sqlalchemy import event
from sqlalchemy.engine import Engine
import sqlite3

@event.listens_for(Engine, "connect")
def set_sqlite_pragma(dbapi_connection, connection_record):
    if isinstance(dbapi_connection, sqlite3.Connection):
        cursor = dbapi_connection.cursor()
        cursor.execute("PRAGMA foreign_keys=ON")
        cursor.close()

login_manager = LoginManager(app)
login_manager.login_view = 'login'
login_manager.login_message = 'Veuillez vous connecter pour accéder à cette page.'
login_manager.login_message_category = 'warning'

# SECURITY: Initialize rate limiter
if RATE_LIMITING_ENABLED:
    limiter = Limiter(
        key_func=get_remote_address,
        app=app,
        default_limits=["200 per day", "50 per hour"],
        storage_uri="memory://"
    )
else:
    limiter = None


# =============================================================================
# DECORATORS
# =============================================================================

def role_required(roles):
    """Décorateur pour restreindre l'accès par rôle"""
    def decorator(f):
        @wraps(f)
        def decorated_function(*args, **kwargs):
            if not current_user.is_authenticated:
                flash('Veuillez vous connecter.', 'warning')
                return redirect(url_for('login'))
            if current_user.role not in roles:
                flash('Vous n\'avez pas les permissions nécessaires.', 'danger')
                return redirect(url_for('dashboard'))
            return f(*args, **kwargs)
        return decorated_function
    return decorator


def log_audit(table_name, record_id, action, old_values=None, new_values=None):
    """Enregistrer une action dans le journal d'audit"""
    audit = AuditLog(
        table_name=table_name,
        record_id=record_id,
        action=action,
        old_values=json.dumps(old_values) if old_values else None,
        new_values=json.dumps(new_values) if new_values else None,
        user=current_user.email if current_user.is_authenticated else 'system',
        ip_address=request.remote_addr if request else None
    )
    db.session.add(audit)


@login_manager.user_loader
def load_user(user_id):
    return Utilisateur.query.get(int(user_id))

# =============================================================================
# MODELS
# =============================================================================

class Devise(db.Model):
    """Currencies / Devises"""
    __tablename__ = 'devises'

    id = db.Column(db.Integer, primary_key=True)
    code = db.Column(db.String(3), unique=True, nullable=False)  # XOF, USD, EUR
    nom = db.Column(db.String(50), nullable=False)
    symbole = db.Column(db.String(5))
    taux_base = db.Column(db.Numeric(15, 6), default=1)  # Taux vs XOF

    def __repr__(self):
        return f'<Devise {self.code}>'


class ExerciceComptable(db.Model):
    """Fiscal Year / Exercice comptable"""
    __tablename__ = 'exercices'

    id = db.Column(db.Integer, primary_key=True)
    annee = db.Column(db.Integer, unique=True, nullable=False)
    date_debut = db.Column(db.Date, nullable=False)
    date_fin = db.Column(db.Date, nullable=False)
    cloture = db.Column(db.Boolean, default=False)

    def __repr__(self):
        return f'<Exercice {self.annee}>'


class CompteComptable(db.Model):
    """Chart of Accounts / Plan comptable SYSCOHADA"""
    __tablename__ = 'comptes'

    id = db.Column(db.Integer, primary_key=True)
    numero = db.Column(db.String(10), unique=True, nullable=False)
    intitule = db.Column(db.String(200), nullable=False)
    classe = db.Column(db.Integer, nullable=False)  # 1-7 SYSCOHADA
    type_compte = db.Column(db.String(20))  # actif, passif, charge, produit
    compte_parent_id = db.Column(db.Integer, db.ForeignKey('comptes.id'))
    actif = db.Column(db.Boolean, default=True)

    # Relations
    compte_parent = db.relationship('CompteComptable', remote_side=[id], backref='sous_comptes')
    details_bancaires = db.relationship('CompteTresorerie', backref='compte_comptable', uselist=False)

    def __repr__(self):
        return f'<Compte {self.numero} - {self.intitule}>'

    @property
    def est_tresorerie(self):
        """Vérifie si c'est un compte de trésorerie (classe 5)"""
        return self.classe == 5


class CompteTresorerie(db.Model):
    """Détails des comptes de trésorerie (banques, caisses, mobile money)
    Lié à un CompteComptable de classe 5 (52x, 57x, 58x)
    """
    __tablename__ = 'comptes_tresorerie'

    id = db.Column(db.Integer, primary_key=True)
    compte_id = db.Column(db.Integer, db.ForeignKey('comptes.id'), unique=True, nullable=False)

    # Type de compte
    type_tresorerie = db.Column(db.String(20), nullable=False)  # banque, caisse, mobile_money

    # Détails bancaires (pour type=banque)
    nom_banque = db.Column(db.String(100))  # Ex: CBAO, BICIS, Ecobank
    numero_compte = db.Column(db.String(50))
    iban = db.Column(db.String(34))
    code_swift = db.Column(db.String(11))
    agence = db.Column(db.String(100))
    adresse_agence = db.Column(db.String(200))
    titulaire = db.Column(db.String(200))  # Nom du titulaire du compte

    # Détails mobile money (pour type=mobile_money)
    operateur = db.Column(db.String(50))  # Wave, Orange Money, Free Money
    numero_telephone = db.Column(db.String(20))
    nom_marchand = db.Column(db.String(100))  # Pour comptes marchands

    # Informations générales
    devise_id = db.Column(db.Integer, db.ForeignKey('devises.id'))
    solde_ouverture = db.Column(db.Numeric(15, 2), default=0)
    date_ouverture = db.Column(db.Date)
    plafond = db.Column(db.Numeric(15, 2))  # Plafond de caisse ou limite
    notes = db.Column(db.Text)
    actif = db.Column(db.Boolean, default=True)

    # Relations
    devise = db.relationship('Devise')

    def __repr__(self):
        return f'<CompteTresorerie {self.type_tresorerie} - {self.compte_comptable.numero if self.compte_comptable else "?"}>'

    @property
    def label(self):
        """Retourne un label descriptif"""
        if self.type_tresorerie == 'banque':
            return f"{self.nom_banque or 'Banque'} - {self.numero_compte[-4:] if self.numero_compte else ''}"
        elif self.type_tresorerie == 'mobile_money':
            return f"{self.operateur or 'Mobile'} - {self.numero_telephone or ''}"
        else:
            return "Caisse"


class Bailleur(db.Model):
    """Donors / Bailleurs de fonds"""
    __tablename__ = 'bailleurs'

    id = db.Column(db.Integer, primary_key=True)
    code = db.Column(db.String(20), unique=True, nullable=False)
    nom = db.Column(db.String(200), nullable=False)
    pays = db.Column(db.String(100))
    contact = db.Column(db.String(200))
    email = db.Column(db.String(100))
    devise_id = db.Column(db.Integer, db.ForeignKey('devises.id'))
    actif = db.Column(db.Boolean, default=True)

    # Relations
    devise = db.relationship('Devise')
    projets = db.relationship('Projet', back_populates='bailleur')

    def __repr__(self):
        return f'<Bailleur {self.code} - {self.nom}>'


class Projet(db.Model):
    """Projects / Projets"""
    __tablename__ = 'projets'

    id = db.Column(db.Integer, primary_key=True)
    code = db.Column(db.String(20), unique=True, nullable=False)
    nom = db.Column(db.String(200), nullable=False)
    description = db.Column(db.Text)
    bailleur_id = db.Column(db.Integer, db.ForeignKey('bailleurs.id'))
    date_debut = db.Column(db.Date)
    date_fin = db.Column(db.Date)
    budget_total = db.Column(db.Numeric(15, 2), default=0)
    devise_id = db.Column(db.Integer, db.ForeignKey('devises.id'))
    statut = db.Column(db.String(20), default='actif')  # actif, cloture, suspendu

    # Relations
    bailleur = db.relationship('Bailleur', back_populates='projets')
    devise = db.relationship('Devise')
    lignes_budget = db.relationship('LigneBudget', back_populates='projet', cascade='all, delete-orphan')

    def __repr__(self):
        return f'<Projet {self.code} - {self.nom}>'


class CategorieBudget(db.Model):
    """Budget Categories / Catégories budgétaires"""
    __tablename__ = 'categories_budget'

    id = db.Column(db.Integer, primary_key=True)
    code = db.Column(db.String(20), unique=True, nullable=False)
    nom = db.Column(db.String(100), nullable=False)
    ordre = db.Column(db.Integer, default=0)

    def __repr__(self):
        return f'<Categorie {self.code} - {self.nom}>'


class LigneBudget(db.Model):
    """Budget Lines / Lignes budgétaires"""
    __tablename__ = 'lignes_budget'

    id = db.Column(db.Integer, primary_key=True)
    projet_id = db.Column(db.Integer, db.ForeignKey('projets.id'), nullable=False)
    categorie_id = db.Column(db.Integer, db.ForeignKey('categories_budget.id'))
    code = db.Column(db.String(20), nullable=False)
    intitule = db.Column(db.String(200), nullable=False)
    annee = db.Column(db.Integer)  # Pour budget multi-années
    quantite = db.Column(db.Numeric(10, 2), default=1)
    unite = db.Column(db.String(50))
    cout_unitaire = db.Column(db.Numeric(15, 2), default=0)
    montant_prevu = db.Column(db.Numeric(15, 2), default=0)

    # Relations
    projet = db.relationship('Projet', back_populates='lignes_budget')
    categorie = db.relationship('CategorieBudget')
    budgets_annuels = db.relationship('BudgetAnnee', back_populates='ligne_budget', cascade='all, delete-orphan')

    def __repr__(self):
        return f'<LigneBudget {self.code} - {self.intitule}>'

    def get_montant_annee(self, annee):
        """Retourne le montant prévu pour une année donnée"""
        for ba in self.budgets_annuels:
            if ba.annee == annee:
                return ba.montant_prevu
        return Decimal('0')

    def get_total_prevu(self):
        """Retourne le total prévu (somme des années ou montant_prevu si pas de détail annuel)"""
        if self.budgets_annuels:
            return sum(ba.montant_prevu or 0 for ba in self.budgets_annuels)
        return self.montant_prevu or Decimal('0')


class BudgetAnnee(db.Model):
    """Budget par année / Annual Budget Breakdown"""
    __tablename__ = 'budgets_annee'

    id = db.Column(db.Integer, primary_key=True)
    ligne_budget_id = db.Column(db.Integer, db.ForeignKey('lignes_budget.id'), nullable=False)
    annee = db.Column(db.Integer, nullable=False)
    montant_prevu = db.Column(db.Numeric(15, 2), default=0)
    commentaire = db.Column(db.String(255))

    # Relations
    ligne_budget = db.relationship('LigneBudget', back_populates='budgets_annuels')

    # Contrainte unique: une seule entrée par ligne/année
    __table_args__ = (
        db.UniqueConstraint('ligne_budget_id', 'annee', name='uq_ligne_annee'),
    )

    def __repr__(self):
        return f'<BudgetAnnee {self.ligne_budget_id} - {self.annee}: {self.montant_prevu}>'


class Journal(db.Model):
    """Accounting Journals / Journaux comptables"""
    __tablename__ = 'journaux'

    id = db.Column(db.Integer, primary_key=True)
    code = db.Column(db.String(10), unique=True, nullable=False)
    nom = db.Column(db.String(100), nullable=False)
    type_journal = db.Column(db.String(20))  # achat, vente, banque, caisse, mobile_money, od

    # Lien optionnel vers un compte de trésorerie (pour journaux banque/caisse)
    compte_tresorerie_id = db.Column(db.Integer, db.ForeignKey('comptes.id'))
    compte_tresorerie = db.relationship('CompteComptable')

    def __repr__(self):
        return f'<Journal {self.code} - {self.nom}>'


class PieceComptable(db.Model):
    """Accounting Entries / Pièces comptables"""
    __tablename__ = 'pieces'

    id = db.Column(db.Integer, primary_key=True)
    numero = db.Column(db.String(20), unique=True, nullable=False)
    date_piece = db.Column(db.Date, nullable=False)
    journal_id = db.Column(db.Integer, db.ForeignKey('journaux.id'), nullable=False)
    exercice_id = db.Column(db.Integer, db.ForeignKey('exercices.id'), nullable=False)
    libelle = db.Column(db.String(200), nullable=False)
    reference = db.Column(db.String(100))  # Numéro facture, chèque, etc.
    devise_id = db.Column(db.Integer, db.ForeignKey('devises.id'))
    taux_change = db.Column(db.Numeric(15, 6), default=1)
    valide = db.Column(db.Boolean, default=False)
    date_creation = db.Column(db.DateTime, default=datetime.utcnow)

    # Relations
    journal = db.relationship('Journal')
    exercice = db.relationship('ExerciceComptable', backref='pieces')
    devise = db.relationship('Devise')
    lignes = db.relationship('LigneEcriture', back_populates='piece', cascade='all, delete-orphan')

    def __repr__(self):
        return f'<Piece {self.numero} - {self.libelle}>'

    @property
    def total_debit(self):
        return sum(l.debit or 0 for l in self.lignes)

    @property
    def total_credit(self):
        return sum(l.credit or 0 for l in self.lignes)

    @property
    def est_equilibree(self):
        return abs(self.total_debit - self.total_credit) < 0.01


class LigneEcriture(db.Model):
    """Journal Entry Lines / Lignes d'écriture"""
    __tablename__ = 'lignes_ecriture'

    id = db.Column(db.Integer, primary_key=True)
    piece_id = db.Column(db.Integer, db.ForeignKey('pieces.id'), nullable=False)
    compte_id = db.Column(db.Integer, db.ForeignKey('comptes.id'), nullable=False)
    projet_id = db.Column(db.Integer, db.ForeignKey('projets.id'))
    ligne_budget_id = db.Column(db.Integer, db.ForeignKey('lignes_budget.id'))
    libelle = db.Column(db.String(200))
    debit = db.Column(db.Numeric(15, 2), default=0)
    credit = db.Column(db.Numeric(15, 2), default=0)

    # Relations
    piece = db.relationship('PieceComptable', back_populates='lignes')
    compte = db.relationship('CompteComptable')
    projet = db.relationship('Projet')
    ligne_budget = db.relationship('LigneBudget')

    def __repr__(self):
        return f'<Ligne {self.compte.numero if self.compte else ""} D:{self.debit} C:{self.credit}>'


class ImputationAnalytique(db.Model):
    """Analytical Imputation / Ventilation analytique multi-projets

    Permet de répartir une charge sur plusieurs projets.
    Ex: Loyer 60% LED + 40% SOR4D
    """
    __tablename__ = 'imputations_analytiques'

    id = db.Column(db.Integer, primary_key=True)
    ligne_ecriture_id = db.Column(db.Integer, db.ForeignKey('lignes_ecriture.id'), nullable=False)
    projet_id = db.Column(db.Integer, db.ForeignKey('projets.id'), nullable=False)
    ligne_budget_id = db.Column(db.Integer, db.ForeignKey('lignes_budget.id'))
    pourcentage = db.Column(db.Numeric(5, 2))  # Ex: 60.00 pour 60%
    montant = db.Column(db.Numeric(15, 2))     # Montant calculé

    # Relations
    ligne_ecriture = db.relationship('LigneEcriture', backref='imputations_analytiques')
    projet = db.relationship('Projet')
    ligne_budget = db.relationship('LigneBudget')

    def __repr__(self):
        return f'<Imputation {self.projet.code if self.projet else ""} {self.pourcentage}%>'


class AuditLog(db.Model):
    """Journal d'audit - Traçabilité des modifications"""
    __tablename__ = 'audit_log'

    id = db.Column(db.Integer, primary_key=True)
    table_name = db.Column(db.String(50), nullable=False)
    record_id = db.Column(db.Integer)
    action = db.Column(db.String(20), nullable=False)  # CREATE, UPDATE, DELETE
    old_values = db.Column(db.Text)    # JSON des anciennes valeurs
    new_values = db.Column(db.Text)    # JSON des nouvelles valeurs
    user = db.Column(db.String(100))
    ip_address = db.Column(db.String(45))
    timestamp = db.Column(db.DateTime, default=datetime.utcnow)

    def __repr__(self):
        return f'<AuditLog {self.table_name} {self.action} {self.timestamp}>'


class PieceJustificative(db.Model):
    """Pièces justificatives / Supporting Documents

    Permet d'attacher des documents scannés aux écritures.
    """
    __tablename__ = 'pieces_justificatives'

    id = db.Column(db.Integer, primary_key=True)
    ligne_ecriture_id = db.Column(db.Integer, db.ForeignKey('lignes_ecriture.id'))
    piece_comptable_id = db.Column(db.Integer, db.ForeignKey('pieces.id'))
    type_piece = db.Column(db.String(50))  # facture, recu, contrat, bon_commande
    numero_piece = db.Column(db.String(50))
    fichier_path = db.Column(db.String(255))  # Chemin du fichier
    fichier_nom = db.Column(db.String(255))   # Nom original du fichier
    date_piece = db.Column(db.Date)
    description = db.Column(db.String(255))
    date_upload = db.Column(db.DateTime, default=datetime.utcnow)
    uploaded_by = db.Column(db.String(100))

    # Relations
    ligne_ecriture = db.relationship('LigneEcriture', backref='pieces_justificatives')
    piece_comptable = db.relationship('PieceComptable', backref='pieces_justificatives')

    def __repr__(self):
        return f'<PieceJustificative {self.type_piece} {self.numero_piece}>'


class Utilisateur(UserMixin, db.Model):
    """Utilisateurs avec rôles et permissions"""
    __tablename__ = 'utilisateurs'

    id = db.Column(db.Integer, primary_key=True)
    email = db.Column(db.String(120), unique=True, nullable=False)
    nom = db.Column(db.String(100), nullable=False)
    prenom = db.Column(db.String(100))
    password_hash = db.Column(db.String(256))
    role = db.Column(db.String(20), default='comptable')  # comptable, directeur, auditeur
    actif = db.Column(db.Boolean, default=True)
    derniere_connexion = db.Column(db.DateTime)
    date_creation = db.Column(db.DateTime, default=datetime.utcnow)
    created_by = db.Column(db.String(100))

    # Permissions par rôle
    ROLES_PERMISSIONS = {
        'comptable': ['saisie_ecritures', 'voir_rapports', 'gerer_projets'],
        'directeur': ['saisie_ecritures', 'voir_rapports', 'gerer_projets',
                      'valider_ecritures', 'cloturer_exercice', 'gerer_utilisateurs'],
        'auditeur': ['voir_rapports', 'voir_audit_trail', 'export_donnees']
    }

    def has_permission(self, permission):
        """Vérifie si l'utilisateur a une permission donnée"""
        return permission in self.ROLES_PERMISSIONS.get(self.role, [])

    def __repr__(self):
        return f'<Utilisateur {self.email} ({self.role})>'


class Alerte(db.Model):
    """Alertes système pour le suivi budgétaire"""
    __tablename__ = 'alertes'

    id = db.Column(db.Integer, primary_key=True)
    type_alerte = db.Column(db.String(50), nullable=False)  # budget_80, ecritures_non_validees, solde_negatif
    niveau = db.Column(db.String(20), default='warning')  # info, warning, danger
    message = db.Column(db.String(500), nullable=False)
    projet_id = db.Column(db.Integer, db.ForeignKey('projets.id'))
    compte_id = db.Column(db.Integer, db.ForeignKey('comptes.id'))
    date_creation = db.Column(db.DateTime, default=datetime.utcnow)
    date_lecture = db.Column(db.DateTime)
    lu_par = db.Column(db.String(100))
    active = db.Column(db.Boolean, default=True)

    # Relations
    projet = db.relationship('Projet')
    compte = db.relationship('CompteComptable')

    def __repr__(self):
        return f'<Alerte {self.type_alerte} {self.niveau}>'


# =============================================================================
# PHASE 1 - TRÉSORERIE & RAPPROCHEMENT
# =============================================================================

class ReconciliationBancaire(db.Model):
    """Réconciliation Bancaire - Manuel Section 3.10.2"""
    __tablename__ = 'reconciliations_bancaires'

    id = db.Column(db.Integer, primary_key=True)
    compte_id = db.Column(db.Integer, db.ForeignKey('comptes.id'), nullable=False)
    date_reconciliation = db.Column(db.Date, nullable=False)
    periode_debut = db.Column(db.Date, nullable=False)
    periode_fin = db.Column(db.Date, nullable=False)
    solde_releve = db.Column(db.Numeric(15, 2), default=0)
    solde_comptable = db.Column(db.Numeric(15, 2), default=0)
    ecart = db.Column(db.Numeric(15, 2), default=0)
    statut = db.Column(db.String(20), default='en_cours')  # en_cours, validee
    cree_par = db.Column(db.String(100))
    valide_par = db.Column(db.String(100))
    date_validation = db.Column(db.DateTime)
    date_creation = db.Column(db.DateTime, default=datetime.utcnow)
    notes = db.Column(db.Text)

    # Relations
    compte = db.relationship('CompteComptable')
    lignes = db.relationship('LigneReconciliation', back_populates='reconciliation', cascade='all, delete-orphan')

    def __repr__(self):
        return f'<Reconciliation {self.compte.numero if self.compte else ""} {self.date_reconciliation}>'

    @property
    def nb_pointees(self):
        return sum(1 for l in self.lignes if l.pointee)

    @property
    def nb_non_pointees(self):
        return sum(1 for l in self.lignes if not l.pointee)


class LigneReconciliation(db.Model):
    """Lignes de réconciliation bancaire"""
    __tablename__ = 'lignes_reconciliation'

    id = db.Column(db.Integer, primary_key=True)
    reconciliation_id = db.Column(db.Integer, db.ForeignKey('reconciliations_bancaires.id'), nullable=False)
    ligne_ecriture_id = db.Column(db.Integer, db.ForeignKey('lignes_ecriture.id'), nullable=False)
    pointee = db.Column(db.Boolean, default=False)
    date_pointage = db.Column(db.DateTime)
    pointe_par = db.Column(db.String(100))

    # Relations
    reconciliation = db.relationship('ReconciliationBancaire', back_populates='lignes')
    ligne_ecriture = db.relationship('LigneEcriture')

    def __repr__(self):
        return f'<LigneReconciliation {self.id} pointee={self.pointee}>'


class Avance(db.Model):
    """Gestion des Avances - Manuel Section 3.11.1
    Justification sous 7 jours, sinon déduction salaire
    """
    __tablename__ = 'avances'

    id = db.Column(db.Integer, primary_key=True)
    numero = db.Column(db.String(20), unique=True, nullable=False)
    date_avance = db.Column(db.Date, nullable=False)
    beneficiaire = db.Column(db.String(100), nullable=False)
    montant = db.Column(db.Numeric(15, 2), nullable=False)
    objet = db.Column(db.String(255), nullable=False)
    projet_id = db.Column(db.Integer, db.ForeignKey('projets.id'))
    statut = db.Column(db.String(20), default='en_attente')  # en_attente, justifiee, soldee, deduite
    date_limite = db.Column(db.Date)  # +7 jours
    montant_justifie = db.Column(db.Numeric(15, 2), default=0)
    montant_rembourse = db.Column(db.Numeric(15, 2), default=0)
    piece_comptable_id = db.Column(db.Integer, db.ForeignKey('pieces.id'))
    piece_justification_id = db.Column(db.Integer, db.ForeignKey('pieces.id'))
    justification_notes = db.Column(db.Text)
    cree_par = db.Column(db.String(100))
    date_creation = db.Column(db.DateTime, default=datetime.utcnow)
    date_justification = db.Column(db.DateTime)
    date_solde = db.Column(db.DateTime)

    # Relations
    projet = db.relationship('Projet')
    piece_comptable = db.relationship('PieceComptable', foreign_keys=[piece_comptable_id])
    piece_justification = db.relationship('PieceComptable', foreign_keys=[piece_justification_id])

    def __repr__(self):
        return f'<Avance {self.numero} {self.beneficiaire} {self.montant}>'

    @property
    def est_en_retard(self):
        if self.statut == 'en_attente' and self.date_limite:
            return date.today() > self.date_limite
        return False

    @property
    def jours_retard(self):
        if self.est_en_retard:
            return (date.today() - self.date_limite).days
        return 0

    @property
    def solde_restant(self):
        return float(self.montant or 0) - float(self.montant_justifie or 0) - float(self.montant_rembourse or 0)


class Immobilisation(db.Model):
    """Registre des Immobilisations - Manuel Section 3.6
    Durées SYSCOA: Informatique 3 ans, véhicules 5 ans, mobilier 10 ans, bâtiments 20 ans
    """
    __tablename__ = 'immobilisations'

    id = db.Column(db.Integer, primary_key=True)
    code = db.Column(db.String(20), unique=True, nullable=False)
    designation = db.Column(db.String(200), nullable=False)
    categorie = db.Column(db.String(50), nullable=False)  # informatique, vehicule, mobilier, batiment
    date_acquisition = db.Column(db.Date, nullable=False)
    valeur_acquisition = db.Column(db.Numeric(15, 2), nullable=False)
    duree_amortissement = db.Column(db.Integer, nullable=False)  # en années
    taux_amortissement = db.Column(db.Numeric(5, 2))  # calculé automatiquement
    compte_immobilisation_id = db.Column(db.Integer, db.ForeignKey('comptes.id'))
    compte_amortissement_id = db.Column(db.Integer, db.ForeignKey('comptes.id'))
    compte_dotation_id = db.Column(db.Integer, db.ForeignKey('comptes.id'))
    projet_id = db.Column(db.Integer, db.ForeignKey('projets.id'))
    localisation = db.Column(db.String(100))
    numero_serie = db.Column(db.String(100))
    fournisseur = db.Column(db.String(200))
    numero_facture = db.Column(db.String(50))
    statut = db.Column(db.String(20), default='actif')  # actif, cede, rebut
    date_sortie = db.Column(db.Date)
    motif_sortie = db.Column(db.String(100))
    valeur_cession = db.Column(db.Numeric(15, 2))
    cree_par = db.Column(db.String(100))
    date_creation = db.Column(db.DateTime, default=datetime.utcnow)

    # Relations
    compte_immobilisation = db.relationship('CompteComptable', foreign_keys=[compte_immobilisation_id])
    compte_amortissement = db.relationship('CompteComptable', foreign_keys=[compte_amortissement_id])
    compte_dotation = db.relationship('CompteComptable', foreign_keys=[compte_dotation_id])
    projet = db.relationship('Projet')
    lignes_amortissement = db.relationship('LigneAmortissement', back_populates='immobilisation', cascade='all, delete-orphan')

    # Durées standard SYSCOA
    DUREES_SYSCOA = {
        'informatique': 3,
        'vehicule': 5,
        'mobilier': 10,
        'batiment': 20
    }

    def __repr__(self):
        return f'<Immobilisation {self.code} {self.designation}>'

    @property
    def amortissement_annuel(self):
        if self.duree_amortissement and self.duree_amortissement > 0:
            return float(self.valeur_acquisition) / self.duree_amortissement
        return 0

    @property
    def cumul_amortissement(self):
        return sum(float(l.dotation or 0) for l in self.lignes_amortissement)

    @property
    def valeur_nette_comptable(self):
        return float(self.valeur_acquisition or 0) - self.cumul_amortissement


class LigneAmortissement(db.Model):
    """Tableau d'amortissement par immobilisation"""
    __tablename__ = 'lignes_amortissement'

    id = db.Column(db.Integer, primary_key=True)
    immobilisation_id = db.Column(db.Integer, db.ForeignKey('immobilisations.id'), nullable=False)
    exercice_id = db.Column(db.Integer, db.ForeignKey('exercices.id'), nullable=False)
    annee = db.Column(db.Integer, nullable=False)
    dotation = db.Column(db.Numeric(15, 2), nullable=False)
    cumul = db.Column(db.Numeric(15, 2), nullable=False)
    valeur_nette = db.Column(db.Numeric(15, 2), nullable=False)
    piece_comptable_id = db.Column(db.Integer, db.ForeignKey('pieces.id'))
    date_calcul = db.Column(db.DateTime, default=datetime.utcnow)

    # Relations
    immobilisation = db.relationship('Immobilisation', back_populates='lignes_amortissement')
    exercice = db.relationship('ExerciceComptable')
    piece_comptable = db.relationship('PieceComptable')

    def __repr__(self):
        return f'<LigneAmortissement {self.annee} {self.dotation}>'


class TauxChange(db.Model):
    """Taux de change mensuels - Manuel Section 3.4
    Taux de change moyen mensuel BCEAO
    """
    __tablename__ = 'taux_change'

    id = db.Column(db.Integer, primary_key=True)
    devise_id = db.Column(db.Integer, db.ForeignKey('devises.id'), nullable=False)
    mois = db.Column(db.Integer, nullable=False)
    annee = db.Column(db.Integer, nullable=False)
    taux = db.Column(db.Numeric(15, 6), nullable=False)
    source = db.Column(db.String(50), default='BCEAO')
    date_saisie = db.Column(db.DateTime, default=datetime.utcnow)
    saisi_par = db.Column(db.String(100))

    # Relations
    devise = db.relationship('Devise')

    # Contrainte unique
    __table_args__ = (
        db.UniqueConstraint('devise_id', 'mois', 'annee', name='uq_taux_devise_periode'),
    )

    def __repr__(self):
        return f'<TauxChange {self.devise.code if self.devise else ""} {self.mois}/{self.annee} {self.taux}>'


class ModeleEcriture(db.Model):
    """Modèles d'écritures récurrentes
    Pour: Loyer mensuel, salaires, abonnements
    """
    __tablename__ = 'modeles_ecritures'

    id = db.Column(db.Integer, primary_key=True)
    nom = db.Column(db.String(100), nullable=False)
    description = db.Column(db.String(255))
    journal_id = db.Column(db.Integer, db.ForeignKey('journaux.id'), nullable=False)
    libelle = db.Column(db.String(200), nullable=False)
    frequence = db.Column(db.String(20))  # mensuel, trimestriel, annuel
    jour_execution = db.Column(db.Integer)  # jour du mois
    actif = db.Column(db.Boolean, default=True)
    derniere_execution = db.Column(db.Date)
    cree_par = db.Column(db.String(100))
    date_creation = db.Column(db.DateTime, default=datetime.utcnow)

    # Relations
    journal = db.relationship('Journal')
    lignes = db.relationship('LigneModeleEcriture', back_populates='modele', cascade='all, delete-orphan')

    def __repr__(self):
        return f'<ModeleEcriture {self.nom}>'


class LigneModeleEcriture(db.Model):
    """Lignes d'un modèle d'écriture"""
    __tablename__ = 'lignes_modele_ecriture'

    id = db.Column(db.Integer, primary_key=True)
    modele_id = db.Column(db.Integer, db.ForeignKey('modeles_ecritures.id'), nullable=False)
    compte_id = db.Column(db.Integer, db.ForeignKey('comptes.id'), nullable=False)
    projet_id = db.Column(db.Integer, db.ForeignKey('projets.id'))
    libelle = db.Column(db.String(200))
    type_montant = db.Column(db.String(10))  # debit, credit
    montant = db.Column(db.Numeric(15, 2), default=0)
    formule = db.Column(db.String(100))  # Pour montants calculés

    # Relations
    modele = db.relationship('ModeleEcriture', back_populates='lignes')
    compte = db.relationship('CompteComptable')
    projet = db.relationship('Projet')

    def __repr__(self):
        return f'<LigneModele {self.compte.numero if self.compte else ""} {self.type_montant} {self.montant}>'


class Fournisseur(db.Model):
    """Registre des Fournisseurs
    Pour traçabilité et suivi des paiements par fournisseur
    """
    __tablename__ = 'fournisseurs'

    id = db.Column(db.Integer, primary_key=True)
    code = db.Column(db.String(20), unique=True, nullable=False)  # Ex: FRN001
    nom = db.Column(db.String(200), nullable=False)
    categorie = db.Column(db.String(50))  # fournitures, services, loyer, telecom, etc.
    contact = db.Column(db.String(100))
    telephone = db.Column(db.String(20))
    email = db.Column(db.String(100))
    adresse = db.Column(db.String(255))
    ville = db.Column(db.String(100))
    ninea = db.Column(db.String(20))  # Numéro d'identification fiscale Sénégal
    compte_comptable_id = db.Column(db.Integer, db.ForeignKey('comptes.id'))  # Compte 401xxx
    notes = db.Column(db.Text)
    actif = db.Column(db.Boolean, default=True)
    cree_par = db.Column(db.String(100))
    date_creation = db.Column(db.DateTime, default=datetime.utcnow)

    # Relations
    compte_comptable = db.relationship('CompteComptable')

    def __repr__(self):
        return f'<Fournisseur {self.code} - {self.nom}>'

    @property
    def total_paye(self):
        """Calcule le total payé à ce fournisseur (crédits sur compte 401)"""
        if not self.compte_comptable_id:
            return 0
        total = db.session.query(db.func.sum(LigneEcriture.credit)).filter(
            LigneEcriture.compte_id == self.compte_comptable_id
        ).scalar()
        return float(total or 0)


# =============================================================================
# MODULES NOTES DE FRAIS & DEMANDES D'ACHAT
# =============================================================================

class NoteFrais(db.Model):
    """Notes de frais employés - Expense Reports
    Workflow: brouillon -> soumis -> approuve/rejete -> rembourse
    """
    __tablename__ = 'notes_frais'

    id = db.Column(db.Integer, primary_key=True)
    numero = db.Column(db.String(20), unique=True, nullable=False)  # NF-2026-001
    employe_id = db.Column(db.Integer, db.ForeignKey('utilisateurs.id'), nullable=False)
    projet_id = db.Column(db.Integer, db.ForeignKey('projets.id'))
    ligne_budget_id = db.Column(db.Integer, db.ForeignKey('lignes_budget.id'))
    date_depense = db.Column(db.Date, nullable=False)
    montant = db.Column(db.Numeric(15, 2), nullable=False)
    categorie = db.Column(db.String(50), nullable=False)  # transport, repas, fournitures, hebergement, communication, autre
    description = db.Column(db.Text, nullable=False)
    justificatif = db.Column(db.String(255))  # chemin fichier uploadé
    statut = db.Column(db.String(20), default='brouillon')  # brouillon, soumis, approuve, rejete, rembourse
    date_soumission = db.Column(db.DateTime)
    validateur_id = db.Column(db.Integer, db.ForeignKey('utilisateurs.id'))
    date_validation = db.Column(db.DateTime)
    motif_rejet = db.Column(db.Text)
    piece_comptable_id = db.Column(db.Integer, db.ForeignKey('pieces.id'))
    a_rembourser = db.Column(db.Boolean, default=True)
    date_remboursement = db.Column(db.DateTime)
    cree_par = db.Column(db.String(100))
    date_creation = db.Column(db.DateTime, default=datetime.utcnow)

    # Relations
    employe = db.relationship('Utilisateur', foreign_keys=[employe_id], backref='notes_frais')
    validateur = db.relationship('Utilisateur', foreign_keys=[validateur_id])
    projet = db.relationship('Projet')
    ligne_budget = db.relationship('LigneBudget')
    piece_comptable = db.relationship('PieceComptable')

    # Catégories de dépenses
    CATEGORIES = [
        ('transport', 'Transport'),
        ('repas', 'Repas & Restauration'),
        ('hebergement', 'Hébergement'),
        ('fournitures', 'Fournitures de bureau'),
        ('communication', 'Communication & Téléphone'),
        ('autre', 'Autre')
    ]

    def __repr__(self):
        return f'<NoteFrais {self.numero} {self.montant}>'

    @property
    def est_modifiable(self):
        return self.statut in ['brouillon', 'rejete']

    @property
    def peut_soumettre(self):
        return self.statut in ['brouillon', 'rejete']

    @property
    def peut_approuver(self):
        return self.statut == 'soumis'


class DemandeAchat(db.Model):
    """Demandes d'achat avec workflow d'approbation
    Workflow: brouillon -> soumis -> approuve/rejete -> commande
    """
    __tablename__ = 'demandes_achat'

    id = db.Column(db.Integer, primary_key=True)
    numero = db.Column(db.String(20), unique=True, nullable=False)  # DA-2026-001
    demandeur_id = db.Column(db.Integer, db.ForeignKey('utilisateurs.id'), nullable=False)
    projet_id = db.Column(db.Integer, db.ForeignKey('projets.id'))
    ligne_budget_id = db.Column(db.Integer, db.ForeignKey('lignes_budget.id'))
    date_demande = db.Column(db.Date, nullable=False)
    objet = db.Column(db.String(200), nullable=False)
    description = db.Column(db.Text)
    montant_estime = db.Column(db.Numeric(15, 2), default=0)
    urgence = db.Column(db.String(20), default='normal')  # normal, urgent, tres_urgent
    statut = db.Column(db.String(20), default='brouillon')  # brouillon, soumis, approuve, rejete, commande
    approbateur_id = db.Column(db.Integer, db.ForeignKey('utilisateurs.id'))
    date_approbation = db.Column(db.DateTime)
    motif_rejet = db.Column(db.Text)
    bon_commande_id = db.Column(db.Integer, db.ForeignKey('bons_commande.id'))
    date_soumission = db.Column(db.DateTime)
    cree_par = db.Column(db.String(100))
    date_creation = db.Column(db.DateTime, default=datetime.utcnow)

    # Seuil d'approbation: en dessous = comptable, au dessus = directeur
    SEUIL_APPROBATION_DIRECTEUR = 500000  # 500,000 FCFA

    # Relations
    demandeur = db.relationship('Utilisateur', foreign_keys=[demandeur_id], backref='demandes_achat')
    approbateur = db.relationship('Utilisateur', foreign_keys=[approbateur_id])
    projet = db.relationship('Projet')
    ligne_budget = db.relationship('LigneBudget')
    lignes = db.relationship('LigneDemandeAchat', back_populates='demande', cascade='all, delete-orphan')
    bon_commande = db.relationship('BonCommande', foreign_keys=[bon_commande_id])

    # Niveaux d'urgence
    URGENCES = [
        ('normal', 'Normal'),
        ('urgent', 'Urgent'),
        ('tres_urgent', 'Très urgent')
    ]

    def __repr__(self):
        return f'<DemandeAchat {self.numero} {self.objet}>'

    @property
    def montant_total(self):
        return sum(float(l.montant_total or 0) for l in self.lignes)

    @property
    def est_modifiable(self):
        return self.statut in ['brouillon', 'rejete']

    @property
    def peut_soumettre(self):
        return self.statut in ['brouillon', 'rejete'] and len(self.lignes) > 0

    @property
    def peut_approuver(self):
        return self.statut == 'soumis'

    @property
    def necessite_approbation_directeur(self):
        return self.montant_total >= self.SEUIL_APPROBATION_DIRECTEUR


class LigneDemandeAchat(db.Model):
    """Lignes d'une demande d'achat"""
    __tablename__ = 'lignes_demande_achat'

    id = db.Column(db.Integer, primary_key=True)
    demande_id = db.Column(db.Integer, db.ForeignKey('demandes_achat.id'), nullable=False)
    designation = db.Column(db.String(200), nullable=False)
    quantite = db.Column(db.Numeric(10, 2), default=1)
    unite = db.Column(db.String(20), default='pièce')  # pièce, kg, litre, forfait, mois
    prix_unitaire_estime = db.Column(db.Numeric(15, 2), default=0)

    # Relations
    demande = db.relationship('DemandeAchat', back_populates='lignes')

    # Unités disponibles
    UNITES = [
        ('piece', 'Pièce'),
        ('kg', 'Kilogramme'),
        ('litre', 'Litre'),
        ('forfait', 'Forfait'),
        ('mois', 'Mois'),
        ('jour', 'Jour'),
        ('heure', 'Heure')
    ]

    def __repr__(self):
        return f'<LigneDemandeAchat {self.designation} x{self.quantite}>'

    @property
    def montant_total(self):
        return float(self.quantite or 0) * float(self.prix_unitaire_estime or 0)


class BonCommande(db.Model):
    """Bons de commande générés à partir des demandes approuvées"""
    __tablename__ = 'bons_commande'

    id = db.Column(db.Integer, primary_key=True)
    numero = db.Column(db.String(20), unique=True, nullable=False)  # BC-2026-001
    demande_achat_id = db.Column(db.Integer, db.ForeignKey('demandes_achat.id'))
    fournisseur_id = db.Column(db.Integer, db.ForeignKey('fournisseurs.id'))
    date_commande = db.Column(db.Date, nullable=False)
    date_livraison_prevue = db.Column(db.Date)
    statut = db.Column(db.String(20), default='emis')  # emis, livre_partiel, livre, facture, annule
    montant_total = db.Column(db.Numeric(15, 2), default=0)
    conditions_paiement = db.Column(db.String(100))
    adresse_livraison = db.Column(db.String(255))
    notes = db.Column(db.Text)
    cree_par = db.Column(db.String(100))
    date_creation = db.Column(db.DateTime, default=datetime.utcnow)

    # Relations
    demande_achat = db.relationship('DemandeAchat', foreign_keys=[demande_achat_id], backref='bons_commande_generes')
    fournisseur = db.relationship('Fournisseur')
    lignes = db.relationship('LigneBonCommande', back_populates='bon_commande', cascade='all, delete-orphan')

    def __repr__(self):
        return f'<BonCommande {self.numero}>'

    @property
    def est_modifiable(self):
        return self.statut == 'emis'


class LigneBonCommande(db.Model):
    """Lignes d'un bon de commande"""
    __tablename__ = 'lignes_bon_commande'

    id = db.Column(db.Integer, primary_key=True)
    bon_commande_id = db.Column(db.Integer, db.ForeignKey('bons_commande.id'), nullable=False)
    designation = db.Column(db.String(200), nullable=False)
    quantite = db.Column(db.Numeric(10, 2), default=1)
    unite = db.Column(db.String(20), default='pièce')
    prix_unitaire = db.Column(db.Numeric(15, 2), default=0)
    quantite_livree = db.Column(db.Numeric(10, 2), default=0)

    # Relations
    bon_commande = db.relationship('BonCommande', back_populates='lignes')

    def __repr__(self):
        return f'<LigneBonCommande {self.designation}>'

    @property
    def montant_total(self):
        return float(self.quantite or 0) * float(self.prix_unitaire or 0)


class ConfigBackup(db.Model):
    """Configuration des backups distants (email, cloud, etc.)"""
    __tablename__ = 'config_backup'

    id = db.Column(db.Integer, primary_key=True)
    type_destination = db.Column(db.String(20), nullable=False)  # email, gdrive, sftp
    actif = db.Column(db.Boolean, default=False)

    # Config Email
    smtp_server = db.Column(db.String(100))
    smtp_port = db.Column(db.Integer, default=587)
    smtp_user = db.Column(db.String(100))
    smtp_password = db.Column(db.String(200))  # En clair pour simplifier (à chiffrer en prod)
    email_destinataire = db.Column(db.String(200))

    date_modification = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)

    def __repr__(self):
        return f'<ConfigBackup {self.type_destination}>'


class Financement(db.Model):
    """Dons et subventions des bailleurs"""
    __tablename__ = 'financements'

    id = db.Column(db.Integer, primary_key=True)
    reference = db.Column(db.String(50))  # Référence interne ou du bailleur
    bailleur_id = db.Column(db.Integer, db.ForeignKey('bailleurs.id'), nullable=False)

    # Type d'affectation: libre, projet, usage
    type_affectation = db.Column(db.String(20), default='libre')
    projet_id = db.Column(db.Integer, db.ForeignKey('projets.id'), nullable=True)
    affectation_libelle = db.Column(db.String(200), nullable=True)  # Ex: "Terrain et bâtiment"

    # Montant et devise
    montant = db.Column(db.Numeric(15, 2), nullable=False)
    devise_id = db.Column(db.Integer, db.ForeignKey('devises.id'))

    # Dates
    date_accord = db.Column(db.Date)
    date_fin = db.Column(db.Date, nullable=True)

    statut = db.Column(db.String(20), default='actif')  # actif, cloture, annule
    notes = db.Column(db.Text)

    date_creation = db.Column(db.DateTime, default=datetime.utcnow)

    # Relations
    bailleur = db.relationship('Bailleur', backref='financements')
    projet = db.relationship('Projet', backref='financements')
    devise = db.relationship('Devise')
    tranches = db.relationship('TrancheFinancement', back_populates='financement',
                               cascade='all, delete-orphan', order_by='TrancheFinancement.numero')

    def __repr__(self):
        return f'<Financement {self.reference} - {self.bailleur.nom if self.bailleur else "?"}>'

    @property
    def montant_recu(self):
        """Total des montants reçus sur toutes les tranches"""
        return sum(float(t.montant_recu or 0) for t in self.tranches)

    @property
    def montant_attendu(self):
        """Total des montants encore attendus"""
        return float(self.montant or 0) - self.montant_recu

    @property
    def pourcentage_recu(self):
        """Pourcentage du financement reçu"""
        if not self.montant or self.montant == 0:
            return 0
        return round((self.montant_recu / float(self.montant)) * 100, 1)

    @property
    def prochaine_tranche(self):
        """Retourne la prochaine tranche attendue"""
        for t in self.tranches:
            if t.statut in ('attendu', 'retard'):
                return t
        return None


class TrancheFinancement(db.Model):
    """Échéancier des versements d'un financement"""
    __tablename__ = 'tranches_financement'

    id = db.Column(db.Integer, primary_key=True)
    financement_id = db.Column(db.Integer, db.ForeignKey('financements.id'), nullable=False)
    numero = db.Column(db.Integer, default=1)

    montant_prevu = db.Column(db.Numeric(15, 2), nullable=False)
    date_prevue = db.Column(db.Date)

    montant_recu = db.Column(db.Numeric(15, 2), default=0)
    date_reception = db.Column(db.Date, nullable=True)

    # Lien vers l'écriture comptable quand reçu
    piece_comptable_id = db.Column(db.Integer, db.ForeignKey('pieces.id'), nullable=True)

    statut = db.Column(db.String(20), default='attendu')  # attendu, recu, partiel, retard

    # Relations
    financement = db.relationship('Financement', back_populates='tranches')
    piece_comptable = db.relationship('PieceComptable')

    def __repr__(self):
        return f'<Tranche {self.numero} - {self.financement.reference if self.financement else "?"}>'

    @property
    def est_en_retard(self):
        """Vérifie si la tranche est en retard"""
        if self.statut == 'recu':
            return False
        if self.date_prevue and self.date_prevue < date.today():
            return True
        return False


# =============================================================================
# HELPER FUNCTIONS
# =============================================================================

def generer_alertes():
    """Génère les alertes système automatiques"""
    alertes = []

    # Alerte: Projets > 80% budget consommé
    projets = Projet.query.filter_by(statut='actif').all()
    for projet in projets:
        total_prevu = sum(float(l.montant_prevu or 0) for l in projet.lignes_budget)
        if total_prevu > 0:
            total_realise = 0
            for ligne in projet.lignes_budget:
                realise = db.session.query(
                    db.func.sum(LigneEcriture.debit)
                ).join(CompteComptable).filter(
                    (LigneEcriture.ligne_budget_id == ligne.id) &
                    (CompteComptable.classe == 6)
                ).scalar() or 0
                total_realise += float(realise)

            taux = (total_realise / total_prevu) * 100
            if taux > 80:
                alertes.append({
                    'type': 'budget_80',
                    'niveau': 'danger' if taux > 100 else 'warning',
                    'message': f"Projet {projet.code}: {taux:.0f}% du budget consommé",
                    'projet': projet
                })

    # Alerte: Écritures non validées > 7 jours
    date_limite = datetime.utcnow() - timedelta(days=7)
    ecritures_non_validees = PieceComptable.query.filter(
        PieceComptable.valide == False,
        PieceComptable.date_creation < date_limite
    ).count()
    if ecritures_non_validees > 0:
        alertes.append({
            'type': 'ecritures_non_validees',
            'niveau': 'warning',
            'message': f"{ecritures_non_validees} écriture(s) non validée(s) depuis plus de 7 jours",
            'projet': None
        })

    # Alerte: Solde bancaire négatif (comptes classe 5)
    comptes_banque = db.session.query(
        CompteComptable.numero,
        CompteComptable.intitule,
        db.func.sum(LigneEcriture.debit).label('total_debit'),
        db.func.sum(LigneEcriture.credit).label('total_credit')
    ).join(
        LigneEcriture, LigneEcriture.compte_id == CompteComptable.id
    ).filter(
        CompteComptable.classe == 5
    ).group_by(CompteComptable.id).all()

    for compte in comptes_banque:
        solde = float(compte.total_debit or 0) - float(compte.total_credit or 0)
        if solde < 0:
            alertes.append({
                'type': 'solde_negatif',
                'niveau': 'danger',
                'message': f"Compte {compte.numero} ({compte.intitule}): solde négatif de {solde:,.0f} FCFA",
                'projet': None
            })

    # Alerte: Avances non justifiées > 7 jours
    avances_retard = Avance.query.filter(
        Avance.statut == 'en_attente',
        Avance.date_limite < date.today()
    ).count()
    if avances_retard > 0:
        alertes.append({
            'type': 'avances_retard',
            'niveau': 'danger',
            'message': f"{avances_retard} avance(s) non justifiée(s) depuis plus de 7 jours (déduction salaire applicable)",
            'projet': None
        })

    return alertes


def calculer_stats_dashboard():
    """Calcule les statistiques pour le dashboard"""
    projets = Projet.query.filter_by(statut='actif').all()

    stats = {
        'nb_projets': len(projets),
        'nb_bailleurs': Bailleur.query.filter_by(actif=True).count(),
        'budget_total': sum(float(p.budget_total or 0) for p in projets),
        'total_realise': 0,
        'ecritures_mois': 0,
        'projets_data': [],
        'solde_banque': 0,
        'solde_caisse': 0
    }

    # Calculer réalisé total
    for projet in projets:
        projet_realise = 0
        projet_prevu = sum(float(l.montant_prevu or 0) for l in projet.lignes_budget)
        for ligne in projet.lignes_budget:
            realise = db.session.query(
                db.func.sum(LigneEcriture.debit)
            ).join(CompteComptable).filter(
                (LigneEcriture.ligne_budget_id == ligne.id) &
                (CompteComptable.classe == 6)
            ).scalar() or 0
            projet_realise += float(realise)

        stats['total_realise'] += projet_realise
        if projet_prevu > 0:
            stats['projets_data'].append({
                'code': projet.code,
                'nom': projet.nom,
                'prevu': projet_prevu,
                'realise': projet_realise,
                'taux': (projet_realise / projet_prevu) * 100
            })

    # Calculer solde banque (comptes 52x)
    solde_banque = db.session.query(
        db.func.sum(LigneEcriture.debit) - db.func.sum(LigneEcriture.credit)
    ).join(CompteComptable).filter(
        CompteComptable.numero.like('52%')
    ).scalar() or 0
    stats['solde_banque'] = float(solde_banque)

    # Calculer solde caisse (comptes 57x)
    solde_caisse = db.session.query(
        db.func.sum(LigneEcriture.debit) - db.func.sum(LigneEcriture.credit)
    ).join(CompteComptable).filter(
        CompteComptable.numero.like('57%')
    ).scalar() or 0
    stats['solde_caisse'] = float(solde_caisse)

    # Écritures ce mois
    debut_mois = date.today().replace(day=1)
    stats['ecritures_mois'] = PieceComptable.query.filter(
        PieceComptable.date_piece >= debut_mois
    ).count()

    # Taux d'exécution global
    if stats['budget_total'] > 0:
        stats['taux_execution'] = (stats['total_realise'] / stats['budget_total']) * 100
    else:
        stats['taux_execution'] = 0

    return stats


# =============================================================================
# ROUTES - AUTHENTICATION
# =============================================================================

# SECURITY: Simple in-memory rate limiting for login
_login_attempts = {}  # {ip: [(timestamp, email), ...]}
LOGIN_MAX_ATTEMPTS = 5
LOGIN_WINDOW_SECONDS = 60


def check_login_rate_limit(ip_address):
    """Check if IP has exceeded login rate limit"""
    now = datetime.utcnow()
    window_start = now - timedelta(seconds=LOGIN_WINDOW_SECONDS)

    if ip_address in _login_attempts:
        # Clean old entries
        _login_attempts[ip_address] = [
            (ts, email) for ts, email in _login_attempts[ip_address]
            if ts > window_start
        ]
        return len(_login_attempts[ip_address]) >= LOGIN_MAX_ATTEMPTS
    return False


def record_login_attempt(ip_address, email):
    """Record a failed login attempt"""
    now = datetime.utcnow()
    if ip_address not in _login_attempts:
        _login_attempts[ip_address] = []
    _login_attempts[ip_address].append((now, email))


def clear_login_attempts(ip_address):
    """Clear login attempts after successful login"""
    if ip_address in _login_attempts:
        del _login_attempts[ip_address]


@app.route('/login', methods=['GET', 'POST'])
def login():
    """Page de connexion"""
    if current_user.is_authenticated:
        return redirect(url_for('dashboard'))

    ip_address = request.remote_addr

    # SECURITY: Check rate limit
    if check_login_rate_limit(ip_address):
        flash('Trop de tentatives de connexion. Veuillez réessayer dans une minute.', 'danger')
        return render_template('auth/login.html')

    if request.method == 'POST':
        email = request.form.get('email')
        password = request.form.get('password')

        user = Utilisateur.query.filter_by(email=email).first()

        if user and user.actif and check_password_hash(user.password_hash, password):
            clear_login_attempts(ip_address)  # Reset on success
            login_user(user, remember=request.form.get('remember'))
            user.derniere_connexion = datetime.utcnow()
            db.session.commit()

            log_audit('utilisateurs', user.id, 'LOGIN')
            db.session.commit()

            # SECURITY: Validate next parameter to prevent open redirect
            next_page = request.args.get('next')
            if next_page:
                # Only allow relative URLs (no scheme/netloc)
                from urllib.parse import urlparse
                parsed = urlparse(next_page)
                if parsed.netloc or parsed.scheme:
                    next_page = None  # Reject external URLs
            flash(f'Bienvenue, {user.prenom or user.nom}!', 'success')
            return redirect(next_page or url_for('dashboard'))
        else:
            record_login_attempt(ip_address, email)
            flash('Email ou mot de passe incorrect.', 'danger')
            # Log failed login attempt
            log_audit('utilisateurs', None, 'LOGIN_FAILED', new_values={'email': email})
            db.session.commit()

    return render_template('auth/login.html')


@app.route('/logout')
@login_required
def logout():
    """Déconnexion"""
    log_audit('utilisateurs', current_user.id, 'LOGOUT')
    db.session.commit()
    logout_user()
    flash('Vous avez été déconnecté.', 'info')
    return redirect(url_for('login'))


@app.route('/admin/utilisateurs')
@login_required
@role_required(['directeur'])
def liste_utilisateurs():
    """Gestion des utilisateurs (Directeur uniquement)"""
    utilisateurs = Utilisateur.query.all()
    return render_template('admin/utilisateurs.html', utilisateurs=utilisateurs)


@app.route('/admin/utilisateurs/nouveau', methods=['GET', 'POST'])
@login_required
@role_required(['directeur'])
def nouveau_utilisateur():
    """Créer un nouvel utilisateur"""
    if request.method == 'POST':
        email = request.form.get('email')

        if Utilisateur.query.filter_by(email=email).first():
            flash('Cet email est déjà utilisé.', 'danger')
        else:
            utilisateur = Utilisateur(
                email=email,
                nom=request.form.get('nom'),
                prenom=request.form.get('prenom'),
                password_hash=generate_password_hash(request.form.get('password')),
                role=request.form.get('role', 'comptable'),
                created_by=current_user.email
            )
            db.session.add(utilisateur)
            log_audit('utilisateurs', None, 'CREATE', new_values={'email': email, 'role': utilisateur.role})
            db.session.commit()
            flash('Utilisateur créé avec succès.', 'success')
            return redirect(url_for('liste_utilisateurs'))

    return render_template('admin/utilisateur_form.html')


@app.route('/admin/utilisateurs/<int:id>/modifier', methods=['GET', 'POST'])
@login_required
@role_required(['directeur'])
def modifier_utilisateur(id):
    """Modifier un utilisateur"""
    utilisateur = Utilisateur.query.get_or_404(id)

    if request.method == 'POST':
        old_values = {'email': utilisateur.email, 'role': utilisateur.role, 'actif': utilisateur.actif}

        utilisateur.nom = request.form.get('nom')
        utilisateur.prenom = request.form.get('prenom')
        utilisateur.role = request.form.get('role')
        utilisateur.actif = request.form.get('actif') == 'on'

        if request.form.get('password'):
            utilisateur.password_hash = generate_password_hash(request.form.get('password'))

        new_values = {'email': utilisateur.email, 'role': utilisateur.role, 'actif': utilisateur.actif}
        log_audit('utilisateurs', id, 'UPDATE', old_values=old_values, new_values=new_values)
        db.session.commit()

        flash('Utilisateur modifié avec succès.', 'success')
        return redirect(url_for('liste_utilisateurs'))

    return render_template('admin/utilisateur_form.html', utilisateur=utilisateur)


@app.route('/admin/audit')
@login_required
@role_required(['directeur', 'auditeur'])
def audit_trail():
    """Consulter le journal d'audit"""
    page = request.args.get('page', 1, type=int)
    logs = AuditLog.query.order_by(AuditLog.timestamp.desc()).paginate(page=page, per_page=50)
    return render_template('admin/audit_trail.html', logs=logs)


# =============================================================================
# ROUTES - DASHBOARD
# =============================================================================

@app.route('/')
@login_required
def dashboard():
    """Tableau de bord principal"""
    projets = Projet.query.filter_by(statut='actif').all()
    bailleurs = Bailleur.query.filter_by(actif=True).all()

    # Statistiques améliorées
    stats = calculer_stats_dashboard()

    # Alertes
    alertes = generer_alertes()

    return render_template('dashboard.html',
                          projets=projets,
                          bailleurs=bailleurs,
                          stats=stats,
                          alertes=alertes)


@app.route('/api/dashboard/monthly-expenses')
@login_required
def api_monthly_expenses():
    """API: Dépenses mensuelles sur les 12 derniers mois"""
    from calendar import month_abbr
    import locale
    try:
        locale.setlocale(locale.LC_TIME, 'fr_FR.UTF-8')
    except:
        pass

    # Calculer les 12 derniers mois
    today = date.today()
    months = []
    for i in range(11, -1, -1):
        # Calculer le mois (en partant de 11 mois avant jusqu'à maintenant)
        month = today.month - i
        year = today.year
        while month <= 0:
            month += 12
            year -= 1
        months.append((year, month))

    labels = []
    values = []

    for year, month in months:
        # Calculer premier et dernier jour du mois
        first_day = date(year, month, 1)
        if month == 12:
            last_day = date(year + 1, 1, 1) - timedelta(days=1)
        else:
            last_day = date(year, month + 1, 1) - timedelta(days=1)

        # Somme des débits sur comptes classe 6 pour ce mois
        total = db.session.query(
            db.func.sum(LigneEcriture.debit)
        ).join(PieceComptable).join(CompteComptable).filter(
            PieceComptable.date_piece >= first_day,
            PieceComptable.date_piece <= last_day,
            CompteComptable.classe == 6
        ).scalar() or 0

        # Label du mois
        try:
            labels.append(first_day.strftime('%b %Y'))
        except:
            labels.append(f"{month}/{year}")
        values.append(float(total))

    return jsonify({'labels': labels, 'values': values})


@app.route('/api/dashboard/category-distribution')
@login_required
def api_category_distribution():
    """API: Répartition des dépenses par catégorie budgétaire"""
    categories = CategorieBudget.query.order_by(CategorieBudget.ordre).all()

    labels = []
    values = []

    for cat in categories:
        # Somme des débits imputés aux lignes budget de cette catégorie
        total = db.session.query(
            db.func.sum(LigneEcriture.debit)
        ).join(LigneBudget).join(CompteComptable).filter(
            LigneBudget.categorie_id == cat.id,
            CompteComptable.classe == 6
        ).scalar() or 0

        if float(total) > 0:
            labels.append(cat.nom)
            values.append(float(total))

    # Si aucune donnée, retourner des valeurs par défaut
    if not values:
        labels = ['Aucune dépense']
        values = [1]

    return jsonify({'labels': labels, 'values': values})


# =============================================================================
# ROUTES - BAILLEURS
# =============================================================================

@app.route('/bailleurs')
@login_required
def liste_bailleurs():
    """Liste des bailleurs"""
    bailleurs = Bailleur.query.all()
    return render_template('bailleurs/liste.html', bailleurs=bailleurs)


@app.route('/bailleurs/nouveau', methods=['GET', 'POST'])
@login_required
@role_required(['comptable', 'directeur'])
def nouveau_bailleur():
    """Créer un nouveau bailleur"""
    devises = Devise.query.all()

    if request.method == 'POST':
        bailleur = Bailleur(
            code=request.form['code'],
            nom=request.form['nom'],
            pays=request.form.get('pays'),
            contact=request.form.get('contact'),
            email=request.form.get('email'),
            devise_id=request.form.get('devise_id') or None
        )
        db.session.add(bailleur)
        db.session.commit()
        flash('Bailleur créé avec succès', 'success')
        return redirect(url_for('liste_bailleurs'))

    return render_template('bailleurs/form.html', devises=devises)


@app.route('/bailleurs/<int:id>/modifier', methods=['GET', 'POST'])
@login_required
@role_required(['comptable', 'directeur'])
def modifier_bailleur(id):
    """Modifier un bailleur"""
    bailleur = Bailleur.query.get_or_404(id)
    devises = Devise.query.all()

    if request.method == 'POST':
        bailleur.code = request.form['code']
        bailleur.nom = request.form['nom']
        bailleur.pays = request.form.get('pays')
        bailleur.contact = request.form.get('contact')
        bailleur.email = request.form.get('email')
        bailleur.devise_id = request.form.get('devise_id') or None
        db.session.commit()
        flash('Bailleur modifié avec succès', 'success')
        return redirect(url_for('liste_bailleurs'))

    return render_template('bailleurs/form.html', bailleur=bailleur, devises=devises)


# =============================================================================
# ROUTES - FINANCEMENTS (Dons et Subventions)
# =============================================================================

@app.route('/financements')
@login_required
def liste_financements():
    """Liste des financements/dons"""
    financements = Financement.query.order_by(Financement.date_creation.desc()).all()

    # Statistiques
    stats = {
        'total': len(financements),
        'actifs': len([f for f in financements if f.statut == 'actif']),
        'montant_total': sum(float(f.montant or 0) for f in financements),
        'montant_recu': sum(f.montant_recu for f in financements),
        'par_type': {
            'libre': len([f for f in financements if f.type_affectation == 'libre']),
            'projet': len([f for f in financements if f.type_affectation == 'projet']),
            'usage': len([f for f in financements if f.type_affectation == 'usage']),
        }
    }

    return render_template('financements/liste.html', financements=financements, stats=stats)


@app.route('/financements/nouveau', methods=['GET', 'POST'])
@login_required
@role_required(['comptable', 'directeur'])
def nouveau_financement():
    """Créer un nouveau financement"""
    if request.method == 'POST':
        financement = Financement(
            reference=request.form.get('reference'),
            bailleur_id=int(request.form['bailleur_id']),
            type_affectation=request.form.get('type_affectation', 'libre'),
            montant=Decimal(request.form['montant'].replace(',', '.').replace(' ', '')),
            devise_id=int(request.form['devise_id']) if request.form.get('devise_id') else None,
            date_accord=datetime.strptime(request.form['date_accord'], '%Y-%m-%d').date() if request.form.get('date_accord') else None,
            date_fin=datetime.strptime(request.form['date_fin'], '%Y-%m-%d').date() if request.form.get('date_fin') else None,
            notes=request.form.get('notes'),
            statut='actif'
        )

        # Affectation selon le type
        if financement.type_affectation == 'projet' and request.form.get('projet_id'):
            financement.projet_id = int(request.form['projet_id'])
        elif financement.type_affectation == 'usage':
            financement.affectation_libelle = request.form.get('affectation_libelle')

        db.session.add(financement)
        db.session.flush()  # Pour obtenir l'ID

        # Créer les tranches si spécifiées
        nb_tranches = int(request.form.get('nb_tranches', 1))
        for i in range(1, nb_tranches + 1):
            montant_tranche = request.form.get(f'tranche_{i}_montant')
            date_tranche = request.form.get(f'tranche_{i}_date')

            if montant_tranche:
                tranche = TrancheFinancement(
                    financement_id=financement.id,
                    numero=i,
                    montant_prevu=Decimal(montant_tranche.replace(',', '.').replace(' ', '')),
                    date_prevue=datetime.strptime(date_tranche, '%Y-%m-%d').date() if date_tranche else None,
                    statut='attendu'
                )
                db.session.add(tranche)

        db.session.commit()
        log_audit('financement', financement.id, 'CREATE', new_values={
            'reference': financement.reference,
            'bailleur': financement.bailleur.nom,
            'montant': str(financement.montant)
        })

        flash(f'Financement "{financement.reference}" créé avec succès', 'success')
        return redirect(url_for('detail_financement', id=financement.id))

    bailleurs = Bailleur.query.order_by(Bailleur.nom).all()
    projets = Projet.query.filter_by(statut='actif').order_by(Projet.code).all()
    devises = Devise.query.all()

    return render_template('financements/form.html',
                           financement=None,
                           bailleurs=bailleurs,
                           projets=projets,
                           devises=devises)


@app.route('/financements/<int:id>')
@login_required
def detail_financement(id):
    """Détail d'un financement"""
    financement = Financement.query.get_or_404(id)
    return render_template('financements/detail.html', financement=financement, today=date.today)


@app.route('/financements/<int:id>/modifier', methods=['GET', 'POST'])
@login_required
@role_required(['comptable', 'directeur'])
def modifier_financement(id):
    """Modifier un financement"""
    financement = Financement.query.get_or_404(id)

    if request.method == 'POST':
        financement.reference = request.form.get('reference')
        financement.bailleur_id = int(request.form['bailleur_id'])
        financement.type_affectation = request.form.get('type_affectation', 'libre')
        financement.montant = Decimal(request.form['montant'].replace(',', '.').replace(' ', ''))
        financement.devise_id = int(request.form['devise_id']) if request.form.get('devise_id') else None
        financement.date_accord = datetime.strptime(request.form['date_accord'], '%Y-%m-%d').date() if request.form.get('date_accord') else None
        financement.date_fin = datetime.strptime(request.form['date_fin'], '%Y-%m-%d').date() if request.form.get('date_fin') else None
        financement.notes = request.form.get('notes')
        financement.statut = request.form.get('statut', 'actif')

        # Affectation selon le type
        if financement.type_affectation == 'projet':
            financement.projet_id = int(request.form['projet_id']) if request.form.get('projet_id') else None
            financement.affectation_libelle = None
        elif financement.type_affectation == 'usage':
            financement.projet_id = None
            financement.affectation_libelle = request.form.get('affectation_libelle')
        else:
            financement.projet_id = None
            financement.affectation_libelle = None

        db.session.commit()
        log_audit('financement', financement.id, 'UPDATE')

        flash('Financement modifié avec succès', 'success')
        return redirect(url_for('detail_financement', id=financement.id))

    bailleurs = Bailleur.query.order_by(Bailleur.nom).all()
    projets = Projet.query.order_by(Projet.code).all()
    devises = Devise.query.all()

    return render_template('financements/form.html',
                           financement=financement,
                           bailleurs=bailleurs,
                           projets=projets,
                           devises=devises)


@app.route('/financements/<int:id>/supprimer', methods=['POST'])
@login_required
@role_required(['directeur'])
def supprimer_financement(id):
    """Supprimer un financement"""
    financement = Financement.query.get_or_404(id)

    # Vérifier qu'aucune tranche n'a été reçue (même partiellement)
    if any(t.statut in ('recu', 'partiel') or (t.montant_recu and t.montant_recu > 0) for t in financement.tranches):
        flash('Impossible de supprimer: des tranches ont déjà été reçues (totalement ou partiellement)', 'danger')
        return redirect(url_for('detail_financement', id=id))

    log_audit('financement', financement.id, 'DELETE', old_values={
        'reference': financement.reference,
        'bailleur': financement.bailleur.nom
    })

    db.session.delete(financement)
    db.session.commit()

    flash('Financement supprimé', 'success')
    return redirect(url_for('liste_financements'))


@app.route('/financements/<int:id>/tranches/ajouter', methods=['POST'])
@login_required
@role_required(['comptable', 'directeur'])
def ajouter_tranche(id):
    """Ajouter une tranche à un financement"""
    financement = Financement.query.get_or_404(id)

    dernier_numero = max([t.numero for t in financement.tranches], default=0)

    tranche = TrancheFinancement(
        financement_id=financement.id,
        numero=dernier_numero + 1,
        montant_prevu=Decimal(request.form['montant_prevu'].replace(',', '.').replace(' ', '')),
        date_prevue=datetime.strptime(request.form['date_prevue'], '%Y-%m-%d').date() if request.form.get('date_prevue') else None,
        statut='attendu'
    )

    db.session.add(tranche)
    db.session.commit()

    flash(f'Tranche {tranche.numero} ajoutée', 'success')
    return redirect(url_for('detail_financement', id=id))


@app.route('/financements/tranches/<int:id>/recevoir', methods=['POST'])
@login_required
@role_required(['comptable', 'directeur'])
def recevoir_tranche(id):
    """Marquer une tranche comme reçue"""
    tranche = TrancheFinancement.query.get_or_404(id)

    montant_recu = request.form.get('montant_recu')
    if montant_recu:
        tranche.montant_recu = Decimal(montant_recu.replace(',', '.').replace(' ', ''))
    else:
        tranche.montant_recu = tranche.montant_prevu

    tranche.date_reception = datetime.strptime(request.form['date_reception'], '%Y-%m-%d').date() if request.form.get('date_reception') else date.today()

    # Statut: recu si montant complet, partiel sinon
    if tranche.montant_recu >= tranche.montant_prevu:
        tranche.statut = 'recu'
    else:
        tranche.statut = 'partiel'

    db.session.commit()

    log_audit('tranche_financement', tranche.id, 'RECEPTION', new_values={
        'financement': tranche.financement.reference,
        'montant_recu': str(tranche.montant_recu)
    })

    flash(f'Tranche {tranche.numero} marquée comme reçue ({tranche.montant_recu})', 'success')
    return redirect(url_for('detail_financement', id=tranche.financement_id))


@app.route('/financements/tranches/<int:id>/supprimer', methods=['POST'])
@login_required
@role_required(['directeur'])
def supprimer_tranche(id):
    """Supprimer une tranche"""
    tranche = TrancheFinancement.query.get_or_404(id)
    financement_id = tranche.financement_id

    if tranche.statut in ('recu', 'partiel') or (tranche.montant_recu and tranche.montant_recu > 0):
        flash('Impossible de supprimer une tranche déjà reçue (totalement ou partiellement)', 'danger')
        return redirect(url_for('detail_financement', id=financement_id))

    db.session.delete(tranche)
    db.session.commit()

    flash('Tranche supprimée', 'success')
    return redirect(url_for('detail_financement', id=financement_id))


@app.route('/revenus/tableau-de-bord')
@login_required
def tableau_bord_revenus():
    """Tableau de bord des revenus"""
    # Financements actifs
    financements = Financement.query.filter_by(statut='actif').all()

    # Statistiques globales
    stats = {
        'nb_financements': len(financements),
        'montant_total': sum(float(f.montant or 0) for f in financements),
        'montant_recu': sum(f.montant_recu for f in financements),
        'montant_attendu': sum(f.montant_attendu for f in financements),
    }
    stats['pourcentage_recu'] = round((stats['montant_recu'] / stats['montant_total'] * 100), 1) if stats['montant_total'] > 0 else 0

    # Par type d'affectation
    par_type = {}
    for type_aff in ['libre', 'projet', 'usage']:
        fins = [f for f in financements if f.type_affectation == type_aff]
        par_type[type_aff] = {
            'count': len(fins),
            'montant': sum(float(f.montant or 0) for f in fins),
            'recu': sum(f.montant_recu for f in fins)
        }

    # Tranches en retard
    tranches_retard = TrancheFinancement.query.join(Financement).filter(
        Financement.statut == 'actif',
        TrancheFinancement.statut.in_(['attendu', 'retard']),
        TrancheFinancement.date_prevue < date.today()
    ).all()

    # Prochaines tranches attendues (30 jours)
    date_limite = date.today() + timedelta(days=30)
    prochaines_tranches = TrancheFinancement.query.join(Financement).filter(
        Financement.statut == 'actif',
        TrancheFinancement.statut == 'attendu',
        TrancheFinancement.date_prevue >= date.today(),
        TrancheFinancement.date_prevue <= date_limite
    ).order_by(TrancheFinancement.date_prevue).all()

    # Top bailleurs
    bailleurs_stats = {}
    for f in financements:
        bailleur_nom = f.bailleur.nom if f.bailleur else 'Inconnu'
        if bailleur_nom not in bailleurs_stats:
            bailleurs_stats[bailleur_nom] = {'montant': 0, 'recu': 0}
        bailleurs_stats[bailleur_nom]['montant'] += float(f.montant or 0)
        bailleurs_stats[bailleur_nom]['recu'] += f.montant_recu

    top_bailleurs = sorted(bailleurs_stats.items(), key=lambda x: x[1]['montant'], reverse=True)[:5]

    return render_template('financements/tableau_bord.html',
                           stats=stats,
                           par_type=par_type,
                           tranches_retard=tranches_retard,
                           prochaines_tranches=prochaines_tranches,
                           top_bailleurs=top_bailleurs,
                           financements=financements,
                           today=date.today)


# =============================================================================
# ROUTES - PROJETS
# =============================================================================

@app.route('/projets')
@login_required
def liste_projets():
    """Liste des projets"""
    projets = Projet.query.all()
    return render_template('projets/liste.html', projets=projets)


@app.route('/projets/nouveau', methods=['GET', 'POST'])
@login_required
@role_required(['comptable', 'directeur'])
def nouveau_projet():
    """Créer un nouveau projet"""
    bailleurs = Bailleur.query.filter_by(actif=True).all()
    devises = Devise.query.all()

    if request.method == 'POST':
        projet = Projet(
            code=request.form['code'],
            nom=request.form['nom'],
            description=request.form.get('description'),
            bailleur_id=request.form.get('bailleur_id') or None,
            date_debut=datetime.strptime(request.form['date_debut'], '%Y-%m-%d').date() if request.form.get('date_debut') else None,
            date_fin=datetime.strptime(request.form['date_fin'], '%Y-%m-%d').date() if request.form.get('date_fin') else None,
            budget_total=request.form.get('budget_total') or 0,
            devise_id=request.form.get('devise_id') or None
        )
        db.session.add(projet)
        db.session.commit()
        flash('Projet créé avec succès', 'success')
        return redirect(url_for('liste_projets'))

    return render_template('projets/form.html', bailleurs=bailleurs, devises=devises)


@app.route('/projets/<int:id>')
@login_required
def detail_projet(id):
    """Détail d'un projet avec budget"""
    projet = Projet.query.get_or_404(id)
    categories = CategorieBudget.query.order_by(CategorieBudget.ordre).all()

    # Déterminer les années du projet
    annee_debut = projet.date_debut.year if projet.date_debut else datetime.now().year
    annee_fin = projet.date_fin.year if projet.date_fin else annee_debut + 2
    annees_disponibles = list(range(annee_debut, annee_fin + 1))

    # Filtrer par année si spécifié
    annee_filtre = request.args.get('annee', type=int)

    # Calcul des réalisés par ligne budgétaire
    # SYSCOHADA: Pour les charges (classe 6), seuls les débits comptent comme dépenses réalisées
    realisations = {}
    budgets_par_annee = {}

    for ligne in projet.lignes_budget:
        # Requête de base pour le réalisé
        query = db.session.query(
            db.func.sum(LigneEcriture.debit)
        ).join(CompteComptable).join(PieceComptable).filter(
            (LigneEcriture.ligne_budget_id == ligne.id) &
            (CompteComptable.classe == 6)
        )

        # Filtrer par année si nécessaire
        if annee_filtre:
            query = query.filter(
                db.extract('year', PieceComptable.date_piece) == annee_filtre
            )

        realise = query.scalar() or 0
        realisations[ligne.id] = float(realise)

        # Budget par année pour cette ligne
        budgets_par_annee[ligne.id] = {
            ba.annee: float(ba.montant_prevu) for ba in ligne.budgets_annuels
        }

    # Calculer les totaux par année
    totaux_annuels = {}
    for annee in annees_disponibles:
        total_prevu = sum(
            budgets_par_annee.get(ligne.id, {}).get(annee, 0)
            for ligne in projet.lignes_budget
        )
        totaux_annuels[annee] = total_prevu

    return render_template('projets/detail.html',
                         projet=projet,
                         categories=categories,
                         realisations=realisations,
                         annees_disponibles=annees_disponibles,
                         annee_filtre=annee_filtre,
                         budgets_par_annee=budgets_par_annee,
                         totaux_annuels=totaux_annuels)


@app.route('/projets/<int:id>/budget/ajouter', methods=['GET', 'POST'])
@login_required
@role_required(['comptable', 'directeur'])
def ajouter_ligne_budget(id):
    """Ajouter une ligne budgétaire"""
    projet = Projet.query.get_or_404(id)
    categories = CategorieBudget.query.order_by(CategorieBudget.ordre).all()

    if request.method == 'POST':
        montant = float(request.form.get('quantite', 1)) * float(request.form.get('cout_unitaire', 0))
        ligne = LigneBudget(
            projet_id=projet.id,
            categorie_id=request.form.get('categorie_id') or None,
            code=request.form['code'],
            intitule=request.form['intitule'],
            annee=request.form.get('annee') or None,
            quantite=request.form.get('quantite') or 1,
            unite=request.form.get('unite'),
            cout_unitaire=request.form.get('cout_unitaire') or 0,
            montant_prevu=montant
        )
        db.session.add(ligne)
        db.session.commit()
        flash('Ligne budgétaire ajoutée', 'success')
        return redirect(url_for('detail_projet', id=id))

    return render_template('projets/ligne_budget_form.html', projet=projet, categories=categories)


@app.route('/projets/<int:projet_id>/budget/<int:ligne_id>/annees', methods=['GET', 'POST'])
@login_required
@role_required(['comptable', 'directeur'])
def gerer_budget_annuel(projet_id, ligne_id):
    """Gérer la répartition annuelle d'une ligne budgétaire"""
    projet = Projet.query.get_or_404(projet_id)
    ligne = LigneBudget.query.get_or_404(ligne_id)

    if ligne.projet_id != projet.id:
        flash('Ligne budgétaire invalide', 'danger')
        return redirect(url_for('detail_projet', id=projet_id))

    # Déterminer les années du projet
    annee_debut = projet.date_debut.year if projet.date_debut else datetime.now().year
    annee_fin = projet.date_fin.year if projet.date_fin else annee_debut + 2
    annees = list(range(annee_debut, annee_fin + 1))

    if request.method == 'POST':
        # Supprimer les anciennes entrées
        BudgetAnnee.query.filter_by(ligne_budget_id=ligne.id).delete()

        # Ajouter les nouvelles entrées
        total = Decimal('0')
        for annee in annees:
            montant = request.form.get(f'montant_{annee}', '0')
            try:
                montant = Decimal(montant) if montant else Decimal('0')
            except:
                montant = Decimal('0')

            if montant > 0:
                ba = BudgetAnnee(
                    ligne_budget_id=ligne.id,
                    annee=annee,
                    montant_prevu=montant
                )
                db.session.add(ba)
                total += montant

        # Mettre à jour le total de la ligne
        ligne.montant_prevu = total
        db.session.commit()

        flash('Répartition annuelle enregistrée', 'success')
        return redirect(url_for('detail_projet', id=projet_id))

    # Charger les budgets existants
    budgets_existants = {ba.annee: ba.montant_prevu for ba in ligne.budgets_annuels}

    return render_template('projets/budget_annuel.html',
                         projet=projet,
                         ligne=ligne,
                         annees=annees,
                         budgets=budgets_existants)


@app.route('/api/projets/<int:projet_id>/budget-annuel')
@login_required
def api_budget_annuel(projet_id):
    """API pour obtenir le budget par année d'un projet"""
    projet = Projet.query.get_or_404(projet_id)
    annee = request.args.get('annee', type=int)

    # Déterminer les années disponibles
    annee_debut = projet.date_debut.year if projet.date_debut else datetime.now().year
    annee_fin = projet.date_fin.year if projet.date_fin else annee_debut + 2
    annees_disponibles = list(range(annee_debut, annee_fin + 1))

    result = {
        'projet_id': projet.id,
        'annees_disponibles': annees_disponibles,
        'lignes': []
    }

    for ligne in projet.lignes_budget:
        ligne_data = {
            'id': ligne.id,
            'code': ligne.code,
            'intitule': ligne.intitule,
            'categorie': ligne.categorie.nom if ligne.categorie else None,
            'total_prevu': float(ligne.get_total_prevu()),
            'budgets_annuels': {}
        }

        for ba in ligne.budgets_annuels:
            ligne_data['budgets_annuels'][ba.annee] = float(ba.montant_prevu)

        # Si filtre par année
        if annee:
            ligne_data['montant_annee'] = float(ligne.get_montant_annee(annee))

        result['lignes'].append(ligne_data)

    return jsonify(result)


# =============================================================================
# ROUTES - COMPTABILITE
# =============================================================================

@app.route('/comptabilite/comptes')
@login_required
def plan_comptable():
    """Plan comptable"""
    comptes = CompteComptable.query.order_by(CompteComptable.numero).all()
    return render_template('comptabilite/plan_comptable.html', comptes=comptes)


@app.route('/comptabilite/ecritures')
@login_required
def liste_ecritures():
    """Liste des écritures comptables avec filtres et pagination"""
    # Base query
    query = PieceComptable.query

    # Filtre par exercice
    exercice_id = request.args.get('exercice_id', type=int)
    if exercice_id:
        query = query.filter(PieceComptable.exercice_id == exercice_id)

    # Filtre par journal
    journal_id = request.args.get('journal_id', type=int)
    if journal_id:
        query = query.filter(PieceComptable.journal_id == journal_id)

    # Filtre par statut de validation
    valide = request.args.get('valide')
    if valide == '1':
        query = query.filter(PieceComptable.valide == True)
    elif valide == '0':
        query = query.filter(PieceComptable.valide == False)

    # Recherche textuelle
    q = request.args.get('q', '').strip()
    if q:
        search = f"%{q}%"
        query = query.filter(
            db.or_(
                PieceComptable.numero.ilike(search),
                PieceComptable.libelle.ilike(search),
                PieceComptable.reference.ilike(search)
            )
        )

    # Pagination (50 par page)
    page = request.args.get('page', 1, type=int)
    per_page = 50
    query = query.order_by(PieceComptable.date_piece.desc())
    pagination = query.paginate(page=page, per_page=per_page, error_out=False)
    pieces = pagination.items
    total = pagination.total

    # Données pour les filtres
    exercices = ExerciceComptable.query.order_by(ExerciceComptable.annee.desc()).all()
    journaux = Journal.query.order_by(Journal.code).all()

    return render_template('comptabilite/ecritures.html',
                           pieces=pieces,
                           pagination=pagination,
                           total=total,
                           exercices=exercices,
                           journaux=journaux)


@app.route('/comptabilite/ecritures/nouvelle', methods=['GET', 'POST'])
@login_required
@role_required(['comptable', 'directeur'])
def nouvelle_ecriture():
    """Saisir une nouvelle écriture - Mode simplifié ou expert"""
    journaux = Journal.query.all()
    exercices = ExerciceComptable.query.filter_by(cloture=False).all()
    comptes = CompteComptable.query.filter_by(actif=True).order_by(CompteComptable.numero).all()
    projets = Projet.query.filter_by(statut='actif').all()
    devises = Devise.query.all()

    if request.method == 'POST':
        operation_type = request.form.get('operation_type', 'expert')

        # Générer numéro de pièce
        dernier = PieceComptable.query.order_by(PieceComptable.id.desc()).first()
        numero = f"PC{datetime.now().year}{(dernier.id + 1 if dernier else 1):05d}"

        # Trouver l'exercice actif
        exercice = ExerciceComptable.query.filter_by(cloture=False).first()
        if not exercice:
            flash("Aucun exercice comptable ouvert.", "danger")
            return redirect(url_for('nouvelle_ecriture'))

        # MODE SIMPLIFIÉ - Le système crée automatiquement l'écriture équilibrée
        if operation_type in ['depense', 'recette', 'virement', 'avance']:
            # SECURITY: Use Decimal for monetary values to avoid floating point errors
            try:
                montant = Decimal(request.form.get('montant', '0'))
            except:
                flash('Montant invalide.', 'danger')
                return redirect(url_for('nouvelle_ecriture'))

            try:
                date_piece = datetime.strptime(request.form['date_piece'], '%Y-%m-%d').date()
            except ValueError:
                flash('Format de date invalide.', 'danger')
                return redirect(url_for('nouvelle_ecriture'))

            # SECURITY: Validate date is within exercise period
            if date_piece < exercice.date_debut or date_piece > exercice.date_fin:
                flash(f'La date doit être comprise entre {exercice.date_debut.strftime("%d/%m/%Y")} et {exercice.date_fin.strftime("%d/%m/%Y")}.', 'danger')
                return redirect(url_for('nouvelle_ecriture'))

            libelle = request.form.get('libelle', '')
            reference = request.form.get('reference', '')
            projet_id = request.form.get('projet_id') or None

            # Déterminer le journal approprié
            journal_codes = {'depense': 'AC', 'recette': 'BQ', 'virement': 'OD', 'avance': 'CA'}
            journal = Journal.query.filter(Journal.code.like(f"{journal_codes.get(operation_type, 'OD')}%")).first()
            if not journal:
                journal = journaux[0] if journaux else None

            piece = PieceComptable(
                numero=numero,
                date_piece=date_piece,
                journal_id=journal.id if journal else 1,
                exercice_id=exercice.id,
                libelle=libelle,
                reference=reference
            )
            db.session.add(piece)
            db.session.flush()

            if operation_type == 'depense':
                # Dépense: Débit charge (6xx), Crédit trésorerie (5xx)
                compte_charge_id = request.form.get('compte_charge')
                compte_tresorerie_id = request.form.get('compte_tresorerie')
                ligne_budget_id = request.form.get('ligne_budget_id') or None

                # Ligne débit (charge) - avec ligne budgétaire pour suivi budget
                ligne_debit = LigneEcriture(
                    piece_id=piece.id,
                    compte_id=compte_charge_id,
                    projet_id=projet_id,
                    libelle=libelle,
                    debit=montant,
                    credit=0,
                    ligne_budget_id=ligne_budget_id
                )
                db.session.add(ligne_debit)
                db.session.flush()

                # Traiter ventilation multi-projets
                ventilation_data = request.form.get('ventilation')
                if ventilation_data:
                    try:
                        ventilations = json.loads(ventilation_data)
                        if ventilations:
                            # SECURITY: Validate ventilation totals 100%
                            total_pct = sum(Decimal(str(v.get('pourcentage', 0))) for v in ventilations)
                            if abs(total_pct - 100) > Decimal('0.01'):
                                flash(f'La ventilation doit totaliser 100% (actuellement {total_pct}%).', 'warning')

                            for v in ventilations:
                                if 'projet_id' not in v or 'pourcentage' not in v:
                                    continue
                                pct = Decimal(str(v['pourcentage']))
                                imputation = ImputationAnalytique(
                                    ligne_ecriture_id=ligne_debit.id,
                                    projet_id=int(v['projet_id']),
                                    pourcentage=pct,
                                    montant=montant * pct / 100
                                )
                                db.session.add(imputation)
                    except json.JSONDecodeError as e:
                        flash(f'Erreur dans les données de ventilation: {str(e)}', 'warning')
                    except (KeyError, ValueError, TypeError) as e:
                        flash(f'Données de ventilation invalides: {str(e)}', 'warning')

                # Ligne crédit (trésorerie)
                ligne_credit = LigneEcriture(
                    piece_id=piece.id,
                    compte_id=compte_tresorerie_id,
                    projet_id=projet_id,
                    libelle=libelle,
                    debit=0,
                    credit=montant
                )
                db.session.add(ligne_credit)

            elif operation_type == 'recette':
                # Recette: Débit trésorerie (5xx), Crédit produit (7xx)
                compte_produit_id = request.form.get('compte_produit')
                compte_tresorerie_id = request.form.get('compte_tresorerie')

                # Ligne débit (trésorerie)
                ligne_debit = LigneEcriture(
                    piece_id=piece.id,
                    compte_id=compte_tresorerie_id,
                    projet_id=projet_id,
                    libelle=libelle,
                    debit=montant,
                    credit=0
                )
                db.session.add(ligne_debit)

                # Ligne crédit (produit)
                ligne_credit = LigneEcriture(
                    piece_id=piece.id,
                    compte_id=compte_produit_id,
                    projet_id=projet_id,
                    libelle=libelle,
                    debit=0,
                    credit=montant
                )
                db.session.add(ligne_credit)

            elif operation_type == 'virement':
                # Virement interne: Débit destination, Crédit source
                compte_source_id = request.form.get('compte_source')
                compte_destination_id = request.form.get('compte_destination')

                # Ligne débit (destination)
                ligne_debit = LigneEcriture(
                    piece_id=piece.id,
                    compte_id=compte_destination_id,
                    libelle=libelle,
                    debit=montant,
                    credit=0
                )
                db.session.add(ligne_debit)

                # Ligne crédit (source)
                ligne_credit = LigneEcriture(
                    piece_id=piece.id,
                    compte_id=compte_source_id,
                    libelle=libelle,
                    debit=0,
                    credit=montant
                )
                db.session.add(ligne_credit)

            elif operation_type == 'avance':
                # Avance: Débit compte personnel (421), Crédit trésorerie (5xx)
                compte_tresorerie_id = request.form.get('compte_tresorerie')
                beneficiaire = request.form.get('beneficiaire', '')

                # Trouver ou créer un compte d'avances personnel (421)
                compte_avance = CompteComptable.query.filter(CompteComptable.numero.like('421%')).first()
                if not compte_avance:
                    compte_avance = CompteComptable.query.filter(CompteComptable.numero.like('42%')).first()
                if not compte_avance:
                    flash("Compte d'avances personnel (421) non trouvé dans le plan comptable.", "danger")
                    db.session.rollback()
                    return redirect(url_for('nouvelle_ecriture'))

                libelle_avance = f"Avance {beneficiaire} - {libelle}"

                # Ligne débit (avance personnel)
                ligne_debit = LigneEcriture(
                    piece_id=piece.id,
                    compte_id=compte_avance.id,
                    projet_id=projet_id,
                    libelle=libelle_avance,
                    debit=montant,
                    credit=0
                )
                db.session.add(ligne_debit)

                # Ligne crédit (trésorerie)
                ligne_credit = LigneEcriture(
                    piece_id=piece.id,
                    compte_id=compte_tresorerie_id,
                    projet_id=projet_id,
                    libelle=libelle_avance,
                    debit=0,
                    credit=montant
                )
                db.session.add(ligne_credit)

            elif operation_type == 'salaires':
                # Salaires: Plusieurs lignes d'écriture
                # Débit 661 (salaires bruts), Débit 664 (charges patronales)
                # Crédit 421 (personnel), Crédit 43x (organismes sociaux), Crédit 5xx (trésorerie)
                compte_tresorerie_id = request.form.get('compte_tresorerie')
                # SECURITY: Use Decimal for monetary values
                try:
                    salaires_bruts = Decimal(request.form.get('salaires_bruts', '0') or '0')
                    charges_patronales = Decimal(request.form.get('charges_patronales', '0') or '0')
                    retenues_salariales = Decimal(request.form.get('retenues_salariales', '0') or '0')
                except:
                    flash('Montants de salaires invalides.', 'danger')
                    return redirect(url_for('nouvelle_ecriture'))
                mois_salaire = request.form.get('mois_salaire', '')
                annee_salaire = request.form.get('annee_salaire', '')

                net_a_payer = salaires_bruts - retenues_salariales

                # Trouver les comptes nécessaires
                compte_661 = CompteComptable.query.filter(CompteComptable.numero.like('661%')).first()
                compte_664 = CompteComptable.query.filter(CompteComptable.numero.like('664%')).first()
                compte_421 = CompteComptable.query.filter(CompteComptable.numero.like('421%')).first()
                compte_43 = CompteComptable.query.filter(CompteComptable.numero.like('43%')).first()

                if not compte_661:
                    flash("Compte 661 (Rémunérations) non trouvé.", "danger")
                    db.session.rollback()
                    return redirect(url_for('nouvelle_ecriture'))

                libelle_salaire = f"Salaires {mois_salaire}/{annee_salaire}"
                piece.libelle = libelle_salaire

                # 1. Débit 661 - Salaires bruts
                ligne_salaires = LigneEcriture(
                    piece_id=piece.id,
                    compte_id=compte_661.id,
                    projet_id=projet_id,
                    libelle=libelle_salaire,
                    debit=salaires_bruts,
                    credit=0
                )
                db.session.add(ligne_salaires)
                db.session.flush()

                # Traiter ventilation multi-projets sur les salaires
                ventilation_data = request.form.get('ventilation')
                if ventilation_data:
                    try:
                        ventilations = json.loads(ventilation_data)
                        if ventilations:
                            for v in ventilations:
                                imputation = ImputationAnalytique(
                                    ligne_ecriture_id=ligne_salaires.id,
                                    projet_id=int(v['projet_id']),
                                    pourcentage=Decimal(str(v['pourcentage'])),
                                    montant=Decimal(str(salaires_bruts * v['pourcentage'] / 100))
                                )
                                db.session.add(imputation)
                    except (json.JSONDecodeError, KeyError):
                        pass

                # 2. Débit 664 - Charges patronales (si > 0)
                if charges_patronales > 0 and compte_664:
                    ligne_charges = LigneEcriture(
                        piece_id=piece.id,
                        compte_id=compte_664.id,
                        projet_id=projet_id,
                        libelle=f"Charges sociales {mois_salaire}/{annee_salaire}",
                        debit=charges_patronales,
                        credit=0
                    )
                    db.session.add(ligne_charges)

                # 3. Crédit 43x - Organismes sociaux (retenues + charges patronales)
                total_organismes = retenues_salariales + charges_patronales
                if total_organismes > 0 and compte_43:
                    ligne_organismes = LigneEcriture(
                        piece_id=piece.id,
                        compte_id=compte_43.id,
                        projet_id=projet_id,
                        libelle=f"Cotisations sociales {mois_salaire}/{annee_salaire}",
                        debit=0,
                        credit=total_organismes
                    )
                    db.session.add(ligne_organismes)

                # 4. Crédit 5xx - Trésorerie (net à payer)
                if net_a_payer > 0:
                    ligne_paiement = LigneEcriture(
                        piece_id=piece.id,
                        compte_id=compte_tresorerie_id,
                        projet_id=projet_id,
                        libelle=f"Paiement salaires {mois_salaire}/{annee_salaire}",
                        debit=0,
                        credit=net_a_payer
                    )
                    db.session.add(ligne_paiement)

            db.session.commit()
            flash(f'Opération enregistrée avec succès (Écriture {numero})', 'success')
            return redirect(url_for('liste_ecritures'))

        # MODE EXPERT - Écriture manuelle classique
        else:
            piece = PieceComptable(
                numero=numero,
                date_piece=datetime.strptime(request.form['date_piece'], '%Y-%m-%d').date(),
                journal_id=request.form['journal_id'],
                exercice_id=request.form['exercice_id'],
                libelle=request.form['libelle'],
                reference=request.form.get('reference'),
                devise_id=request.form.get('devise_id') or None,
                taux_change=request.form.get('taux_change') or 1
            )
            db.session.add(piece)
            db.session.flush()

            # Ajouter les lignes
            comptes_ids = request.form.getlist('compte_id[]')
            projets_ids = request.form.getlist('projet_id[]')
            debits = request.form.getlist('debit[]')
            credits = request.form.getlist('credit[]')

            for i in range(len(comptes_ids)):
                if comptes_ids[i]:
                    ligne = LigneEcriture(
                        piece_id=piece.id,
                        compte_id=comptes_ids[i],
                        projet_id=projets_ids[i] if i < len(projets_ids) and projets_ids[i] else None,
                        libelle=request.form.get('libelle', ''),
                        debit=float(debits[i]) if debits[i] else 0,
                        credit=float(credits[i]) if credits[i] else 0
                    )
                    db.session.add(ligne)

            # VALIDATION SYSCOHADA : Vérifier équilibre Débit = Crédit
            db.session.flush()
            if not piece.est_equilibree:
                db.session.rollback()
                flash("Écriture déséquilibrée - Total Débit ≠ Total Crédit. L'écriture n'a pas été enregistrée.", "danger")
                return redirect(url_for('nouvelle_ecriture'))

            db.session.commit()
            flash(f'Écriture {numero} créée avec succès', 'success')
            return redirect(url_for('liste_ecritures'))

    # Charger les lignes budgétaires pour le suivi budget
    lignes_budget = LigneBudget.query.join(Projet).filter(Projet.statut == 'actif').all()

    return render_template('comptabilite/ecriture_form.html',
                         journaux=journaux,
                         exercices=exercices,
                         comptes=comptes,
                         projets=projets,
                         devises=devises,
                         lignes_budget=lignes_budget,
                         today=date.today().strftime('%Y-%m-%d'))


@app.route('/comptabilite/ecritures/<int:id>')
@login_required
def detail_ecriture(id):
    """Détail d'une écriture comptable"""
    piece = PieceComptable.query.get_or_404(id)
    return render_template('comptabilite/ecriture_detail.html', piece=piece)


@app.route('/comptabilite/ecritures/<int:id>/modifier', methods=['GET', 'POST'])
@login_required
@role_required(['comptable', 'directeur'])
def modifier_ecriture(id):
    """Modifier une écriture comptable non validée"""
    piece = PieceComptable.query.get_or_404(id)

    if piece.valide:
        flash('Impossible de modifier une écriture validée.', 'danger')
        return redirect(url_for('detail_ecriture', id=id))

    # Check if exercise is closed
    if piece.exercice.cloture:
        flash('Impossible de modifier une écriture sur un exercice clôturé.', 'danger')
        return redirect(url_for('detail_ecriture', id=id))

    journaux = Journal.query.all()
    exercices = ExerciceComptable.query.filter_by(cloture=False).all()
    comptes = CompteComptable.query.filter_by(actif=True).order_by(CompteComptable.numero).all()
    projets = Projet.query.filter_by(statut='actif').all()
    devises = Devise.query.all()
    lignes_budget = LigneBudget.query.all()

    if request.method == 'POST':
        # Store old values for audit
        old_values = {
            'date_piece': str(piece.date_piece),
            'libelle': piece.libelle,
            'reference': piece.reference
        }

        # Parse and validate date
        try:
            new_date = datetime.strptime(request.form['date_piece'], '%Y-%m-%d').date()
        except ValueError:
            flash('Format de date invalide.', 'danger')
            return redirect(url_for('modifier_ecriture', id=id))

        # SECURITY: Validate target exercise is not closed
        new_exercice_id = request.form['exercice_id']
        new_exercice = ExerciceComptable.query.get(new_exercice_id)
        if not new_exercice:
            flash('Exercice invalide.', 'danger')
            return redirect(url_for('modifier_ecriture', id=id))

        if new_exercice.cloture:
            flash('Impossible de déplacer une écriture vers un exercice clôturé.', 'danger')
            return redirect(url_for('modifier_ecriture', id=id))

        # SECURITY: Validate date is within exercise period
        if new_date < new_exercice.date_debut or new_date > new_exercice.date_fin:
            flash(f'La date doit être comprise entre {new_exercice.date_debut.strftime("%d/%m/%Y")} et {new_exercice.date_fin.strftime("%d/%m/%Y")}.', 'danger')
            return redirect(url_for('modifier_ecriture', id=id))

        # Update piece
        piece.date_piece = new_date
        piece.journal_id = request.form['journal_id']
        piece.exercice_id = new_exercice_id
        piece.libelle = request.form['libelle']
        piece.reference = request.form.get('reference')
        piece.devise_id = request.form.get('devise_id') or None
        piece.taux_change = request.form.get('taux_change') or 1

        # Delete existing lines
        for ligne in piece.lignes:
            db.session.delete(ligne)

        # Add new lines
        comptes_ids = request.form.getlist('compte_id[]')
        projets_ids = request.form.getlist('projet_id[]')
        libelles = request.form.getlist('ligne_libelle[]')
        debits = request.form.getlist('debit[]')
        credits = request.form.getlist('credit[]')
        lignes_budget_ids = request.form.getlist('ligne_budget_id[]')

        for i in range(len(comptes_ids)):
            if comptes_ids[i]:
                ligne = LigneEcriture(
                    piece_id=piece.id,
                    compte_id=comptes_ids[i],
                    projet_id=projets_ids[i] if projets_ids[i] else None,
                    libelle=libelles[i] if i < len(libelles) else '',
                    debit=Decimal(debits[i] or 0),
                    credit=Decimal(credits[i] or 0),
                    ligne_budget_id=lignes_budget_ids[i] if i < len(lignes_budget_ids) and lignes_budget_ids[i] else None
                )
                db.session.add(ligne)

        # Validate balance
        db.session.flush()
        if not piece.est_equilibree:
            db.session.rollback()
            flash("Écriture déséquilibrée - Total Débit ≠ Total Crédit.", "danger")
            return redirect(url_for('modifier_ecriture', id=id))

        # Audit log
        new_values = {
            'date_piece': str(piece.date_piece),
            'libelle': piece.libelle,
            'reference': piece.reference
        }
        log_audit('pieces', id, 'UPDATE', old_values=old_values, new_values=new_values)

        db.session.commit()
        flash(f'Écriture {piece.numero} modifiée avec succès.', 'success')
        return redirect(url_for('detail_ecriture', id=id))

    return render_template('comptabilite/ecriture_edit.html',
                           piece=piece,
                           journaux=journaux,
                           exercices=exercices,
                           comptes=comptes,
                           projets=projets,
                           devises=devises,
                           lignes_budget=lignes_budget)


@app.route('/comptabilite/ecritures/<int:id>/valider', methods=['POST'])
@login_required
@role_required(['directeur'])
def valider_ecriture(id):
    """Valider une écriture comptable (Directeur uniquement)"""
    piece = PieceComptable.query.get_or_404(id)

    if piece.valide:
        flash('Cette écriture est déjà validée.', 'warning')
        return redirect(url_for('detail_ecriture', id=id))

    if not piece.est_equilibree:
        flash('Impossible de valider une écriture déséquilibrée.', 'danger')
        return redirect(url_for('detail_ecriture', id=id))

    piece.valide = True
    log_audit('pieces', id, 'VALIDATE', new_values={'valide': True})
    db.session.commit()

    flash(f'Écriture {piece.numero} validée avec succès.', 'success')
    return redirect(url_for('detail_ecriture', id=id))


@app.route('/comptabilite/ecritures/<int:id>/invalider', methods=['POST'])
@login_required
@role_required(['directeur'])
def invalider_ecriture(id):
    """Invalider une écriture comptable (Directeur uniquement)"""
    piece = PieceComptable.query.get_or_404(id)

    if not piece.valide:
        flash('Cette écriture n\'est pas validée.', 'warning')
        return redirect(url_for('detail_ecriture', id=id))

    # Vérifier que l'exercice n'est pas clôturé
    if piece.exercice.cloture:
        flash('Impossible de modifier une écriture d\'un exercice clôturé.', 'danger')
        return redirect(url_for('detail_ecriture', id=id))

    piece.valide = False
    log_audit('pieces', id, 'INVALIDATE', new_values={'valide': False})
    db.session.commit()

    flash(f'Écriture {piece.numero} invalidée.', 'warning')
    return redirect(url_for('detail_ecriture', id=id))


@app.route('/comptabilite/ecritures/<int:id>/dupliquer', methods=['POST'])
@login_required
@role_required(['comptable', 'directeur'])
def dupliquer_ecriture(id):
    """Dupliquer une écriture comptable existante"""
    piece_origine = PieceComptable.query.get_or_404(id)

    # Vérifier qu'un exercice est ouvert
    exercice = ExerciceComptable.query.filter_by(cloture=False).first()
    if not exercice:
        flash('Aucun exercice ouvert. Impossible de dupliquer.', 'danger')
        return redirect(url_for('detail_ecriture', id=id))

    # Générer nouveau numéro
    dernier = PieceComptable.query.order_by(PieceComptable.id.desc()).first()
    numero = f"PC{datetime.now().year}{(dernier.id + 1 if dernier else 1):05d}"

    # Créer la nouvelle pièce
    nouvelle_piece = PieceComptable(
        numero=numero,
        date_piece=date.today(),
        journal_id=piece_origine.journal_id,
        exercice_id=exercice.id,
        libelle=piece_origine.libelle,
        reference=None,  # Nouvelle référence à saisir
        devise_id=piece_origine.devise_id,
        taux_change=piece_origine.taux_change,
        valide=False  # Nouvelle écriture non validée
    )
    db.session.add(nouvelle_piece)
    db.session.flush()

    # Dupliquer les lignes
    for ligne_origine in piece_origine.lignes:
        nouvelle_ligne = LigneEcriture(
            piece_id=nouvelle_piece.id,
            compte_id=ligne_origine.compte_id,
            projet_id=ligne_origine.projet_id,
            libelle=ligne_origine.libelle,
            debit=ligne_origine.debit,
            credit=ligne_origine.credit,
            ligne_budget_id=ligne_origine.ligne_budget_id
        )
        db.session.add(nouvelle_ligne)
        db.session.flush()

        # Dupliquer les imputations analytiques si présentes
        if ligne_origine.imputations_analytiques:
            for imp in ligne_origine.imputations_analytiques:
                nouvelle_imp = ImputationAnalytique(
                    ligne_ecriture_id=nouvelle_ligne.id,
                    projet_id=imp.projet_id,
                    pourcentage=imp.pourcentage,
                    montant=imp.montant
                )
                db.session.add(nouvelle_imp)

    db.session.commit()

    log_audit('pieces', nouvelle_piece.id, 'CREATE',
              new_values={'duplique_de': piece_origine.numero})

    flash(f'Écriture dupliquée avec succès. Nouvelle écriture: {numero}', 'success')
    return redirect(url_for('modifier_ecriture', id=nouvelle_piece.id))


@app.route('/comptabilite/ecritures/valider-lot', methods=['POST'])
@login_required
@role_required(['directeur'])
def valider_lot_ecritures():
    """Valider plusieurs écritures en lot"""
    ids = request.form.getlist('piece_ids')
    count = 0

    for piece_id in ids:
        piece = PieceComptable.query.get(piece_id)
        if piece and not piece.valide and piece.est_equilibree:
            piece.valide = True
            log_audit('pieces', piece.id, 'VALIDATE', new_values={'valide': True})
            count += 1

    db.session.commit()
    flash(f'{count} écriture(s) validée(s) avec succès.', 'success')
    return redirect(url_for('liste_ecritures'))


# =============================================================================
# ROUTES - PIECES JUSTIFICATIVES
# =============================================================================

ALLOWED_EXTENSIONS = {'pdf', 'png', 'jpg', 'jpeg', 'gif'}

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


@app.route('/comptabilite/ecritures/<int:id>/pieces-justificatives')
@login_required
def liste_pieces_justificatives(id):
    """Liste des pièces justificatives d'une écriture"""
    piece = PieceComptable.query.get_or_404(id)
    return render_template('comptabilite/pieces_justificatives.html', piece=piece)


@app.route('/comptabilite/ecritures/<int:id>/upload', methods=['POST'])
@login_required
@role_required(['comptable', 'directeur'])
def upload_piece_justificative(id):
    """Upload d'une pièce justificative"""
    piece = PieceComptable.query.get_or_404(id)

    if 'fichier' not in request.files:
        flash('Aucun fichier sélectionné.', 'danger')
        return redirect(url_for('detail_ecriture', id=id))

    fichier = request.files['fichier']

    if fichier.filename == '':
        flash('Aucun fichier sélectionné.', 'danger')
        return redirect(url_for('detail_ecriture', id=id))

    if fichier and allowed_file(fichier.filename):
        # Créer un nom de fichier sécurisé
        filename = secure_filename(fichier.filename)
        # Ajouter un timestamp pour éviter les doublons
        timestamp = datetime.now().strftime('%Y%m%d%H%M%S')
        filename = f"{piece.numero}_{timestamp}_{filename}"

        # Créer le dossier par année/mois si nécessaire
        year_month = piece.date_piece.strftime('%Y/%m')
        upload_path = os.path.join(app.config['UPLOAD_FOLDER'], year_month)
        os.makedirs(upload_path, exist_ok=True)

        # Sauvegarder le fichier
        filepath = os.path.join(upload_path, filename)
        fichier.save(filepath)

        # Créer l'enregistrement en base
        pj = PieceJustificative(
            piece_comptable_id=piece.id,
            type_piece=request.form.get('type_piece', 'autre'),
            numero_piece=request.form.get('numero_piece'),
            fichier_path=os.path.join(year_month, filename),
            fichier_nom=fichier.filename,
            date_piece=datetime.strptime(request.form.get('date_piece'), '%Y-%m-%d').date() if request.form.get('date_piece') else None,
            description=request.form.get('description'),
            uploaded_by=current_user.email
        )
        db.session.add(pj)
        log_audit('pieces_justificatives', None, 'CREATE', new_values={'fichier': filename, 'piece_id': piece.id})
        db.session.commit()

        flash('Pièce justificative ajoutée avec succès.', 'success')
    else:
        flash('Type de fichier non autorisé. Formats acceptés: PDF, PNG, JPG, GIF', 'danger')

    return redirect(url_for('detail_ecriture', id=id))


@app.route('/uploads/<path:filename>')
@login_required
def uploaded_file(filename):
    """Servir les fichiers uploadés - SECURITY: Path traversal protection"""
    # Sanitize filename to prevent path traversal
    safe_filename = secure_filename(os.path.basename(filename))
    if not safe_filename:
        flash('Fichier non trouvé.', 'danger')
        return redirect(url_for('dashboard'))

    # Build safe path and verify it's within UPLOAD_FOLDER
    filepath = os.path.join(app.config['UPLOAD_FOLDER'], safe_filename)
    real_path = os.path.realpath(filepath)
    upload_folder = os.path.realpath(app.config['UPLOAD_FOLDER'])

    if not real_path.startswith(upload_folder):
        flash('Accès non autorisé.', 'danger')
        return redirect(url_for('dashboard'))

    if not os.path.exists(real_path):
        flash('Fichier non trouvé.', 'danger')
        return redirect(url_for('dashboard'))

    return send_file(real_path)


@app.route('/pieces-justificatives/<int:id>/supprimer', methods=['POST'])
@login_required
@role_required(['comptable', 'directeur'])
def supprimer_piece_justificative(id):
    """Supprimer une pièce justificative"""
    pj = PieceJustificative.query.get_or_404(id)
    piece_id = pj.piece_comptable_id

    # Supprimer le fichier physique
    filepath = os.path.join(app.config['UPLOAD_FOLDER'], pj.fichier_path)
    if os.path.exists(filepath):
        os.remove(filepath)

    log_audit('pieces_justificatives', id, 'DELETE', old_values={'fichier': pj.fichier_nom})
    db.session.delete(pj)
    db.session.commit()

    flash('Pièce justificative supprimée.', 'success')
    return redirect(url_for('detail_ecriture', id=piece_id))


# =============================================================================
# ROUTES - RECONCILIATION BANCAIRE
# =============================================================================

@app.route('/comptabilite/reconciliation-bancaire')
@login_required
def liste_reconciliations():
    """Liste des réconciliations bancaires"""
    reconciliations = ReconciliationBancaire.query.order_by(
        ReconciliationBancaire.date_reconciliation.desc()
    ).all()
    return render_template('comptabilite/reconciliations.html', reconciliations=reconciliations)


@app.route('/comptabilite/reconciliation-bancaire/nouvelle', methods=['GET', 'POST'])
@login_required
@role_required(['comptable', 'directeur'])
def nouvelle_reconciliation():
    """Créer une nouvelle réconciliation bancaire"""
    # Comptes bancaires (classe 5, sous-comptes 52x)
    comptes_banque = CompteComptable.query.filter(
        CompteComptable.numero.like('52%'),
        CompteComptable.actif == True
    ).order_by(CompteComptable.numero).all()

    if request.method == 'POST':
        compte_id = request.form.get('compte_id')
        periode_debut = datetime.strptime(request.form.get('periode_debut'), '%Y-%m-%d').date()
        periode_fin = datetime.strptime(request.form.get('periode_fin'), '%Y-%m-%d').date()
        solde_releve = Decimal(request.form.get('solde_releve', '0'))

        # Calculer le solde comptable
        compte = CompteComptable.query.get(compte_id)
        solde_comptable_query = db.session.query(
            db.func.sum(LigneEcriture.debit) - db.func.sum(LigneEcriture.credit)
        ).join(PieceComptable).filter(
            LigneEcriture.compte_id == compte_id,
            PieceComptable.date_piece <= periode_fin
        ).scalar() or 0

        reconciliation = ReconciliationBancaire(
            compte_id=compte_id,
            date_reconciliation=date.today(),
            periode_debut=periode_debut,
            periode_fin=periode_fin,
            solde_releve=solde_releve,
            solde_comptable=Decimal(str(solde_comptable_query)),
            ecart=solde_releve - Decimal(str(solde_comptable_query)),
            cree_par=current_user.email
        )
        db.session.add(reconciliation)
        db.session.flush()

        # Récupérer les écritures de la période pour ce compte
        lignes_ecritures = db.session.query(LigneEcriture).join(PieceComptable).filter(
            LigneEcriture.compte_id == compte_id,
            PieceComptable.date_piece >= periode_debut,
            PieceComptable.date_piece <= periode_fin
        ).all()

        for ligne in lignes_ecritures:
            ligne_recon = LigneReconciliation(
                reconciliation_id=reconciliation.id,
                ligne_ecriture_id=ligne.id,
                pointee=False
            )
            db.session.add(ligne_recon)

        log_audit('reconciliations_bancaires', reconciliation.id, 'CREATE',
                  new_values={'compte': compte.numero, 'periode': f'{periode_debut} - {periode_fin}'})
        db.session.commit()

        flash('Réconciliation bancaire créée.', 'success')
        return redirect(url_for('detail_reconciliation', id=reconciliation.id))

    return render_template('comptabilite/reconciliation_form.html', comptes=comptes_banque)


@app.route('/comptabilite/reconciliation-bancaire/<int:id>')
@login_required
def detail_reconciliation(id):
    """Détail d'une réconciliation bancaire"""
    reconciliation = ReconciliationBancaire.query.get_or_404(id)
    return render_template('comptabilite/reconciliation_detail.html', reconciliation=reconciliation)


@app.route('/comptabilite/reconciliation-bancaire/<int:id>/pointer', methods=['POST'])
@login_required
@role_required(['comptable', 'directeur'])
def pointer_lignes_reconciliation(id):
    """Pointer des lignes de réconciliation"""
    reconciliation = ReconciliationBancaire.query.get_or_404(id)

    if reconciliation.statut == 'validee':
        flash('Impossible de modifier une réconciliation validée.', 'danger')
        return redirect(url_for('detail_reconciliation', id=id))

    lignes_ids = request.form.getlist('lignes[]')

    # Mettre à jour le statut de pointage
    for ligne in reconciliation.lignes:
        if str(ligne.id) in lignes_ids:
            if not ligne.pointee:
                ligne.pointee = True
                ligne.date_pointage = datetime.utcnow()
                ligne.pointe_par = current_user.email
        else:
            ligne.pointee = False
            ligne.date_pointage = None
            ligne.pointe_par = None

    # Recalculer l'écart
    total_pointe = sum(
        float(l.ligne_ecriture.debit or 0) - float(l.ligne_ecriture.credit or 0)
        for l in reconciliation.lignes if l.pointee
    )
    reconciliation.ecart = reconciliation.solde_releve - Decimal(str(total_pointe))

    db.session.commit()
    flash('Pointage mis à jour.', 'success')
    return redirect(url_for('detail_reconciliation', id=id))


@app.route('/comptabilite/reconciliation-bancaire/<int:id>/valider', methods=['POST'])
@login_required
@role_required(['directeur'])
def valider_reconciliation(id):
    """Valider une réconciliation bancaire"""
    reconciliation = ReconciliationBancaire.query.get_or_404(id)

    if reconciliation.statut == 'validee':
        flash('Cette réconciliation est déjà validée.', 'warning')
        return redirect(url_for('detail_reconciliation', id=id))

    reconciliation.statut = 'validee'
    reconciliation.valide_par = current_user.email
    reconciliation.date_validation = datetime.utcnow()

    log_audit('reconciliations_bancaires', id, 'VALIDATE',
              new_values={'statut': 'validee', 'valide_par': current_user.email})
    db.session.commit()

    flash('Réconciliation validée.', 'success')
    return redirect(url_for('detail_reconciliation', id=id))


@app.route('/comptabilite/reconciliation-bancaire/<int:id>/pdf')
@login_required
def export_reconciliation_pdf(id):
    """Export PDF du rapport de réconciliation"""
    try:
        from xhtml2pdf import pisa
        from io import BytesIO
    except ImportError:
        flash("xhtml2pdf n'est pas installé.", "danger")
        return redirect(url_for('detail_reconciliation', id=id))

    reconciliation = ReconciliationBancaire.query.get_or_404(id)

    html = render_template('comptabilite/reconciliation_pdf.html',
                          reconciliation=reconciliation,
                          date_generation=datetime.now())

    try:
        pdf_buffer = BytesIO()
        pisa_status = pisa.CreatePDF(html, dest=pdf_buffer)
        if pisa_status.err:
            flash("Erreur lors de la génération PDF.", "danger")
            return redirect(url_for('detail_reconciliation', id=id))
        pdf_buffer.seek(0)
    except Exception as e:
        flash(f"Erreur lors de la génération PDF: {str(e)}", "danger")
        return redirect(url_for('detail_reconciliation', id=id))

    return Response(pdf_buffer.getvalue(),
                   mimetype='application/pdf',
                   headers={'Content-Disposition': f'attachment; filename=reconciliation_{reconciliation.compte.numero}_{reconciliation.date_reconciliation}.pdf'})


# =============================================================================
# ROUTES - GESTION DES AVANCES
# =============================================================================

@app.route('/comptabilite/avances')
@login_required
def liste_avances():
    """Liste des avances"""
    statut = request.args.get('statut', '')
    beneficiaire = request.args.get('beneficiaire', '')

    query = Avance.query

    if statut:
        query = query.filter(Avance.statut == statut)
    if beneficiaire:
        query = query.filter(Avance.beneficiaire.ilike(f'%{beneficiaire}%'))

    avances = query.order_by(Avance.date_avance.desc()).all()

    # Compter les avances en retard
    nb_retard = sum(1 for a in avances if a.est_en_retard)

    return render_template('comptabilite/avances.html',
                          avances=avances,
                          nb_retard=nb_retard)


@app.route('/comptabilite/avances/nouvelle', methods=['GET', 'POST'])
@login_required
@role_required(['comptable', 'directeur'])
def nouvelle_avance():
    """Créer une nouvelle avance"""
    projets = Projet.query.filter_by(statut='actif').all()

    if request.method == 'POST':
        # Générer numéro
        derniere = Avance.query.order_by(Avance.id.desc()).first()
        numero = f"AV{datetime.now().year}{(derniere.id + 1 if derniere else 1):04d}"

        date_avance = datetime.strptime(request.form.get('date_avance'), '%Y-%m-%d').date()

        avance = Avance(
            numero=numero,
            date_avance=date_avance,
            beneficiaire=request.form.get('beneficiaire'),
            montant=Decimal(request.form.get('montant')),
            objet=request.form.get('objet'),
            projet_id=request.form.get('projet_id') or None,
            date_limite=date_avance + timedelta(days=7),
            cree_par=current_user.email
        )
        db.session.add(avance)

        log_audit('avances', None, 'CREATE',
                  new_values={'numero': numero, 'beneficiaire': avance.beneficiaire, 'montant': str(avance.montant)})
        db.session.commit()

        flash(f'Avance {numero} créée. Date limite de justification: {avance.date_limite}', 'success')
        return redirect(url_for('liste_avances'))

    return render_template('comptabilite/avance_form.html',
                          projets=projets,
                          today=date.today().strftime('%Y-%m-%d'))


@app.route('/comptabilite/avances/<int:id>')
@login_required
def detail_avance(id):
    """Détail d'une avance"""
    avance = Avance.query.get_or_404(id)
    return render_template('comptabilite/avance_detail.html', avance=avance)


@app.route('/comptabilite/avances/<int:id>/justifier', methods=['GET', 'POST'])
@login_required
@role_required(['comptable', 'directeur'])
def justifier_avance(id):
    """Justifier une avance"""
    avance = Avance.query.get_or_404(id)

    if avance.statut not in ['en_attente', 'justifiee']:
        flash('Cette avance ne peut plus être justifiée.', 'danger')
        return redirect(url_for('detail_avance', id=id))

    if request.method == 'POST':
        montant_justifie = Decimal(request.form.get('montant_justifie', '0'))
        montant_rembourse = Decimal(request.form.get('montant_rembourse', '0'))

        avance.montant_justifie = montant_justifie
        avance.montant_rembourse = montant_rembourse
        avance.justification_notes = request.form.get('notes')
        avance.date_justification = datetime.utcnow()

        if avance.solde_restant <= 0:
            avance.statut = 'soldee'
            avance.date_solde = datetime.utcnow()
        else:
            avance.statut = 'justifiee'

        log_audit('avances', id, 'JUSTIFIER',
                  new_values={'montant_justifie': str(montant_justifie), 'statut': avance.statut})
        db.session.commit()

        flash('Justification enregistrée.', 'success')
        return redirect(url_for('detail_avance', id=id))

    return render_template('comptabilite/avance_justifier.html', avance=avance)


@app.route('/comptabilite/avances/<int:id>/deduire', methods=['POST'])
@login_required
@role_required(['directeur'])
def deduire_avance(id):
    """Marquer une avance comme déduite du salaire"""
    avance = Avance.query.get_or_404(id)

    if avance.statut == 'soldee' or avance.statut == 'deduite':
        flash('Cette avance est déjà soldée ou déduite.', 'warning')
        return redirect(url_for('detail_avance', id=id))

    avance.statut = 'deduite'
    avance.date_solde = datetime.utcnow()

    log_audit('avances', id, 'DEDUIRE',
              new_values={'statut': 'deduite', 'solde_restant': str(avance.solde_restant)})
    db.session.commit()

    flash(f'Avance marquée comme déduite. Montant à déduire: {avance.solde_restant:,.0f} FCFA', 'warning')
    return redirect(url_for('detail_avance', id=id))


# =============================================================================
# ROUTES - PETITE CAISSE
# =============================================================================

@app.route('/comptabilite/petite-caisse')
@login_required
def petite_caisse():
    """Gestion de la petite caisse"""
    # Comptes de caisse (57x)
    comptes_caisse = CompteComptable.query.filter(
        CompteComptable.numero.like('57%'),
        CompteComptable.actif == True
    ).order_by(CompteComptable.numero).all()

    compte_id = request.args.get('compte_id', type=int)
    if not compte_id and comptes_caisse:
        compte_id = comptes_caisse[0].id

    # Calculer le solde actuel
    if compte_id:
        solde = db.session.query(
            db.func.sum(LigneEcriture.debit) - db.func.sum(LigneEcriture.credit)
        ).filter(LigneEcriture.compte_id == compte_id).scalar() or 0

        # Mouvements récents
        mouvements = db.session.query(LigneEcriture).join(PieceComptable).filter(
            LigneEcriture.compte_id == compte_id
        ).order_by(PieceComptable.date_piece.desc()).limit(50).all()
    else:
        solde = 0
        mouvements = []

    return render_template('comptabilite/petite_caisse.html',
                          comptes=comptes_caisse,
                          compte_id=compte_id,
                          solde=float(solde),
                          mouvements=mouvements)


# =============================================================================
# ROUTES - JOURNAUX COMPTABLES
# =============================================================================

@app.route('/admin/journaux')
@login_required
@role_required(['comptable', 'directeur'])
def liste_journaux():
    """Liste des journaux comptables"""
    journaux = Journal.query.order_by(Journal.code).all()

    # Comptes de trésorerie pour l'affichage
    comptes_tresorerie = CompteComptable.query.filter(
        CompteComptable.classe == 5,
        CompteComptable.actif == True
    ).order_by(CompteComptable.numero).all()

    return render_template('admin/journaux.html',
                          journaux=journaux,
                          comptes_tresorerie=comptes_tresorerie)


@app.route('/admin/journaux/nouveau', methods=['GET', 'POST'])
@login_required
@role_required(['directeur'])
def nouveau_journal():
    """Créer un nouveau journal"""
    comptes_tresorerie = CompteComptable.query.filter(
        CompteComptable.classe == 5,
        CompteComptable.actif == True
    ).order_by(CompteComptable.numero).all()

    if request.method == 'POST':
        code = request.form.get('code', '').upper()
        nom = request.form.get('nom')
        type_journal = request.form.get('type_journal')
        compte_tresorerie_id = request.form.get('compte_tresorerie_id') or None

        # Vérifier unicité du code
        if Journal.query.filter_by(code=code).first():
            flash(f"Le code {code} existe déjà.", "danger")
            return redirect(url_for('nouveau_journal'))

        journal = Journal(
            code=code,
            nom=nom,
            type_journal=type_journal,
            compte_tresorerie_id=compte_tresorerie_id
        )
        db.session.add(journal)
        db.session.commit()

        flash(f"Journal {code} créé avec succès.", "success")
        return redirect(url_for('liste_journaux'))

    return render_template('admin/journal_form.html',
                          journal=None,
                          comptes_tresorerie=comptes_tresorerie)


@app.route('/admin/journaux/<int:id>/modifier', methods=['GET', 'POST'])
@login_required
@role_required(['directeur'])
def modifier_journal(id):
    """Modifier un journal"""
    journal = Journal.query.get_or_404(id)
    comptes_tresorerie = CompteComptable.query.filter(
        CompteComptable.classe == 5,
        CompteComptable.actif == True
    ).order_by(CompteComptable.numero).all()

    if request.method == 'POST':
        journal.code = request.form.get('code', '').upper()
        journal.nom = request.form.get('nom')
        journal.type_journal = request.form.get('type_journal')
        journal.compte_tresorerie_id = request.form.get('compte_tresorerie_id') or None

        db.session.commit()
        flash(f"Journal {journal.code} mis à jour.", "success")
        return redirect(url_for('liste_journaux'))

    return render_template('admin/journal_form.html',
                          journal=journal,
                          comptes_tresorerie=comptes_tresorerie)


@app.route('/admin/journaux/<int:id>/supprimer', methods=['POST'])
@login_required
@role_required(['directeur'])
def supprimer_journal(id):
    """Supprimer un journal (si pas d'écritures)"""
    journal = Journal.query.get_or_404(id)

    # Vérifier s'il y a des écritures
    nb_ecritures = PieceComptable.query.filter_by(journal_id=journal.id).count()
    if nb_ecritures > 0:
        flash(f"Impossible de supprimer : {nb_ecritures} écriture(s) utilisent ce journal.", "danger")
        return redirect(url_for('liste_journaux'))

    db.session.delete(journal)
    db.session.commit()
    flash(f"Journal {journal.code} supprimé.", "success")
    return redirect(url_for('liste_journaux'))


# =============================================================================
# ROUTES - COMPTES DE TRESORERIE (Banques, Caisses, Wave)
# =============================================================================

@app.route('/tresorerie/comptes')
@login_required
def liste_comptes_tresorerie():
    """Liste des comptes de trésorerie avec détails"""
    # Comptes de trésorerie (classe 5)
    comptes = CompteComptable.query.filter(
        CompteComptable.classe == 5,
        CompteComptable.actif == True
    ).order_by(CompteComptable.numero).all()

    # Calculer les soldes
    comptes_avec_soldes = []
    for compte in comptes:
        solde_debit = db.session.query(db.func.sum(LigneEcriture.debit)).filter(
            LigneEcriture.compte_id == compte.id
        ).scalar() or 0
        solde_credit = db.session.query(db.func.sum(LigneEcriture.credit)).filter(
            LigneEcriture.compte_id == compte.id
        ).scalar() or 0
        solde = float(solde_debit) - float(solde_credit)

        comptes_avec_soldes.append({
            'compte': compte,
            'details': compte.details_bancaires,
            'solde': solde
        })

    return render_template('tresorerie/comptes.html', comptes=comptes_avec_soldes)


@app.route('/tresorerie/comptes/<int:id>/details', methods=['GET', 'POST'])
@login_required
@role_required(['comptable', 'directeur'])
def details_compte_tresorerie(id):
    """Ajouter/modifier les détails d'un compte de trésorerie"""
    compte = CompteComptable.query.get_or_404(id)

    if compte.classe != 5:
        flash("Ce compte n'est pas un compte de trésorerie.", "danger")
        return redirect(url_for('liste_comptes_tresorerie'))

    details = compte.details_bancaires or CompteTresorerie(compte_id=compte.id)
    devises = Devise.query.all()

    if request.method == 'POST':
        details.compte_id = compte.id
        details.type_tresorerie = request.form.get('type_tresorerie', 'banque')

        # Détails bancaires
        if details.type_tresorerie == 'banque':
            details.nom_banque = request.form.get('nom_banque')
            details.numero_compte = request.form.get('numero_compte')
            details.iban = request.form.get('iban')
            details.code_swift = request.form.get('code_swift')
            details.agence = request.form.get('agence')
            details.adresse_agence = request.form.get('adresse_agence')
            details.titulaire = request.form.get('titulaire')

        # Détails mobile money
        elif details.type_tresorerie == 'mobile_money':
            details.operateur = request.form.get('operateur')
            details.numero_telephone = request.form.get('numero_telephone')
            details.nom_marchand = request.form.get('nom_marchand')

        # Informations générales
        details.devise_id = request.form.get('devise_id') or None
        details.solde_ouverture = request.form.get('solde_ouverture') or 0
        if request.form.get('date_ouverture'):
            details.date_ouverture = datetime.strptime(request.form['date_ouverture'], '%Y-%m-%d').date()
        details.plafond = request.form.get('plafond') or None
        details.notes = request.form.get('notes')

        if not compte.details_bancaires:
            db.session.add(details)

        db.session.commit()
        flash(f"Détails du compte {compte.numero} mis à jour.", "success")
        return redirect(url_for('liste_comptes_tresorerie'))

    return render_template('tresorerie/compte_details.html',
                          compte=compte,
                          details=details,
                          devises=devises)


@app.route('/tresorerie/nouveau-compte', methods=['GET', 'POST'])
@login_required
@role_required(['comptable', 'directeur'])
def nouveau_compte_tresorerie():
    """Créer un nouveau compte de trésorerie"""
    devises = Devise.query.all()

    if request.method == 'POST':
        numero = request.form.get('numero')
        intitule = request.form.get('intitule')
        type_tresorerie = request.form.get('type_tresorerie', 'banque')

        # Vérifier que le numéro commence par 5
        if not numero.startswith('5'):
            flash("Le numéro d'un compte de trésorerie doit commencer par 5.", "danger")
            return redirect(url_for('nouveau_compte_tresorerie'))

        # Vérifier l'unicité
        if CompteComptable.query.filter_by(numero=numero).first():
            flash(f"Le compte {numero} existe déjà.", "danger")
            return redirect(url_for('nouveau_compte_tresorerie'))

        # Créer le compte comptable
        compte = CompteComptable(
            numero=numero,
            intitule=intitule,
            classe=5,
            type_compte='actif',
            actif=True
        )
        db.session.add(compte)
        db.session.flush()

        # Créer les détails de trésorerie
        details = CompteTresorerie(
            compte_id=compte.id,
            type_tresorerie=type_tresorerie
        )

        if type_tresorerie == 'banque':
            details.nom_banque = request.form.get('nom_banque')
            details.numero_compte = request.form.get('numero_compte')
            details.iban = request.form.get('iban')
            details.code_swift = request.form.get('code_swift')
            details.titulaire = request.form.get('titulaire')
        elif type_tresorerie == 'mobile_money':
            details.operateur = request.form.get('operateur')
            details.numero_telephone = request.form.get('numero_telephone')
            details.nom_marchand = request.form.get('nom_marchand')

        details.devise_id = request.form.get('devise_id') or None
        details.solde_ouverture = request.form.get('solde_ouverture') or 0
        details.notes = request.form.get('notes')

        db.session.add(details)
        db.session.commit()

        flash(f"Compte {numero} - {intitule} créé avec succès.", "success")
        return redirect(url_for('liste_comptes_tresorerie'))

    return render_template('tresorerie/nouveau_compte.html', devises=devises)


# =============================================================================
# ROUTES - FOURNISSEURS
# =============================================================================

@app.route('/fournisseurs')
@login_required
def liste_fournisseurs():
    """Liste des fournisseurs"""
    categorie = request.args.get('categorie', '')
    recherche = request.args.get('q', '')

    query = Fournisseur.query.filter_by(actif=True)

    if categorie:
        query = query.filter(Fournisseur.categorie == categorie)
    if recherche:
        query = query.filter(
            db.or_(
                Fournisseur.nom.ilike(f'%{recherche}%'),
                Fournisseur.code.ilike(f'%{recherche}%')
            )
        )

    fournisseurs = query.order_by(Fournisseur.nom).all()

    # Catégories disponibles
    categories = db.session.query(Fournisseur.categorie).filter(
        Fournisseur.categorie.isnot(None),
        Fournisseur.actif == True
    ).distinct().all()
    categories = [c[0] for c in categories if c[0]]

    return render_template('fournisseurs/liste.html',
                          fournisseurs=fournisseurs,
                          categories=categories)


@app.route('/fournisseurs/nouveau', methods=['GET', 'POST'])
@login_required
@role_required(['comptable', 'directeur'])
def nouveau_fournisseur():
    """Créer un nouveau fournisseur"""
    comptes_401 = CompteComptable.query.filter(
        CompteComptable.numero.like('401%'),
        CompteComptable.actif == True
    ).order_by(CompteComptable.numero).all()

    if request.method == 'POST':
        # Générer code
        dernier = Fournisseur.query.order_by(Fournisseur.id.desc()).first()
        numero = (dernier.id + 1) if dernier else 1
        code = f"FRN{numero:03d}"

        fournisseur = Fournisseur(
            code=code,
            nom=request.form.get('nom'),
            categorie=request.form.get('categorie'),
            contact=request.form.get('contact'),
            telephone=request.form.get('telephone'),
            email=request.form.get('email'),
            adresse=request.form.get('adresse'),
            ville=request.form.get('ville'),
            ninea=request.form.get('ninea'),
            compte_comptable_id=request.form.get('compte_comptable_id') or None,
            notes=request.form.get('notes'),
            cree_par=current_user.email
        )

        db.session.add(fournisseur)
        db.session.commit()

        log_audit('fournisseurs', fournisseur.id, 'CREATE', None, {'nom': fournisseur.nom})

        flash(f'Fournisseur {fournisseur.nom} créé avec succès.', 'success')
        return redirect(url_for('liste_fournisseurs'))

    return render_template('fournisseurs/form.html', fournisseur=None, comptes_401=comptes_401)


@app.route('/fournisseurs/<int:id>/modifier', methods=['GET', 'POST'])
@login_required
@role_required(['comptable', 'directeur'])
def modifier_fournisseur(id):
    """Modifier un fournisseur"""
    fournisseur = Fournisseur.query.get_or_404(id)
    comptes_401 = CompteComptable.query.filter(
        CompteComptable.numero.like('401%'),
        CompteComptable.actif == True
    ).order_by(CompteComptable.numero).all()

    if request.method == 'POST':
        old_values = {'nom': fournisseur.nom}

        fournisseur.nom = request.form.get('nom')
        fournisseur.categorie = request.form.get('categorie')
        fournisseur.contact = request.form.get('contact')
        fournisseur.telephone = request.form.get('telephone')
        fournisseur.email = request.form.get('email')
        fournisseur.adresse = request.form.get('adresse')
        fournisseur.ville = request.form.get('ville')
        fournisseur.ninea = request.form.get('ninea')
        fournisseur.compte_comptable_id = request.form.get('compte_comptable_id') or None
        fournisseur.notes = request.form.get('notes')

        db.session.commit()

        log_audit('fournisseurs', fournisseur.id, 'UPDATE', old_values, {'nom': fournisseur.nom})

        flash('Fournisseur mis à jour.', 'success')
        return redirect(url_for('liste_fournisseurs'))

    return render_template('fournisseurs/form.html', fournisseur=fournisseur, comptes_401=comptes_401)


@app.route('/fournisseurs/<int:id>')
@login_required
def details_fournisseur(id):
    """Détails et historique d'un fournisseur"""
    fournisseur = Fournisseur.query.get_or_404(id)

    # Historique des paiements (écritures mentionnant ce fournisseur dans le libellé)
    # ou liées au compte comptable du fournisseur
    ecritures = []
    if fournisseur.compte_comptable_id:
        ecritures = LigneEcriture.query.filter(
            LigneEcriture.compte_id == fournisseur.compte_comptable_id
        ).join(PieceComptable).order_by(PieceComptable.date_piece.desc()).limit(50).all()

    # Calculer total
    total_debit = sum(float(e.debit or 0) for e in ecritures)
    total_credit = sum(float(e.credit or 0) for e in ecritures)

    return render_template('fournisseurs/details.html',
                          fournisseur=fournisseur,
                          ecritures=ecritures,
                          total_debit=total_debit,
                          total_credit=total_credit)


@app.route('/fournisseurs/<int:id>/supprimer', methods=['POST'])
@login_required
@role_required(['directeur'])
def supprimer_fournisseur(id):
    """Désactiver un fournisseur (soft delete)"""
    fournisseur = Fournisseur.query.get_or_404(id)
    fournisseur.actif = False
    db.session.commit()

    log_audit('fournisseurs', fournisseur.id, 'DELETE', {'nom': fournisseur.nom}, None)

    flash('Fournisseur supprimé.', 'success')
    return redirect(url_for('liste_fournisseurs'))


@app.route('/api/fournisseurs/search')
@login_required
def api_search_fournisseurs():
    """API pour recherche de fournisseurs (autocomplétion)"""
    q = request.args.get('q', '')
    if len(q) < 2:
        return jsonify([])

    fournisseurs = Fournisseur.query.filter(
        Fournisseur.actif == True,
        db.or_(
            Fournisseur.nom.ilike(f'%{q}%'),
            Fournisseur.code.ilike(f'%{q}%')
        )
    ).limit(10).all()

    return jsonify([{
        'id': f.id,
        'code': f.code,
        'nom': f.nom,
        'categorie': f.categorie
    } for f in fournisseurs])


# =============================================================================
# ROUTES - NOTES DE FRAIS
# =============================================================================

@app.route('/notes-frais')
@login_required
def liste_notes_frais():
    """Liste des notes de frais"""
    statut = request.args.get('statut', '')
    projet_id = request.args.get('projet_id', '')

    # Si l'utilisateur n'est pas comptable/directeur, ne voir que ses propres notes
    if current_user.role == 'auditeur':
        query = NoteFrais.query
    elif current_user.role in ['comptable', 'directeur']:
        query = NoteFrais.query
    else:
        query = NoteFrais.query.filter(NoteFrais.employe_id == current_user.id)

    if statut:
        query = query.filter(NoteFrais.statut == statut)
    if projet_id:
        query = query.filter(NoteFrais.projet_id == int(projet_id))

    notes = query.order_by(NoteFrais.date_creation.desc()).all()
    projets = Projet.query.filter_by(statut='actif').all()

    # Statistiques
    stats = {
        'total': len(notes),
        'en_attente': sum(1 for n in notes if n.statut == 'soumis'),
        'montant_en_attente': sum(float(n.montant or 0) for n in notes if n.statut == 'soumis')
    }

    return render_template('notes_frais/liste.html',
                          notes=notes,
                          projets=projets,
                          stats=stats,
                          statut_filtre=statut,
                          projet_filtre=projet_id)


@app.route('/notes-frais/nouvelle', methods=['GET', 'POST'])
@login_required
def nouvelle_note_frais():
    """Créer une nouvelle note de frais"""
    projets = Projet.query.filter_by(statut='actif').all()

    if request.method == 'POST':
        # Générer numéro
        annee = datetime.now().year
        derniere = NoteFrais.query.filter(
            NoteFrais.numero.like(f'NF-{annee}-%')
        ).order_by(NoteFrais.id.desc()).first()

        num = 1
        if derniere:
            try:
                num = int(derniere.numero.split('-')[-1]) + 1
            except ValueError:
                pass
        numero = f"NF-{annee}-{num:04d}"

        # Créer la note de frais
        note = NoteFrais(
            numero=numero,
            employe_id=current_user.id,
            projet_id=request.form.get('projet_id') or None,
            ligne_budget_id=request.form.get('ligne_budget_id') or None,
            date_depense=datetime.strptime(request.form.get('date_depense'), '%Y-%m-%d').date(),
            montant=Decimal(request.form.get('montant')),
            categorie=request.form.get('categorie'),
            description=request.form.get('description'),
            a_rembourser=request.form.get('a_rembourser') == 'on',
            cree_par=current_user.email
        )

        # Gérer l'upload du justificatif
        if 'justificatif' in request.files:
            fichier = request.files['justificatif']
            if fichier and fichier.filename:
                from werkzeug.utils import secure_filename
                filename = secure_filename(fichier.filename)
                # Créer un nom unique
                ext = filename.rsplit('.', 1)[-1] if '.' in filename else ''
                unique_filename = f"nf_{numero}_{datetime.now().strftime('%Y%m%d%H%M%S')}.{ext}"
                filepath = os.path.join(app.config['UPLOAD_FOLDER'], 'notes_frais', unique_filename)
                os.makedirs(os.path.dirname(filepath), exist_ok=True)
                fichier.save(filepath)
                note.justificatif = f"notes_frais/{unique_filename}"

        db.session.add(note)
        db.session.commit()

        log_audit('notes_frais', note.id, 'CREATE', None, {
            'numero': note.numero,
            'montant': str(note.montant),
            'categorie': note.categorie
        })

        flash(f'Note de frais {numero} créée avec succès.', 'success')
        return redirect(url_for('detail_note_frais', id=note.id))

    return render_template('notes_frais/formulaire.html',
                          note=None,
                          projets=projets,
                          categories=NoteFrais.CATEGORIES)


@app.route('/notes-frais/<int:id>')
@login_required
def detail_note_frais(id):
    """Détail d'une note de frais"""
    note = NoteFrais.query.get_or_404(id)

    # Vérifier accès
    if current_user.role not in ['comptable', 'directeur', 'auditeur'] and note.employe_id != current_user.id:
        flash("Vous n'avez pas accès à cette note de frais.", 'danger')
        return redirect(url_for('liste_notes_frais'))

    return render_template('notes_frais/detail.html', note=note)


@app.route('/notes-frais/<int:id>/modifier', methods=['GET', 'POST'])
@login_required
def modifier_note_frais(id):
    """Modifier une note de frais (brouillon ou rejetée)"""
    note = NoteFrais.query.get_or_404(id)

    # Vérifier accès et statut
    if note.employe_id != current_user.id and current_user.role != 'directeur':
        flash("Vous n'avez pas le droit de modifier cette note de frais.", 'danger')
        return redirect(url_for('liste_notes_frais'))

    if not note.est_modifiable:
        flash("Cette note de frais ne peut plus être modifiée.", 'warning')
        return redirect(url_for('detail_note_frais', id=id))

    projets = Projet.query.filter_by(statut='actif').all()

    if request.method == 'POST':
        old_values = {
            'montant': str(note.montant),
            'categorie': note.categorie,
            'description': note.description
        }

        note.projet_id = request.form.get('projet_id') or None
        note.ligne_budget_id = request.form.get('ligne_budget_id') or None
        note.date_depense = datetime.strptime(request.form.get('date_depense'), '%Y-%m-%d').date()
        note.montant = Decimal(request.form.get('montant'))
        note.categorie = request.form.get('categorie')
        note.description = request.form.get('description')
        note.a_rembourser = request.form.get('a_rembourser') == 'on'

        # Remettre en brouillon si rejetée
        if note.statut == 'rejete':
            note.statut = 'brouillon'
            note.motif_rejet = None

        # Gérer nouveau justificatif
        if 'justificatif' in request.files:
            fichier = request.files['justificatif']
            if fichier and fichier.filename:
                from werkzeug.utils import secure_filename
                filename = secure_filename(fichier.filename)
                ext = filename.rsplit('.', 1)[-1] if '.' in filename else ''
                unique_filename = f"nf_{note.numero}_{datetime.now().strftime('%Y%m%d%H%M%S')}.{ext}"
                filepath = os.path.join(app.config['UPLOAD_FOLDER'], 'notes_frais', unique_filename)
                os.makedirs(os.path.dirname(filepath), exist_ok=True)
                fichier.save(filepath)
                note.justificatif = f"notes_frais/{unique_filename}"

        db.session.commit()

        log_audit('notes_frais', note.id, 'UPDATE', old_values, {
            'montant': str(note.montant),
            'categorie': note.categorie,
            'description': note.description
        })

        flash('Note de frais modifiée avec succès.', 'success')
        return redirect(url_for('detail_note_frais', id=id))

    return render_template('notes_frais/formulaire.html',
                          note=note,
                          projets=projets,
                          categories=NoteFrais.CATEGORIES)


@app.route('/notes-frais/<int:id>/soumettre', methods=['POST'])
@login_required
def soumettre_note_frais(id):
    """Soumettre une note de frais pour approbation"""
    note = NoteFrais.query.get_or_404(id)

    if note.employe_id != current_user.id:
        flash("Vous ne pouvez soumettre que vos propres notes de frais.", 'danger')
        return redirect(url_for('liste_notes_frais'))

    if not note.peut_soumettre:
        flash("Cette note de frais ne peut pas être soumise.", 'warning')
        return redirect(url_for('detail_note_frais', id=id))

    note.statut = 'soumis'
    note.date_soumission = datetime.utcnow()
    db.session.commit()

    log_audit('notes_frais', note.id, 'UPDATE', {'statut': 'brouillon'}, {'statut': 'soumis'})

    flash(f'Note de frais {note.numero} soumise pour approbation.', 'success')
    return redirect(url_for('detail_note_frais', id=id))


@app.route('/notes-frais/<int:id>/approuver', methods=['POST'])
@login_required
@role_required(['comptable', 'directeur'])
def approuver_note_frais(id):
    """Approuver une note de frais"""
    note = NoteFrais.query.get_or_404(id)

    if not note.peut_approuver:
        flash("Cette note de frais ne peut pas être approuvée.", 'warning')
        return redirect(url_for('detail_note_frais', id=id))

    note.statut = 'approuve'
    note.validateur_id = current_user.id
    note.date_validation = datetime.utcnow()
    db.session.commit()

    log_audit('notes_frais', note.id, 'UPDATE', {'statut': 'soumis'}, {'statut': 'approuve'})

    flash(f'Note de frais {note.numero} approuvée.', 'success')
    return redirect(url_for('detail_note_frais', id=id))


@app.route('/notes-frais/<int:id>/rejeter', methods=['POST'])
@login_required
@role_required(['comptable', 'directeur'])
def rejeter_note_frais(id):
    """Rejeter une note de frais"""
    note = NoteFrais.query.get_or_404(id)

    if not note.peut_approuver:
        flash("Cette note de frais ne peut pas être rejetée.", 'warning')
        return redirect(url_for('detail_note_frais', id=id))

    motif = request.form.get('motif_rejet', '')
    if not motif:
        flash("Veuillez indiquer un motif de rejet.", 'warning')
        return redirect(url_for('detail_note_frais', id=id))

    note.statut = 'rejete'
    note.validateur_id = current_user.id
    note.date_validation = datetime.utcnow()
    note.motif_rejet = motif
    db.session.commit()

    log_audit('notes_frais', note.id, 'UPDATE', {'statut': 'soumis'}, {'statut': 'rejete', 'motif': motif})

    flash(f'Note de frais {note.numero} rejetée.', 'warning')
    return redirect(url_for('detail_note_frais', id=id))


@app.route('/notes-frais/<int:id>/rembourser', methods=['POST'])
@login_required
@role_required(['comptable', 'directeur'])
def rembourser_note_frais(id):
    """Marquer une note de frais comme remboursée"""
    note = NoteFrais.query.get_or_404(id)

    if note.statut != 'approuve':
        flash("Seules les notes approuvées peuvent être remboursées.", 'warning')
        return redirect(url_for('detail_note_frais', id=id))

    note.statut = 'rembourse'
    note.date_remboursement = datetime.utcnow()
    db.session.commit()

    log_audit('notes_frais', note.id, 'UPDATE', {'statut': 'approuve'}, {'statut': 'rembourse'})

    flash(f'Note de frais {note.numero} marquée comme remboursée.', 'success')
    return redirect(url_for('detail_note_frais', id=id))


@app.route('/notes-frais/<int:id>/pdf')
@login_required
def pdf_note_frais(id):
    """Générer PDF de la note de frais"""
    note = NoteFrais.query.get_or_404(id)

    # Vérifier accès
    if current_user.role not in ['comptable', 'directeur', 'auditeur'] and note.employe_id != current_user.id:
        flash("Vous n'avez pas accès à cette note de frais.", 'danger')
        return redirect(url_for('liste_notes_frais'))

    html = render_template('notes_frais/pdf.html', note=note, org=ORG_INFO)

    try:
        from weasyprint import HTML
        pdf = HTML(string=html, base_url=request.host_url).write_pdf()
        response = Response(pdf, mimetype='application/pdf')
        response.headers['Content-Disposition'] = f'inline; filename=note_frais_{note.numero}.pdf'
        return response
    except ImportError:
        # Fallback: retourner HTML
        return html


@app.route('/api/lignes-budget/<int:projet_id>')
@login_required
def api_lignes_budget_projet(projet_id):
    """API pour obtenir les lignes budget d'un projet"""
    lignes = LigneBudget.query.filter_by(projet_id=projet_id).order_by(LigneBudget.code).all()
    return jsonify([{
        'id': l.id,
        'code': l.code,
        'intitule': l.intitule,
        'montant_prevu': float(l.montant_prevu or 0)
    } for l in lignes])


# =============================================================================
# ROUTES - DEMANDES D'ACHAT
# =============================================================================

@app.route('/achats/demandes')
@login_required
def liste_demandes_achat():
    """Liste des demandes d'achat"""
    statut = request.args.get('statut', '')
    projet_id = request.args.get('projet_id', '')

    query = DemandeAchat.query

    if statut:
        query = query.filter(DemandeAchat.statut == statut)
    if projet_id:
        query = query.filter(DemandeAchat.projet_id == int(projet_id))

    demandes = query.order_by(DemandeAchat.date_creation.desc()).all()
    projets = Projet.query.filter_by(statut='actif').all()

    # Statistiques
    stats = {
        'total': len(demandes),
        'en_attente': sum(1 for d in demandes if d.statut == 'soumis'),
        'montant_en_attente': sum(d.montant_total for d in demandes if d.statut == 'soumis')
    }

    return render_template('achats/demandes_liste.html',
                          demandes=demandes,
                          projets=projets,
                          stats=stats,
                          statut_filtre=statut,
                          projet_filtre=projet_id)


@app.route('/achats/demandes/nouvelle', methods=['GET', 'POST'])
@login_required
def nouvelle_demande_achat():
    """Créer une nouvelle demande d'achat"""
    projets = Projet.query.filter_by(statut='actif').all()

    if request.method == 'POST':
        # Générer numéro
        annee = datetime.now().year
        derniere = DemandeAchat.query.filter(
            DemandeAchat.numero.like(f'DA-{annee}-%')
        ).order_by(DemandeAchat.id.desc()).first()

        num = 1
        if derniere:
            try:
                num = int(derniere.numero.split('-')[-1]) + 1
            except ValueError:
                pass
        numero = f"DA-{annee}-{num:04d}"

        # Créer la demande
        demande = DemandeAchat(
            numero=numero,
            demandeur_id=current_user.id,
            projet_id=request.form.get('projet_id') or None,
            ligne_budget_id=request.form.get('ligne_budget_id') or None,
            date_demande=datetime.strptime(request.form.get('date_demande'), '%Y-%m-%d').date(),
            objet=request.form.get('objet'),
            description=request.form.get('description'),
            urgence=request.form.get('urgence', 'normal'),
            cree_par=current_user.email
        )

        db.session.add(demande)
        db.session.flush()  # Pour obtenir l'ID

        # Ajouter les lignes
        designations = request.form.getlist('designation[]')
        quantites = request.form.getlist('quantite[]')
        unites = request.form.getlist('unite[]')
        prix = request.form.getlist('prix_unitaire[]')

        for i, designation in enumerate(designations):
            if designation.strip():
                ligne = LigneDemandeAchat(
                    demande_id=demande.id,
                    designation=designation,
                    quantite=Decimal(quantites[i]) if quantites[i] else 1,
                    unite=unites[i] if i < len(unites) else 'piece',
                    prix_unitaire_estime=Decimal(prix[i]) if prix[i] else 0
                )
                db.session.add(ligne)

        # Mettre à jour le montant estimé
        demande.montant_estime = Decimal(str(demande.montant_total))

        db.session.commit()

        log_audit('demandes_achat', demande.id, 'CREATE', None, {
            'numero': demande.numero,
            'objet': demande.objet,
            'montant_estime': str(demande.montant_estime)
        })

        flash(f'Demande d\'achat {numero} créée avec succès.', 'success')
        return redirect(url_for('detail_demande_achat', id=demande.id))

    return render_template('achats/demande_formulaire.html',
                          demande=None,
                          projets=projets,
                          urgences=DemandeAchat.URGENCES,
                          unites=LigneDemandeAchat.UNITES)


@app.route('/achats/demandes/<int:id>')
@login_required
def detail_demande_achat(id):
    """Détail d'une demande d'achat"""
    demande = DemandeAchat.query.get_or_404(id)
    fournisseurs = Fournisseur.query.filter_by(actif=True).order_by(Fournisseur.nom).all()
    return render_template('achats/demande_detail.html', demande=demande, fournisseurs=fournisseurs)


@app.route('/achats/demandes/<int:id>/modifier', methods=['GET', 'POST'])
@login_required
def modifier_demande_achat(id):
    """Modifier une demande d'achat"""
    demande = DemandeAchat.query.get_or_404(id)

    if not demande.est_modifiable:
        flash("Cette demande ne peut plus être modifiée.", 'warning')
        return redirect(url_for('detail_demande_achat', id=id))

    projets = Projet.query.filter_by(statut='actif').all()

    if request.method == 'POST':
        old_values = {
            'objet': demande.objet,
            'montant_estime': str(demande.montant_estime)
        }

        demande.projet_id = request.form.get('projet_id') or None
        demande.ligne_budget_id = request.form.get('ligne_budget_id') or None
        demande.date_demande = datetime.strptime(request.form.get('date_demande'), '%Y-%m-%d').date()
        demande.objet = request.form.get('objet')
        demande.description = request.form.get('description')
        demande.urgence = request.form.get('urgence', 'normal')

        # Remettre en brouillon si rejetée
        if demande.statut == 'rejete':
            demande.statut = 'brouillon'
            demande.motif_rejet = None

        # Supprimer les anciennes lignes et recréer
        LigneDemandeAchat.query.filter_by(demande_id=demande.id).delete()

        designations = request.form.getlist('designation[]')
        quantites = request.form.getlist('quantite[]')
        unites = request.form.getlist('unite[]')
        prix = request.form.getlist('prix_unitaire[]')

        for i, designation in enumerate(designations):
            if designation.strip():
                ligne = LigneDemandeAchat(
                    demande_id=demande.id,
                    designation=designation,
                    quantite=Decimal(quantites[i]) if quantites[i] else 1,
                    unite=unites[i] if i < len(unites) else 'piece',
                    prix_unitaire_estime=Decimal(prix[i]) if prix[i] else 0
                )
                db.session.add(ligne)

        demande.montant_estime = Decimal(str(demande.montant_total))

        db.session.commit()

        log_audit('demandes_achat', demande.id, 'UPDATE', old_values, {
            'objet': demande.objet,
            'montant_estime': str(demande.montant_estime)
        })

        flash('Demande d\'achat modifiée avec succès.', 'success')
        return redirect(url_for('detail_demande_achat', id=id))

    return render_template('achats/demande_formulaire.html',
                          demande=demande,
                          projets=projets,
                          urgences=DemandeAchat.URGENCES,
                          unites=LigneDemandeAchat.UNITES)


@app.route('/achats/demandes/<int:id>/soumettre', methods=['POST'])
@login_required
def soumettre_demande_achat(id):
    """Soumettre une demande d'achat pour approbation"""
    demande = DemandeAchat.query.get_or_404(id)

    if not demande.peut_soumettre:
        flash("Cette demande ne peut pas être soumise (ajoutez au moins une ligne).", 'warning')
        return redirect(url_for('detail_demande_achat', id=id))

    demande.statut = 'soumis'
    demande.date_soumission = datetime.utcnow()
    db.session.commit()

    log_audit('demandes_achat', demande.id, 'UPDATE', {'statut': 'brouillon'}, {'statut': 'soumis'})

    if demande.necessite_approbation_directeur:
        flash(f'Demande {demande.numero} soumise. Montant >= 500,000 FCFA : approbation du directeur requise.', 'info')
    else:
        flash(f'Demande {demande.numero} soumise pour approbation.', 'success')

    return redirect(url_for('detail_demande_achat', id=id))


@app.route('/achats/demandes/<int:id>/approuver', methods=['POST'])
@login_required
@role_required(['comptable', 'directeur'])
def approuver_demande_achat(id):
    """Approuver une demande d'achat"""
    demande = DemandeAchat.query.get_or_404(id)

    if not demande.peut_approuver:
        flash("Cette demande ne peut pas être approuvée.", 'warning')
        return redirect(url_for('detail_demande_achat', id=id))

    # Vérifier le seuil d'approbation
    if demande.necessite_approbation_directeur and current_user.role != 'directeur':
        flash("Cette demande nécessite l'approbation du directeur (montant >= 500,000 FCFA).", 'warning')
        return redirect(url_for('detail_demande_achat', id=id))

    demande.statut = 'approuve'
    demande.approbateur_id = current_user.id
    demande.date_approbation = datetime.utcnow()
    db.session.commit()

    log_audit('demandes_achat', demande.id, 'UPDATE', {'statut': 'soumis'}, {'statut': 'approuve'})

    flash(f'Demande {demande.numero} approuvée. Vous pouvez maintenant générer un bon de commande.', 'success')
    return redirect(url_for('detail_demande_achat', id=id))


@app.route('/achats/demandes/<int:id>/rejeter', methods=['POST'])
@login_required
@role_required(['comptable', 'directeur'])
def rejeter_demande_achat(id):
    """Rejeter une demande d'achat"""
    demande = DemandeAchat.query.get_or_404(id)

    if not demande.peut_approuver:
        flash("Cette demande ne peut pas être rejetée.", 'warning')
        return redirect(url_for('detail_demande_achat', id=id))

    motif = request.form.get('motif_rejet', '')
    if not motif:
        flash("Veuillez indiquer un motif de rejet.", 'warning')
        return redirect(url_for('detail_demande_achat', id=id))

    demande.statut = 'rejete'
    demande.approbateur_id = current_user.id
    demande.date_approbation = datetime.utcnow()
    demande.motif_rejet = motif
    db.session.commit()

    log_audit('demandes_achat', demande.id, 'UPDATE', {'statut': 'soumis'}, {'statut': 'rejete', 'motif': motif})

    flash(f'Demande {demande.numero} rejetée.', 'warning')
    return redirect(url_for('detail_demande_achat', id=id))


@app.route('/achats/demandes/<int:id>/generer-bc', methods=['POST'])
@login_required
@role_required(['comptable', 'directeur'])
def generer_bon_commande(id):
    """Générer un bon de commande à partir d'une demande approuvée"""
    demande = DemandeAchat.query.get_or_404(id)

    if demande.statut != 'approuve':
        flash("Seules les demandes approuvées peuvent générer un bon de commande.", 'warning')
        return redirect(url_for('detail_demande_achat', id=id))

    fournisseur_id = request.form.get('fournisseur_id')
    if not fournisseur_id:
        flash("Veuillez sélectionner un fournisseur.", 'warning')
        return redirect(url_for('detail_demande_achat', id=id))

    # Générer numéro
    annee = datetime.now().year
    dernier = BonCommande.query.filter(
        BonCommande.numero.like(f'BC-{annee}-%')
    ).order_by(BonCommande.id.desc()).first()

    num = 1
    if dernier:
        try:
            num = int(dernier.numero.split('-')[-1]) + 1
        except ValueError:
            pass
    numero = f"BC-{annee}-{num:04d}"

    # Créer le bon de commande
    bon = BonCommande(
        numero=numero,
        demande_achat_id=demande.id,
        fournisseur_id=int(fournisseur_id),
        date_commande=date.today(),
        date_livraison_prevue=datetime.strptime(request.form.get('date_livraison'), '%Y-%m-%d').date() if request.form.get('date_livraison') else None,
        conditions_paiement=request.form.get('conditions_paiement'),
        adresse_livraison=request.form.get('adresse_livraison'),
        notes=request.form.get('notes'),
        cree_par=current_user.email
    )

    db.session.add(bon)
    db.session.flush()

    # Copier les lignes de la demande
    montant_total = 0
    for ligne_da in demande.lignes:
        ligne_bc = LigneBonCommande(
            bon_commande_id=bon.id,
            designation=ligne_da.designation,
            quantite=ligne_da.quantite,
            unite=ligne_da.unite,
            prix_unitaire=ligne_da.prix_unitaire_estime
        )
        db.session.add(ligne_bc)
        montant_total += ligne_bc.montant_total

    bon.montant_total = Decimal(str(montant_total))

    # Mettre à jour la demande
    demande.statut = 'commande'
    demande.bon_commande_id = bon.id

    db.session.commit()

    log_audit('bons_commande', bon.id, 'CREATE', None, {
        'numero': bon.numero,
        'demande': demande.numero,
        'montant_total': str(bon.montant_total)
    })

    flash(f'Bon de commande {numero} généré avec succès.', 'success')
    return redirect(url_for('detail_bon_commande', id=bon.id))


@app.route('/achats/bons-commande')
@login_required
def liste_bons_commande():
    """Liste des bons de commande"""
    statut = request.args.get('statut', '')

    query = BonCommande.query
    if statut:
        query = query.filter(BonCommande.statut == statut)

    bons = query.order_by(BonCommande.date_creation.desc()).all()

    return render_template('achats/bons_commande_liste.html', bons=bons, statut_filtre=statut)


@app.route('/achats/bons-commande/<int:id>')
@login_required
def detail_bon_commande(id):
    """Détail d'un bon de commande"""
    bon = BonCommande.query.get_or_404(id)
    return render_template('achats/bon_commande_detail.html', bon=bon)


@app.route('/achats/bons-commande/<int:id>/pdf')
@login_required
def pdf_bon_commande(id):
    """Générer PDF du bon de commande"""
    bon = BonCommande.query.get_or_404(id)

    html = render_template('achats/bon_commande_pdf.html', bon=bon, org=ORG_INFO)

    try:
        from weasyprint import HTML
        pdf = HTML(string=html, base_url=request.host_url).write_pdf()
        response = Response(pdf, mimetype='application/pdf')
        response.headers['Content-Disposition'] = f'inline; filename=bon_commande_{bon.numero}.pdf'
        return response
    except ImportError:
        return html


@app.route('/achats/bons-commande/<int:id>/livrer', methods=['POST'])
@login_required
@role_required(['comptable', 'directeur'])
def livrer_bon_commande(id):
    """Mettre à jour le statut de livraison d'un bon de commande"""
    bon = BonCommande.query.get_or_404(id)

    statut = request.form.get('statut_livraison', 'livre')
    if statut in ['livre_partiel', 'livre']:
        bon.statut = statut
        db.session.commit()

        log_audit('bons_commande', bon.id, 'UPDATE', {'statut': 'emis'}, {'statut': statut})

        flash(f'Bon de commande {bon.numero} marqué comme {statut.replace("_", " ")}.', 'success')

    return redirect(url_for('detail_bon_commande', id=id))


# =============================================================================
# ROUTES - IMMOBILISATIONS
# =============================================================================

@app.route('/comptabilite/immobilisations')
@login_required
def liste_immobilisations():
    """Liste des immobilisations"""
    categorie = request.args.get('categorie', '')
    statut = request.args.get('statut', 'actif')

    query = Immobilisation.query

    if categorie:
        query = query.filter(Immobilisation.categorie == categorie)
    if statut:
        query = query.filter(Immobilisation.statut == statut)

    immobilisations = query.order_by(Immobilisation.code).all()

    # Totaux
    total_acquisition = sum(float(i.valeur_acquisition or 0) for i in immobilisations)
    total_amortissement = sum(i.cumul_amortissement for i in immobilisations)
    total_vnc = sum(i.valeur_nette_comptable for i in immobilisations)

    return render_template('comptabilite/immobilisations.html',
                          immobilisations=immobilisations,
                          total_acquisition=total_acquisition,
                          total_amortissement=total_amortissement,
                          total_vnc=total_vnc)


@app.route('/comptabilite/immobilisations/nouvelle', methods=['GET', 'POST'])
@login_required
@role_required(['comptable', 'directeur'])
def nouvelle_immobilisation():
    """Créer une nouvelle immobilisation"""
    projets = Projet.query.filter_by(statut='actif').all()
    comptes_immo = CompteComptable.query.filter(
        CompteComptable.numero.like('2%'),
        CompteComptable.actif == True
    ).order_by(CompteComptable.numero).all()

    if request.method == 'POST':
        # Générer code
        categorie = request.form.get('categorie')
        prefix = {'informatique': 'IT', 'vehicule': 'VH', 'mobilier': 'MB', 'batiment': 'BT'}.get(categorie, 'IM')
        derniere = Immobilisation.query.filter(Immobilisation.code.like(f'{prefix}%')).order_by(Immobilisation.id.desc()).first()
        num = 1
        if derniere:
            try:
                num = int(derniere.code[2:]) + 1
            except ValueError:
                pass
        code = f"{prefix}{num:04d}"

        duree = int(request.form.get('duree_amortissement'))

        immobilisation = Immobilisation(
            code=code,
            designation=request.form.get('designation'),
            categorie=categorie,
            date_acquisition=datetime.strptime(request.form.get('date_acquisition'), '%Y-%m-%d').date(),
            valeur_acquisition=Decimal(request.form.get('valeur_acquisition')),
            duree_amortissement=duree,
            taux_amortissement=Decimal(str(100 / duree)) if duree > 0 else 0,
            compte_immobilisation_id=request.form.get('compte_immobilisation_id') or None,
            compte_amortissement_id=request.form.get('compte_amortissement_id') or None,
            compte_dotation_id=request.form.get('compte_dotation_id') or None,
            projet_id=request.form.get('projet_id') or None,
            localisation=request.form.get('localisation'),
            numero_serie=request.form.get('numero_serie'),
            fournisseur=request.form.get('fournisseur'),
            numero_facture=request.form.get('numero_facture'),
            cree_par=current_user.email
        )
        db.session.add(immobilisation)

        log_audit('immobilisations', None, 'CREATE',
                  new_values={'code': code, 'designation': immobilisation.designation, 'valeur': str(immobilisation.valeur_acquisition)})
        db.session.commit()

        flash(f'Immobilisation {code} créée.', 'success')
        return redirect(url_for('liste_immobilisations'))

    return render_template('comptabilite/immobilisation_form.html',
                          projets=projets,
                          comptes=comptes_immo,
                          durees_syscoa=Immobilisation.DUREES_SYSCOA,
                          today=date.today().strftime('%Y-%m-%d'))


@app.route('/comptabilite/immobilisations/<int:id>')
@login_required
def detail_immobilisation(id):
    """Détail d'une immobilisation avec tableau d'amortissement"""
    immobilisation = Immobilisation.query.get_or_404(id)
    return render_template('comptabilite/immobilisation_detail.html', immobilisation=immobilisation)


@app.route('/comptabilite/immobilisations/<int:id>/calculer-amortissement', methods=['POST'])
@login_required
@role_required(['comptable', 'directeur'])
def calculer_amortissement(id):
    """Calculer l'amortissement pour l'exercice en cours"""
    immobilisation = Immobilisation.query.get_or_404(id)

    if immobilisation.statut != 'actif':
        flash('Cette immobilisation n\'est plus active.', 'danger')
        return redirect(url_for('detail_immobilisation', id=id))

    # Récupérer l'exercice en cours
    exercice = ExerciceComptable.query.filter_by(cloture=False).first()
    if not exercice:
        flash('Aucun exercice ouvert.', 'danger')
        return redirect(url_for('detail_immobilisation', id=id))

    # Vérifier si déjà calculé pour cet exercice
    existant = LigneAmortissement.query.filter_by(
        immobilisation_id=id,
        exercice_id=exercice.id
    ).first()

    if existant:
        flash('L\'amortissement a déjà été calculé pour cet exercice.', 'warning')
        return redirect(url_for('detail_immobilisation', id=id))

    # Calculer la dotation
    dotation = Decimal(str(immobilisation.amortissement_annuel))
    cumul_precedent = Decimal(str(immobilisation.cumul_amortissement))
    cumul_nouveau = cumul_precedent + dotation
    vnc = Decimal(str(immobilisation.valeur_acquisition)) - cumul_nouveau

    # Prorata temporis si première année
    if immobilisation.date_acquisition.year == exercice.annee:
        mois_restants = 12 - immobilisation.date_acquisition.month + 1
        dotation = dotation * mois_restants / 12
        cumul_nouveau = cumul_precedent + dotation
        vnc = Decimal(str(immobilisation.valeur_acquisition)) - cumul_nouveau

    ligne_amort = LigneAmortissement(
        immobilisation_id=id,
        exercice_id=exercice.id,
        annee=exercice.annee,
        dotation=dotation,
        cumul=cumul_nouveau,
        valeur_nette=vnc
    )
    db.session.add(ligne_amort)

    log_audit('lignes_amortissement', None, 'CREATE',
              new_values={'immobilisation': immobilisation.code, 'exercice': exercice.annee, 'dotation': str(dotation)})
    db.session.commit()

    flash(f'Amortissement {exercice.annee} calculé: {dotation:,.0f} FCFA', 'success')
    return redirect(url_for('detail_immobilisation', id=id))


@app.route('/comptabilite/immobilisations/<int:id>/sortie', methods=['GET', 'POST'])
@login_required
@role_required(['directeur'])
def sortie_immobilisation(id):
    """Sortie d'une immobilisation (cession ou mise au rebut)"""
    immobilisation = Immobilisation.query.get_or_404(id)

    if immobilisation.statut != 'actif':
        flash('Cette immobilisation est déjà sortie.', 'danger')
        return redirect(url_for('detail_immobilisation', id=id))

    if request.method == 'POST':
        immobilisation.statut = request.form.get('motif_sortie')  # cede ou rebut
        immobilisation.date_sortie = datetime.strptime(request.form.get('date_sortie'), '%Y-%m-%d').date()
        immobilisation.motif_sortie = request.form.get('motif_sortie')
        if request.form.get('valeur_cession'):
            immobilisation.valeur_cession = Decimal(request.form.get('valeur_cession'))

        log_audit('immobilisations', id, 'SORTIE',
                  new_values={'statut': immobilisation.statut, 'date_sortie': str(immobilisation.date_sortie)})
        db.session.commit()

        flash('Sortie d\'immobilisation enregistrée.', 'success')
        return redirect(url_for('detail_immobilisation', id=id))

    return render_template('comptabilite/immobilisation_sortie.html',
                          immobilisation=immobilisation,
                          today=date.today().strftime('%Y-%m-%d'))


# =============================================================================
# ROUTES - IMPORT EXCEL
# =============================================================================

@app.route('/comptabilite/import')
@login_required
@role_required(['comptable', 'directeur'])
def import_ecritures_page():
    """Page d'import d'écritures depuis Excel"""
    return render_template('comptabilite/import_ecritures.html')


@app.route('/comptabilite/import/template')
@login_required
def download_import_template():
    """Télécharger le template Excel d'import"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        from io import BytesIO
    except ImportError:
        flash("openpyxl n'est pas installé.", "danger")
        return redirect(url_for('import_ecritures_page'))

    wb = Workbook()
    ws = wb.active
    ws.title = "Écritures"

    # En-têtes
    headers = ['Date', 'Journal', 'Libellé', 'Référence', 'Compte', 'Projet', 'Débit', 'Crédit']
    header_fill = PatternFill(start_color="4a7c59", end_color="4a7c59", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font

    # Exemple
    exemple = ['2025-01-15', 'BQ', 'Paiement fournisseur', 'FAC-001', '401', 'LED', '500000', '']
    for col, val in enumerate(exemple, 1):
        ws.cell(row=2, column=col, value=val)

    exemple2 = ['2025-01-15', 'BQ', 'Paiement fournisseur', 'FAC-001', '521', 'LED', '', '500000']
    for col, val in enumerate(exemple2, 1):
        ws.cell(row=3, column=col, value=val)

    # Ajuster largeur
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 10
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 15

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return Response(output.read(),
                   mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                   headers={'Content-Disposition': 'attachment; filename=template_import_ecritures.xlsx'})


@app.route('/comptabilite/import/upload', methods=['POST'])
@login_required
@role_required(['comptable', 'directeur'])
def upload_import_ecritures():
    """Traiter le fichier Excel d'import"""
    if 'fichier' not in request.files:
        flash('Aucun fichier sélectionné.', 'danger')
        return redirect(url_for('import_ecritures_page'))

    fichier = request.files['fichier']
    if fichier.filename == '' or not fichier.filename.endswith('.xlsx'):
        flash('Veuillez sélectionner un fichier .xlsx', 'danger')
        return redirect(url_for('import_ecritures_page'))

    try:
        from openpyxl import load_workbook
        from io import BytesIO
    except ImportError:
        flash("openpyxl n'est pas installé.", "danger")
        return redirect(url_for('import_ecritures_page'))

    wb = load_workbook(BytesIO(fichier.read()))
    ws = wb.active

    exercice = ExerciceComptable.query.filter_by(cloture=False).first()
    if not exercice:
        flash('Aucun exercice ouvert.', 'danger')
        return redirect(url_for('import_ecritures_page'))

    erreurs = []
    pieces_creees = 0
    lignes_creees = 0

    # Regrouper les lignes par date+journal+libelle+reference
    ecritures = {}
    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row[0]:  # Skip empty rows
            continue

        date_str, journal_code, libelle, reference, compte_num, projet_code, debit, credit = row[:8]

        key = (str(date_str), journal_code, libelle, reference)
        if key not in ecritures:
            ecritures[key] = []

        ecritures[key].append({
            'row': row_num,
            'compte_num': str(compte_num) if compte_num else None,
            'projet_code': projet_code,
            'debit': float(debit) if debit else 0,
            'credit': float(credit) if credit else 0
        })

    # Créer les pièces
    for (date_str, journal_code, libelle, reference), lignes in ecritures.items():
        try:
            # Parser la date
            if isinstance(date_str, str):
                date_piece = datetime.strptime(date_str, '%Y-%m-%d').date()
            else:
                date_piece = date_str

            # Trouver le journal
            journal = Journal.query.filter_by(code=journal_code).first()
            if not journal:
                erreurs.append(f"Ligne {lignes[0]['row']}: Journal '{journal_code}' introuvable")
                continue

            # Générer numéro
            dernier = PieceComptable.query.order_by(PieceComptable.id.desc()).first()
            numero = f"IMP{datetime.now().year}{(dernier.id + 1 if dernier else 1):05d}"

            piece = PieceComptable(
                numero=numero,
                date_piece=date_piece,
                journal_id=journal.id,
                exercice_id=exercice.id,
                libelle=libelle,
                reference=reference
            )
            db.session.add(piece)
            db.session.flush()

            total_debit = 0
            total_credit = 0

            for ligne_data in lignes:
                # Trouver le compte
                compte = CompteComptable.query.filter_by(numero=ligne_data['compte_num']).first()
                if not compte:
                    erreurs.append(f"Ligne {ligne_data['row']}: Compte '{ligne_data['compte_num']}' introuvable")
                    continue

                # Trouver le projet si spécifié
                projet_id = None
                if ligne_data['projet_code']:
                    projet = Projet.query.filter_by(code=ligne_data['projet_code']).first()
                    if projet:
                        projet_id = projet.id

                ligne = LigneEcriture(
                    piece_id=piece.id,
                    compte_id=compte.id,
                    projet_id=projet_id,
                    libelle=libelle,
                    debit=ligne_data['debit'],
                    credit=ligne_data['credit']
                )
                db.session.add(ligne)
                total_debit += ligne_data['debit']
                total_credit += ligne_data['credit']
                lignes_creees += 1

            # Vérifier équilibre
            if abs(total_debit - total_credit) > 0.01:
                erreurs.append(f"Écriture du {date_piece}: Déséquilibrée (D={total_debit:,.0f} C={total_credit:,.0f})")
                db.session.rollback()
                continue

            pieces_creees += 1

        except Exception as e:
            erreurs.append(f"Erreur: {str(e)}")
            continue

    if pieces_creees > 0:
        db.session.commit()

    if erreurs:
        for err in erreurs[:10]:  # Limiter à 10 erreurs affichées
            flash(err, 'warning')
        if len(erreurs) > 10:
            flash(f'... et {len(erreurs) - 10} autres erreurs', 'warning')

    flash(f'Import terminé: {pieces_creees} pièce(s) créée(s), {lignes_creees} ligne(s)', 'success')
    return redirect(url_for('liste_ecritures'))


# =============================================================================
# ROUTES - MODELES D'ECRITURES RECURRENTES
# =============================================================================

@app.route('/comptabilite/modeles')
@login_required
def liste_modeles_ecritures():
    """Liste des modèles d'écritures récurrentes"""
    modeles = ModeleEcriture.query.order_by(ModeleEcriture.nom).all()
    return render_template('comptabilite/modeles_ecritures.html', modeles=modeles)


@app.route('/comptabilite/modeles/nouveau', methods=['GET', 'POST'])
@login_required
@role_required(['comptable', 'directeur'])
def nouveau_modele_ecriture():
    """Créer un nouveau modèle d'écriture"""
    journaux = Journal.query.all()
    comptes = CompteComptable.query.order_by(CompteComptable.numero).all()
    projets = Projet.query.filter_by(statut='actif').all()

    if request.method == 'POST':
        modele = ModeleEcriture(
            nom=request.form['nom'],
            description=request.form.get('description'),
            journal_id=request.form['journal_id'],
            libelle=request.form['libelle'],
            frequence=request.form.get('frequence'),
            jour_execution=request.form.get('jour_execution') or None,
            actif=request.form.get('actif') == 'on',
            cree_par=current_user.email
        )
        db.session.add(modele)
        db.session.flush()

        # Ajouter les lignes
        comptes_ids = request.form.getlist('ligne_compte_id[]')
        types = request.form.getlist('ligne_type[]')
        montants = request.form.getlist('ligne_montant[]')
        libelles = request.form.getlist('ligne_libelle[]')
        projets_ids = request.form.getlist('ligne_projet_id[]')

        for i in range(len(comptes_ids)):
            if comptes_ids[i]:
                ligne = LigneModeleEcriture(
                    modele_id=modele.id,
                    compte_id=int(comptes_ids[i]),
                    type_montant=types[i] if i < len(types) else 'debit',
                    montant=Decimal(montants[i]) if i < len(montants) and montants[i] else 0,
                    libelle=libelles[i] if i < len(libelles) else '',
                    projet_id=int(projets_ids[i]) if i < len(projets_ids) and projets_ids[i] else None
                )
                db.session.add(ligne)

        db.session.commit()
        flash(f'Modèle "{modele.nom}" créé avec succès', 'success')
        return redirect(url_for('liste_modeles_ecritures'))

    return render_template('comptabilite/modele_form.html',
                           journaux=journaux, comptes=comptes, projets=projets)


@app.route('/comptabilite/modeles/<int:id>')
@login_required
def detail_modele_ecriture(id):
    """Détail d'un modèle d'écriture"""
    modele = ModeleEcriture.query.get_or_404(id)
    return render_template('comptabilite/modele_detail.html', modele=modele)


@app.route('/comptabilite/modeles/<int:id>/modifier', methods=['GET', 'POST'])
@login_required
@role_required(['comptable', 'directeur'])
def modifier_modele_ecriture(id):
    """Modifier un modèle d'écriture"""
    modele = ModeleEcriture.query.get_or_404(id)
    journaux = Journal.query.all()
    comptes = CompteComptable.query.order_by(CompteComptable.numero).all()
    projets = Projet.query.filter_by(statut='actif').all()

    if request.method == 'POST':
        modele.nom = request.form['nom']
        modele.description = request.form.get('description')
        modele.journal_id = request.form['journal_id']
        modele.libelle = request.form['libelle']
        modele.frequence = request.form.get('frequence')
        modele.jour_execution = request.form.get('jour_execution') or None
        modele.actif = request.form.get('actif') == 'on'

        # Supprimer anciennes lignes et recréer
        LigneModeleEcriture.query.filter_by(modele_id=modele.id).delete()

        comptes_ids = request.form.getlist('ligne_compte_id[]')
        types = request.form.getlist('ligne_type[]')
        montants = request.form.getlist('ligne_montant[]')
        libelles = request.form.getlist('ligne_libelle[]')
        projets_ids = request.form.getlist('ligne_projet_id[]')

        for i in range(len(comptes_ids)):
            if comptes_ids[i]:
                ligne = LigneModeleEcriture(
                    modele_id=modele.id,
                    compte_id=int(comptes_ids[i]),
                    type_montant=types[i] if i < len(types) else 'debit',
                    montant=Decimal(montants[i]) if i < len(montants) and montants[i] else 0,
                    libelle=libelles[i] if i < len(libelles) else '',
                    projet_id=int(projets_ids[i]) if i < len(projets_ids) and projets_ids[i] else None
                )
                db.session.add(ligne)

        db.session.commit()
        flash(f'Modèle "{modele.nom}" modifié avec succès', 'success')
        return redirect(url_for('liste_modeles_ecritures'))

    return render_template('comptabilite/modele_form.html',
                           modele=modele, journaux=journaux, comptes=comptes, projets=projets)


@app.route('/comptabilite/modeles/<int:id>/supprimer', methods=['POST'])
@login_required
@role_required(['directeur'])
def supprimer_modele_ecriture(id):
    """Supprimer un modèle d'écriture"""
    modele = ModeleEcriture.query.get_or_404(id)
    nom = modele.nom
    db.session.delete(modele)
    db.session.commit()
    flash(f'Modèle "{nom}" supprimé', 'success')
    return redirect(url_for('liste_modeles_ecritures'))


@app.route('/comptabilite/modeles/<int:id>/generer', methods=['POST'])
@login_required
@role_required(['comptable', 'directeur'])
def generer_ecriture_modele(id):
    """Générer une écriture à partir d'un modèle"""
    modele = ModeleEcriture.query.get_or_404(id)

    exercice = ExerciceComptable.query.filter_by(cloture=False).first()
    if not exercice:
        flash('Aucun exercice ouvert.', 'danger')
        return redirect(url_for('liste_modeles_ecritures'))

    # Générer numéro
    dernier = PieceComptable.query.order_by(PieceComptable.id.desc()).first()
    numero = f"PC{datetime.now().year}{(dernier.id + 1 if dernier else 1):05d}"

    piece = PieceComptable(
        numero=numero,
        date_piece=date.today(),
        journal_id=modele.journal_id,
        exercice_id=exercice.id,
        libelle=modele.libelle
    )
    db.session.add(piece)
    db.session.flush()

    for ligne_modele in modele.lignes:
        ligne = LigneEcriture(
            piece_id=piece.id,
            compte_id=ligne_modele.compte_id,
            projet_id=ligne_modele.projet_id,
            libelle=ligne_modele.libelle,
            debit=ligne_modele.montant if ligne_modele.type_montant == 'debit' else 0,
            credit=ligne_modele.montant if ligne_modele.type_montant == 'credit' else 0
        )
        db.session.add(ligne)

    modele.derniere_execution = date.today()
    db.session.commit()

    flash(f'Écriture {numero} créée depuis le modèle "{modele.nom}"', 'success')
    return redirect(url_for('detail_ecriture', id=piece.id))


# =============================================================================
# ROUTES - TAUX DE CHANGE
# =============================================================================

@app.route('/admin/taux-change')
@login_required
@role_required(['comptable', 'directeur'])
def liste_taux_change():
    """Liste des taux de change"""
    annee = request.args.get('annee', date.today().year, type=int)

    taux = TauxChange.query.filter(
        TauxChange.annee == annee
    ).order_by(TauxChange.devise_id, TauxChange.mois).all()

    devises = Devise.query.filter(Devise.code != 'XOF').all()
    annees = db.session.query(db.func.distinct(TauxChange.annee)).order_by(TauxChange.annee.desc()).all()
    annees = [a[0] for a in annees] or [date.today().year]

    return render_template('admin/taux_change.html',
                          taux=taux,
                          devises=devises,
                          annee=annee,
                          annees=annees)


@app.route('/admin/taux-change/nouveau', methods=['POST'])
@login_required
@role_required(['comptable', 'directeur'])
def nouveau_taux_change():
    """Ajouter un nouveau taux de change"""
    devise_id = request.form.get('devise_id')
    mois = int(request.form.get('mois'))
    annee = int(request.form.get('annee'))
    taux_value = Decimal(request.form.get('taux'))

    # Vérifier si existe déjà
    existant = TauxChange.query.filter_by(
        devise_id=devise_id,
        mois=mois,
        annee=annee
    ).first()

    if existant:
        existant.taux = taux_value
        existant.date_saisie = datetime.utcnow()
        existant.saisi_par = current_user.email
        flash('Taux de change mis à jour.', 'success')
    else:
        taux = TauxChange(
            devise_id=devise_id,
            mois=mois,
            annee=annee,
            taux=taux_value,
            saisi_par=current_user.email
        )
        db.session.add(taux)
        flash('Taux de change ajouté.', 'success')

    db.session.commit()
    return redirect(url_for('liste_taux_change', annee=annee))


# =============================================================================
# ROUTES - CLOTURE EXERCICE
# =============================================================================

@app.route('/admin/exercices')
@login_required
@role_required(['directeur'])
def liste_exercices():
    """Liste des exercices comptables"""
    exercices = ExerciceComptable.query.order_by(ExerciceComptable.annee.desc()).all()
    return render_template('admin/exercices.html', exercices=exercices)


@app.route('/admin/exercices/nouveau', methods=['GET', 'POST'])
@login_required
@role_required(['directeur'])
def nouvel_exercice():
    """Créer un nouvel exercice comptable"""
    if request.method == 'POST':
        annee = int(request.form.get('annee'))

        if ExerciceComptable.query.filter_by(annee=annee).first():
            flash(f'L\'exercice {annee} existe déjà.', 'danger')
        else:
            exercice = ExerciceComptable(
                annee=annee,
                date_debut=date(annee, 1, 1),
                date_fin=date(annee, 12, 31),
                cloture=False
            )
            db.session.add(exercice)
            log_audit('exercices', None, 'CREATE', new_values={'annee': annee})
            db.session.commit()
            flash(f'Exercice {annee} créé avec succès.', 'success')
            return redirect(url_for('liste_exercices'))

    return render_template('admin/exercice_form.html')


@app.route('/admin/exercices/<int:id>/cloturer', methods=['GET', 'POST'])
@login_required
@role_required(['directeur'])
def cloturer_exercice(id):
    """Clôturer un exercice comptable"""
    exercice = ExerciceComptable.query.get_or_404(id)

    if exercice.cloture:
        flash('Cet exercice est déjà clôturé.', 'warning')
        return redirect(url_for('liste_exercices'))

    # Vérifications avant clôture
    ecritures_non_validees = PieceComptable.query.filter_by(
        exercice_id=id,
        valide=False
    ).count()

    if request.method == 'POST':
        if ecritures_non_validees > 0 and not request.form.get('force'):
            flash(f'Il reste {ecritures_non_validees} écriture(s) non validée(s).', 'danger')
            return redirect(url_for('cloturer_exercice', id=id))

        # Effectuer la clôture
        exercice.cloture = True

        # Calculer le résultat de l'exercice
        # Produits (classe 7) - Charges (classe 6)
        produits = db.session.query(
            db.func.sum(LigneEcriture.credit) - db.func.sum(LigneEcriture.debit)
        ).join(CompteComptable).join(PieceComptable).filter(
            CompteComptable.classe == 7,
            PieceComptable.exercice_id == id
        ).scalar() or 0

        charges = db.session.query(
            db.func.sum(LigneEcriture.debit) - db.func.sum(LigneEcriture.credit)
        ).join(CompteComptable).join(PieceComptable).filter(
            CompteComptable.classe == 6,
            PieceComptable.exercice_id == id
        ).scalar() or 0

        resultat = float(produits) - float(charges)

        log_audit('exercices', id, 'CLOTURE', new_values={
            'cloture': True,
            'resultat': resultat,
            'ecritures_non_validees_ignorees': ecritures_non_validees if request.form.get('force') else 0
        })
        db.session.commit()

        flash(f'Exercice {exercice.annee} clôturé. Résultat: {resultat:,.0f} FCFA', 'success')
        return redirect(url_for('liste_exercices'))

    # Statistiques pour la page de confirmation
    stats = {
        'nb_ecritures': PieceComptable.query.filter_by(exercice_id=id).count(),
        'nb_non_validees': ecritures_non_validees,
        'total_debit': db.session.query(
            db.func.sum(LigneEcriture.debit)
        ).join(PieceComptable).filter(
            PieceComptable.exercice_id == id
        ).scalar() or 0,
        'total_credit': db.session.query(
            db.func.sum(LigneEcriture.credit)
        ).join(PieceComptable).filter(
            PieceComptable.exercice_id == id
        ).scalar() or 0
    }

    return render_template('admin/cloture_exercice.html', exercice=exercice, stats=stats)


# =============================================================================
# ROUTES - MOT DE PASSE OUBLIE
# =============================================================================

import secrets

# Stockage temporaire des tokens (en production, utiliser Redis ou la BD)
password_reset_tokens = {}


@app.route('/mot-de-passe-oublie', methods=['GET', 'POST'])
def mot_de_passe_oublie():
    """Demande de réinitialisation de mot de passe"""
    if current_user.is_authenticated:
        return redirect(url_for('dashboard'))

    if request.method == 'POST':
        email = request.form.get('email')
        user = Utilisateur.query.filter_by(email=email).first()

        if user and user.actif:
            # Générer un token
            token = secrets.token_urlsafe(32)
            password_reset_tokens[token] = {
                'user_id': user.id,
                'expires': datetime.utcnow() + timedelta(hours=1)
            }

            # En production, envoyer un email avec le lien
            reset_url = url_for('reinitialiser_mot_de_passe', token=token, _external=True)

            # Pour le développement, afficher le lien
            flash(f'Un lien de réinitialisation a été généré. En production, il serait envoyé par email.', 'info')
            flash(f'Lien (dev): {reset_url}', 'warning')

            log_audit('utilisateurs', user.id, 'PASSWORD_RESET_REQUEST')
            db.session.commit()
        else:
            # Ne pas révéler si l'email existe ou non (sécurité)
            flash('Si cette adresse email est associée à un compte, un lien de réinitialisation a été envoyé.', 'info')

        return redirect(url_for('login'))

    return render_template('auth/mot_de_passe_oublie.html')


@app.route('/reinitialiser-mot-de-passe/<token>', methods=['GET', 'POST'])
def reinitialiser_mot_de_passe(token):
    """Réinitialisation du mot de passe avec token"""
    if current_user.is_authenticated:
        return redirect(url_for('dashboard'))

    # Vérifier le token
    token_data = password_reset_tokens.get(token)
    if not token_data or token_data['expires'] < datetime.utcnow():
        flash('Ce lien de réinitialisation est invalide ou a expiré.', 'danger')
        return redirect(url_for('mot_de_passe_oublie'))

    user = Utilisateur.query.get(token_data['user_id'])
    if not user:
        flash('Utilisateur introuvable.', 'danger')
        return redirect(url_for('login'))

    if request.method == 'POST':
        password = request.form.get('password')
        password_confirm = request.form.get('password_confirm')

        if len(password) < 6:
            flash('Le mot de passe doit contenir au moins 6 caractères.', 'danger')
        elif password != password_confirm:
            flash('Les mots de passe ne correspondent pas.', 'danger')
        else:
            user.password_hash = generate_password_hash(password)
            log_audit('utilisateurs', user.id, 'PASSWORD_RESET_COMPLETE')
            db.session.commit()

            # Supprimer le token utilisé
            del password_reset_tokens[token]

            flash('Votre mot de passe a été réinitialisé avec succès. Vous pouvez maintenant vous connecter.', 'success')
            return redirect(url_for('login'))

    return render_template('auth/reinitialiser_mot_de_passe.html', token=token, email=user.email)


@app.route('/changer-mot-de-passe', methods=['GET', 'POST'])
@login_required
def changer_mot_de_passe():
    """Changer son propre mot de passe"""
    if request.method == 'POST':
        current_password = request.form.get('current_password')
        new_password = request.form.get('new_password')
        confirm_password = request.form.get('confirm_password')

        if not check_password_hash(current_user.password_hash, current_password):
            flash('Mot de passe actuel incorrect.', 'danger')
        elif len(new_password) < 6:
            flash('Le nouveau mot de passe doit contenir au moins 6 caractères.', 'danger')
        elif new_password != confirm_password:
            flash('Les nouveaux mots de passe ne correspondent pas.', 'danger')
        else:
            current_user.password_hash = generate_password_hash(new_password)
            log_audit('utilisateurs', current_user.id, 'PASSWORD_CHANGE')
            db.session.commit()
            flash('Votre mot de passe a été changé avec succès.', 'success')
            return redirect(url_for('dashboard'))

    return render_template('auth/changer_mot_de_passe.html')


# =============================================================================
# ROUTES - RAPPORTS
# =============================================================================

@app.route('/rapports')
@login_required
def rapports():
    """Page des rapports"""
    projets = Projet.query.all()
    exercices = ExerciceComptable.query.order_by(ExerciceComptable.annee.desc()).all()
    return render_template('rapports/index.html', projets=projets, exercices=exercices)


@app.route('/rapports/balance')
@login_required
def balance_generale():
    """Balance générale"""
    exercice_id = request.args.get('exercice_id')
    inclure_non_validees = request.args.get('inclure_non_validees', 'false') == 'true'

    # Requête pour calculer les soldes par compte
    query = db.session.query(
        CompteComptable.numero,
        CompteComptable.intitule,
        db.func.sum(LigneEcriture.debit).label('total_debit'),
        db.func.sum(LigneEcriture.credit).label('total_credit')
    ).join(
        LigneEcriture, LigneEcriture.compte_id == CompteComptable.id
    ).join(
        PieceComptable, PieceComptable.id == LigneEcriture.piece_id
    )

    # SECURITY: Par défaut, n'inclure que les écritures validées
    if not inclure_non_validees:
        query = query.filter(PieceComptable.valide == True)

    if exercice_id:
        query = query.filter(PieceComptable.exercice_id == exercice_id)

    query = query.group_by(CompteComptable.id).order_by(CompteComptable.numero)

    balance = []
    for row in query.all():
        debit = float(row.total_debit or 0)
        credit = float(row.total_credit or 0)
        solde_debit = debit - credit if debit > credit else 0
        solde_credit = credit - debit if credit > debit else 0
        balance.append({
            'numero': row.numero,
            'intitule': row.intitule,
            'debit': debit,
            'credit': credit,
            'solde_debit': solde_debit,
            'solde_credit': solde_credit
        })

    exercices = ExerciceComptable.query.order_by(ExerciceComptable.annee.desc()).all()
    return render_template('rapports/balance.html', balance=balance, exercices=exercices, exercice_id=exercice_id)


# =============================================================================
# HELPER FUNCTIONS FOR REPORTS
# =============================================================================

MOIS_NOMS = ['', 'Janvier', 'Février', 'Mars', 'Avril', 'Mai', 'Juin',
             'Juillet', 'Août', 'Septembre', 'Octobre', 'Novembre', 'Décembre']


def get_periode_label(periode, annee, trimestre, mois, date_debut, date_fin):
    """Génère un label lisible pour la période"""
    if periode == 'year' and annee:
        return f"Année {annee}"
    elif periode == 'quarter' and annee and trimestre:
        return f"T{trimestre} {annee}"
    elif periode == 'month' and annee and mois:
        return f"{MOIS_NOMS[mois]} {annee}"
    elif periode == 'custom' and date_debut and date_fin:
        return f"{date_debut.strftime('%d/%m/%Y')} - {date_fin.strftime('%d/%m/%Y')}"
    return "Toute la période"


def get_categorie_nom(categorie_id):
    """Retourne le nom d'une catégorie à partir de son ID"""
    if categorie_id:
        cat = CategorieBudget.query.get(categorie_id)
        return cat.nom if cat else None
    return None


def get_ligne_nom(ligne_budget_id):
    """Retourne le nom d'une ligne budgétaire à partir de son ID"""
    if ligne_budget_id:
        ligne = LigneBudget.query.get(ligne_budget_id)
        return f"{ligne.code} - {ligne.intitule}" if ligne else None
    return None


def parse_report_filters():
    """Parse common report filters from request args"""
    # Filtre période
    periode = request.args.get('periode', 'all')  # all, year, quarter, month, custom
    annee = request.args.get('annee', type=int)
    trimestre = request.args.get('trimestre', type=int)  # 1, 2, 3, 4
    mois = request.args.get('mois', type=int)
    date_debut_str = request.args.get('date_debut')
    date_fin_str = request.args.get('date_fin')

    # Filtre section (catégorie budgétaire)
    categorie_id = request.args.get('categorie_id', type=int)

    # Filtre ligne budgétaire spécifique
    ligne_budget_id = request.args.get('ligne_budget_id', type=int)

    # Calculer les dates de filtre
    date_filter_start, date_filter_end = None, None
    annee = annee or datetime.now().year  # Défaut: année courante

    if periode == 'year':
        date_filter_start = date(annee, 1, 1)
        date_filter_end = date(annee, 12, 31)
    elif periode == 'quarter' and trimestre:
        mois_debut = (trimestre - 1) * 3 + 1
        date_filter_start = date(annee, mois_debut, 1)
        mois_fin = mois_debut + 2
        if mois_fin == 12:
            date_filter_end = date(annee, 12, 31)
        else:
            date_filter_end = date(annee, mois_fin + 1, 1) - timedelta(days=1)
    elif periode == 'month' and mois:
        date_filter_start = date(annee, mois, 1)
        if mois == 12:
            date_filter_end = date(annee, 12, 31)
        else:
            date_filter_end = date(annee, mois + 1, 1) - timedelta(days=1)
    elif periode == 'custom' and date_debut_str and date_fin_str:
        try:
            date_filter_start = datetime.strptime(date_debut_str, '%Y-%m-%d').date()
            date_filter_end = datetime.strptime(date_fin_str, '%Y-%m-%d').date()
        except ValueError:
            pass  # Ignore invalid date format

    return {
        'periode': periode,
        'annee': annee,
        'trimestre': trimestre,
        'mois': mois,
        'date_debut_str': date_debut_str,
        'date_fin_str': date_fin_str,
        'categorie_id': categorie_id,
        'ligne_budget_id': ligne_budget_id,
        'date_filter_start': date_filter_start,
        'date_filter_end': date_filter_end
    }


def calculate_rapport_data(projet, filters):
    """Calculate report data with filters applied"""
    categories = CategorieBudget.query.order_by(CategorieBudget.ordre).all()

    # Filtrer les catégories si une catégorie spécifique est demandée
    if filters['categorie_id']:
        categories = [c for c in categories if c.id == filters['categorie_id']]

    rapport = []
    total_prevu = 0
    total_realise = 0

    for cat in categories:
        lignes_cat = [l for l in projet.lignes_budget if l.categorie_id == cat.id]

        # Si filtre ligne spécifique, ne garder que cette ligne
        if filters['ligne_budget_id']:
            lignes_cat = [l for l in lignes_cat if l.id == filters['ligne_budget_id']]

        if not lignes_cat:
            continue

        cat_data = {
            'categorie': cat,
            'lignes': [],
            'total_prevu': 0,
            'total_realise': 0
        }

        for ligne in lignes_cat:
            # Requête réalisé avec jointure PieceComptable pour filtrer par date
            query = db.session.query(
                db.func.sum(LigneEcriture.debit)
            ).join(CompteComptable).join(
                PieceComptable, LigneEcriture.piece_id == PieceComptable.id
            ).filter(
                (LigneEcriture.ligne_budget_id == ligne.id) &
                (CompteComptable.classe == 6)
            )

            # Appliquer filtre de date si spécifié
            if filters['date_filter_start'] and filters['date_filter_end']:
                query = query.filter(
                    PieceComptable.date_piece >= filters['date_filter_start'],
                    PieceComptable.date_piece <= filters['date_filter_end']
                )

            realise = query.scalar() or 0
            realise = float(realise)

            # Si filtre par année et BudgetAnnee existe, utiliser le montant annuel
            prevu = float(ligne.montant_prevu or 0)
            if filters['periode'] == 'year' and filters['annee']:
                budget_annee = ligne.get_montant_annee(filters['annee'])
                if budget_annee > 0:
                    prevu = float(budget_annee)

            cat_data['lignes'].append({
                'ligne': ligne,
                'prevu': prevu,
                'realise': realise,
                'ecart': prevu - realise,
                'taux': (realise / prevu * 100) if prevu > 0 else 0
            })
            cat_data['total_prevu'] += prevu
            cat_data['total_realise'] += realise

        total_prevu += cat_data['total_prevu']
        total_realise += cat_data['total_realise']
        rapport.append(cat_data)

    return rapport, total_prevu, total_realise


@app.route('/rapports/projet/<int:id>')
@login_required
def rapport_projet(id):
    """Rapport bailleur - Budget vs Réalisé avec filtres"""
    projet = Projet.query.get_or_404(id)

    # Parse filters
    filters = parse_report_filters()

    # Calculate report data
    rapport, total_prevu, total_realise = calculate_rapport_data(projet, filters)

    # Get all categories for filter dropdown
    categories_all = CategorieBudget.query.order_by(CategorieBudget.ordre).all()

    # Calculate available years for filter
    current_year = datetime.now().year
    annee_debut = projet.date_debut.year if projet.date_debut else current_year - 2
    annee_fin = projet.date_fin.year if projet.date_fin else current_year + 2
    annees_disponibles = list(range(annee_debut, annee_fin + 1))

    # Generate labels for active filters display
    periode_label = get_periode_label(
        filters['periode'], filters['annee'], filters['trimestre'],
        filters['mois'], filters['date_filter_start'], filters['date_filter_end']
    )
    categorie_nom = get_categorie_nom(filters['categorie_id'])
    ligne_nom = get_ligne_nom(filters['ligne_budget_id'])

    return render_template('rapports/projet.html',
                          projet=projet,
                          rapport=rapport,
                          total_prevu=total_prevu,
                          total_realise=total_realise,
                          # Filter options
                          categories_all=categories_all,
                          annees_disponibles=annees_disponibles,
                          mois_noms=MOIS_NOMS,
                          # Current filter values
                          periode=filters['periode'],
                          annee=filters['annee'],
                          trimestre=filters['trimestre'],
                          mois=filters['mois'],
                          date_debut=filters['date_debut_str'] or '',
                          date_fin=filters['date_fin_str'] or '',
                          categorie_id=filters['categorie_id'],
                          ligne_budget_id=filters['ligne_budget_id'],
                          # Labels for display
                          periode_label=periode_label,
                          categorie_nom=categorie_nom,
                          ligne_nom=ligne_nom)


@app.route('/rapports/projet/<int:id>/pdf')
@login_required
def export_projet_pdf(id):
    """Export PDF du rapport bailleur avec filtres"""
    try:
        from xhtml2pdf import pisa
        from io import BytesIO
    except ImportError:
        flash("xhtml2pdf n'est pas installé. Utilisez: pip install xhtml2pdf", "danger")
        return redirect(url_for('rapport_projet', id=id))

    projet = Projet.query.get_or_404(id)

    # Parse filters (same as rapport_projet)
    filters = parse_report_filters()

    # Calculate report data with filters
    rapport, total_prevu, total_realise = calculate_rapport_data(projet, filters)

    # Generate labels for filter display in PDF
    periode_label = get_periode_label(
        filters['periode'], filters['annee'], filters['trimestre'],
        filters['mois'], filters['date_filter_start'], filters['date_filter_end']
    )
    categorie_nom = get_categorie_nom(filters['categorie_id'])
    ligne_nom = get_ligne_nom(filters['ligne_budget_id'])

    html = render_template('rapports/projet_pdf.html',
                          projet=projet,
                          rapport=rapport,
                          total_prevu=total_prevu,
                          total_realise=total_realise,
                          date_generation=datetime.now(),
                          # Filter info for display
                          periode=filters['periode'],
                          periode_label=periode_label,
                          categorie_nom=categorie_nom,
                          ligne_nom=ligne_nom)

    try:
        pdf_buffer = BytesIO()
        pisa_status = pisa.CreatePDF(html, dest=pdf_buffer)
        if pisa_status.err:
            flash("Erreur lors de la génération PDF.", "danger")
            return redirect(url_for('rapport_projet', id=id))
        pdf_buffer.seek(0)
    except Exception as e:
        flash(f"Erreur lors de la génération PDF: {str(e)}", "danger")
        return redirect(url_for('rapport_projet', id=id))

    # Generate filename with filter info
    filename_suffix = ''
    if filters['periode'] == 'year' and filters['annee']:
        filename_suffix = f"_{filters['annee']}"
    elif filters['periode'] == 'quarter' and filters['annee'] and filters['trimestre']:
        filename_suffix = f"_{filters['annee']}-T{filters['trimestre']}"
    elif filters['periode'] == 'month' and filters['annee'] and filters['mois']:
        filename_suffix = f"_{filters['annee']}-{filters['mois']:02d}"

    return Response(pdf_buffer.getvalue(),
                   mimetype='application/pdf',
                   headers={'Content-Disposition': f'attachment; filename=rapport_{projet.code}{filename_suffix}_{date.today()}.pdf'})


@app.route('/rapports/projet/<int:id>/excel')
@login_required
def export_projet_excel(id):
    """Export Excel du rapport bailleur avec filtres"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from io import BytesIO
    except ImportError:
        flash("openpyxl n'est pas installé. Utilisez: pip install openpyxl", "danger")
        return redirect(url_for('rapport_projet', id=id))

    projet = Projet.query.get_or_404(id)

    # Parse filters (same as rapport_projet)
    filters = parse_report_filters()

    # Calculate report data with filters
    rapport, total_prevu, total_realise = calculate_rapport_data(projet, filters)

    # Generate labels for filter display
    periode_label = get_periode_label(
        filters['periode'], filters['annee'], filters['trimestre'],
        filters['mois'], filters['date_filter_start'], filters['date_filter_end']
    )
    categorie_nom = get_categorie_nom(filters['categorie_id'])
    ligne_nom = get_ligne_nom(filters['ligne_budget_id'])

    wb = Workbook()
    ws = wb.active
    ws.title = "Rapport Budget"

    # Styles
    header_font = Font(bold=True, size=12)
    title_font = Font(bold=True, size=14)
    cat_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    info_fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # En-tête
    ws.merge_cells('A1:F1')
    ws['A1'] = f"RAPPORT BAILLEUR - {projet.code}"
    ws['A1'].font = title_font

    ws['A2'] = f"Projet: {projet.nom}"
    ws['A3'] = f"Bailleur: {projet.bailleur.nom if projet.bailleur else 'N/A'}"
    ws['A4'] = f"Date: {date.today().strftime('%d/%m/%Y')}"

    # Afficher les filtres actifs
    row = 5
    if filters['periode'] != 'all' or filters['categorie_id'] or filters['ligne_budget_id']:
        ws['A5'] = "Filtres appliqués:"
        ws['A5'].font = Font(bold=True)
        ws['A5'].fill = info_fill

        filter_parts = []
        if periode_label and periode_label != "Toute la période":
            filter_parts.append(f"Période: {periode_label}")
        if categorie_nom:
            filter_parts.append(f"Section: {categorie_nom}")
        if ligne_nom:
            filter_parts.append(f"Ligne: {ligne_nom}")

        ws['B5'] = " | ".join(filter_parts)
        ws['B5'].fill = info_fill
        for col in range(1, 7):
            ws.cell(row=5, column=col).fill = info_fill
        row = 7
    else:
        row = 6

    # En-têtes tableau
    headers = ['Code', 'Description', 'Budget prévu', 'Réalisé', 'Écart', 'Taux (%)']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = header_font
        cell.border = thin_border

    row += 1

    # Données du rapport (déjà filtrées)
    for cat_data in rapport:
        cat = cat_data['categorie']

        # Ligne catégorie
        ws.cell(row=row, column=1, value=cat.nom).font = Font(bold=True)
        ws.cell(row=row, column=1).fill = cat_fill
        for col in range(1, 7):
            ws.cell(row=row, column=col).fill = cat_fill
            ws.cell(row=row, column=col).border = thin_border
        row += 1

        for item in cat_data['lignes']:
            ligne = item['ligne']
            prevu = item['prevu']
            realise = item['realise']
            taux = item['taux'] / 100  # Convert to decimal for Excel percentage format

            ws.cell(row=row, column=1, value=ligne.code).border = thin_border
            ws.cell(row=row, column=2, value=ligne.intitule).border = thin_border
            ws.cell(row=row, column=3, value=prevu).border = thin_border
            ws.cell(row=row, column=3).number_format = '#,##0'
            ws.cell(row=row, column=4, value=realise).border = thin_border
            ws.cell(row=row, column=4).number_format = '#,##0'
            ws.cell(row=row, column=5, value=prevu - realise).border = thin_border
            ws.cell(row=row, column=5).number_format = '#,##0'
            ws.cell(row=row, column=6, value=taux).border = thin_border
            ws.cell(row=row, column=6).number_format = '0.0%'

            row += 1

    # Total général
    row += 1
    ws.cell(row=row, column=1, value="TOTAL GÉNÉRAL").font = Font(bold=True)
    ws.cell(row=row, column=3, value=total_prevu).font = Font(bold=True)
    ws.cell(row=row, column=3).number_format = '#,##0'
    ws.cell(row=row, column=4, value=total_realise).font = Font(bold=True)
    ws.cell(row=row, column=4).number_format = '#,##0'
    ws.cell(row=row, column=5, value=total_prevu - total_realise).font = Font(bold=True)
    ws.cell(row=row, column=5).number_format = '#,##0'
    if total_prevu > 0:
        ws.cell(row=row, column=6, value=total_realise / total_prevu).font = Font(bold=True)
        ws.cell(row=row, column=6).number_format = '0.0%'

    # Ajuster largeur colonnes
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 12

    # Ajouter une feuille "Critères" avec les détails des filtres
    if filters['periode'] != 'all' or filters['categorie_id'] or filters['ligne_budget_id']:
        ws_criteres = wb.create_sheet(title="Critères")
        ws_criteres['A1'] = "Critères de filtrage du rapport"
        ws_criteres['A1'].font = title_font

        ws_criteres['A3'] = "Projet:"
        ws_criteres['B3'] = projet.nom
        ws_criteres['A4'] = "Code:"
        ws_criteres['B4'] = projet.code
        ws_criteres['A5'] = "Période du rapport:"
        ws_criteres['B5'] = periode_label if periode_label else "Toute la période"
        ws_criteres['A6'] = "Section budgétaire:"
        ws_criteres['B6'] = categorie_nom if categorie_nom else "Toutes les sections"
        ws_criteres['A7'] = "Ligne budgétaire:"
        ws_criteres['B7'] = ligne_nom if ligne_nom else "Toutes les lignes"
        ws_criteres['A8'] = "Date de génération:"
        ws_criteres['B8'] = datetime.now().strftime('%d/%m/%Y %H:%M')

        ws_criteres.column_dimensions['A'].width = 25
        ws_criteres.column_dimensions['B'].width = 50

    # Sauvegarder
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    # Generate filename with filter info
    filename_suffix = ''
    if filters['periode'] == 'year' and filters['annee']:
        filename_suffix = f"_{filters['annee']}"
    elif filters['periode'] == 'quarter' and filters['annee'] and filters['trimestre']:
        filename_suffix = f"_{filters['annee']}-T{filters['trimestre']}"
    elif filters['periode'] == 'month' and filters['annee'] and filters['mois']:
        filename_suffix = f"_{filters['annee']}-{filters['mois']:02d}"

    return Response(output.read(),
                   mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                   headers={'Content-Disposition': f'attachment; filename=rapport_{projet.code}{filename_suffix}_{date.today()}.xlsx'})


@app.route('/rapports/reconciliation')
@login_required
def reconciliation_analytique():
    """Réconciliation Analytique ↔ Comptabilité Générale"""
    exercice_id = request.args.get('exercice_id')
    exercices = ExerciceComptable.query.order_by(ExerciceComptable.annee.desc()).all()

    if not exercice_id and exercices:
        exercice_id = exercices[0].id

    # Total charges classe 6 (comptabilité générale)
    total_compta_generale = db.session.query(
        db.func.sum(LigneEcriture.debit)
    ).join(CompteComptable).join(PieceComptable).filter(
        CompteComptable.classe == 6
    )
    if exercice_id:
        total_compta_generale = total_compta_generale.filter(
            PieceComptable.exercice_id == exercice_id
        )
    total_compta_generale = float(total_compta_generale.scalar() or 0)

    # Total par projet (analytique)
    projets_data = []
    total_analytique = 0

    for projet in Projet.query.all():
        projet_total = 0
        for ligne in projet.lignes_budget:
            realise = db.session.query(
                db.func.sum(LigneEcriture.debit)
            ).join(CompteComptable).filter(
                (LigneEcriture.ligne_budget_id == ligne.id) &
                (CompteComptable.classe == 6)
            ).scalar() or 0
            projet_total += float(realise)

        if projet_total > 0:
            projets_data.append({
                'projet': projet,
                'total': projet_total
            })
            total_analytique += projet_total

    # Écart de réconciliation
    ecart = total_compta_generale - total_analytique

    # Charges non imputées (sans ligne_budget_id)
    charges_non_imputees = db.session.query(
        db.func.sum(LigneEcriture.debit)
    ).join(CompteComptable).filter(
        CompteComptable.classe == 6,
        LigneEcriture.ligne_budget_id == None
    ).scalar() or 0

    return render_template('rapports/reconciliation.html',
                          exercices=exercices,
                          exercice_id=exercice_id,
                          total_compta_generale=total_compta_generale,
                          total_analytique=total_analytique,
                          projets_data=projets_data,
                          ecart=ecart,
                          charges_non_imputees=float(charges_non_imputees))


@app.route('/rapports/etats-financiers')
@login_required
def etats_financiers():
    """États financiers SYSCOHADA"""
    exercice_id = request.args.get('exercice_id')
    exercices = ExerciceComptable.query.order_by(ExerciceComptable.annee.desc()).all()

    if not exercice_id and exercices:
        exercice_id = exercices[0].id

    # Calculer Actif (classes 2-5)
    actif = calculer_soldes_classe([2, 3, 4, 5], exercice_id, 'actif')

    # Calculer Passif (classes 1, 4)
    passif = calculer_soldes_classe([1, 4], exercice_id, 'passif')

    # Calculer Charges (classe 6)
    charges = calculer_soldes_classe([6], exercice_id)

    # Calculer Produits (classe 7)
    produits = calculer_soldes_classe([7], exercice_id)

    resultat = sum(p['solde'] for p in produits) - sum(c['solde'] for c in charges)

    return render_template('rapports/etats_financiers.html',
                         actif=actif,
                         passif=passif,
                         charges=charges,
                         produits=produits,
                         resultat=resultat,
                         exercices=exercices,
                         exercice_id=exercice_id)


def calculer_soldes_classe(classes, exercice_id=None, type_solde=None, inclure_non_validees=False):
    """Calculer les soldes pour une classe de comptes
    SECURITY: Par défaut, n'inclut que les écritures validées
    """
    query = db.session.query(
        CompteComptable.numero,
        CompteComptable.intitule,
        db.func.sum(LigneEcriture.debit).label('total_debit'),
        db.func.sum(LigneEcriture.credit).label('total_credit')
    ).join(
        LigneEcriture, LigneEcriture.compte_id == CompteComptable.id
    ).join(
        PieceComptable, PieceComptable.id == LigneEcriture.piece_id
    ).filter(
        CompteComptable.classe.in_(classes)
    )

    # SECURITY: Par défaut, n'inclure que les écritures validées pour états financiers
    if not inclure_non_validees:
        query = query.filter(PieceComptable.valide == True)

    if exercice_id:
        query = query.filter(PieceComptable.exercice_id == exercice_id)

    query = query.group_by(CompteComptable.id).order_by(CompteComptable.numero)

    resultats = []
    for row in query.all():
        debit = float(row.total_debit or 0)
        credit = float(row.total_credit or 0)

        if type_solde == 'actif':
            solde = debit - credit
        elif type_solde == 'passif':
            solde = credit - debit
        else:
            solde = debit - credit

        if abs(solde) > 0.01:
            resultats.append({
                'numero': row.numero,
                'intitule': row.intitule,
                'solde': solde
            })

    return resultats


# =============================================================================
# ROUTES - SAUVEGARDES ET BACKUPS
# =============================================================================

# Configuration des backups
BACKUP_FOLDER = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'backups')
BACKUP_DAILY_FOLDER = os.path.join(BACKUP_FOLDER, 'daily')
BACKUP_WEEKLY_FOLDER = os.path.join(BACKUP_FOLDER, 'weekly')
BACKUP_MANUAL_FOLDER = os.path.join(BACKUP_FOLDER, 'manual')

# Créer les dossiers de backup s'ils n'existent pas
for folder in [BACKUP_FOLDER, BACKUP_DAILY_FOLDER, BACKUP_WEEKLY_FOLDER, BACKUP_MANUAL_FOLDER]:
    os.makedirs(folder, exist_ok=True)


def get_db_path():
    """Obtenir le chemin de la base de données SQLite"""
    db_uri = app.config['SQLALCHEMY_DATABASE_URI']
    if db_uri.startswith('sqlite:///'):
        db_file = db_uri.replace('sqlite:///', '')
        # Si chemin relatif, le rendre absolu par rapport au dossier instance
        if not os.path.isabs(db_file):
            db_file = os.path.join(app.instance_path, db_file)
        return db_file
    return None


def create_backup(backup_type='manual'):
    """Créer une sauvegarde de la base de données"""
    db_path = get_db_path()
    if not db_path:
        return None, "Backup uniquement disponible pour SQLite"

    if not os.path.exists(db_path):
        return None, f"Base de données non trouvée: {db_path}"

    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')

    if backup_type == 'daily':
        backup_dir = BACKUP_DAILY_FOLDER
        filename = f"backup_{date.today().isoformat()}.db"
    elif backup_type == 'weekly':
        backup_dir = BACKUP_WEEKLY_FOLDER
        week_num = date.today().isocalendar()[1]
        filename = f"backup_week_{week_num:02d}.db"
    else:
        backup_dir = BACKUP_MANUAL_FOLDER
        filename = f"backup_{timestamp}.db"

    backup_path = os.path.join(backup_dir, filename)

    try:
        shutil.copy2(db_path, backup_path)
        size = os.path.getsize(backup_path)
        return {
            'filename': filename,
            'path': backup_path,
            'size': size,
            'created': datetime.now(),
            'type': backup_type
        }, None
    except Exception as e:
        return None, str(e)


def list_backups():
    """Lister toutes les sauvegardes disponibles"""
    backups = []

    for backup_type, folder in [('manual', BACKUP_MANUAL_FOLDER),
                                 ('daily', BACKUP_DAILY_FOLDER),
                                 ('weekly', BACKUP_WEEKLY_FOLDER)]:
        if os.path.exists(folder):
            for filename in os.listdir(folder):
                if filename.endswith('.db'):
                    filepath = os.path.join(folder, filename)
                    stat = os.stat(filepath)
                    backups.append({
                        'filename': filename,
                        'path': filepath,
                        'type': backup_type,
                        'size': stat.st_size,
                        'created': datetime.fromtimestamp(stat.st_mtime)
                    })

    # Trier par date décroissante
    backups.sort(key=lambda x: x['created'], reverse=True)
    return backups


def cleanup_old_backups():
    """Nettoyer les anciennes sauvegardes (rotation)"""
    # Garder 7 derniers jours de backups daily
    daily_backups = sorted(glob_module.glob(os.path.join(BACKUP_DAILY_FOLDER, '*.db')),
                          key=os.path.getmtime, reverse=True)
    for old_backup in daily_backups[7:]:
        os.remove(old_backup)

    # Garder 4 dernières semaines de backups weekly
    weekly_backups = sorted(glob_module.glob(os.path.join(BACKUP_WEEKLY_FOLDER, '*.db')),
                           key=os.path.getmtime, reverse=True)
    for old_backup in weekly_backups[4:]:
        os.remove(old_backup)


def envoyer_backup_email(backup_path, destinataire=None):
    """Envoyer un backup par email"""
    config = ConfigBackup.query.filter_by(type_destination='email', actif=True).first()
    if not config:
        return False, "Configuration email non définie ou inactive"

    if not config.smtp_server or not config.smtp_user:
        return False, "Configuration SMTP incomplète"

    dest = destinataire or config.email_destinataire
    if not dest:
        return False, "Aucun destinataire défini"

    try:
        msg = MIMEMultipart()
        msg['From'] = config.smtp_user
        msg['To'] = dest
        msg['Subject'] = f'CREATES - Sauvegarde du {date.today().strftime("%d/%m/%Y")}'

        # Corps du message
        file_size = os.path.getsize(backup_path) / 1024 / 1024
        body = f"""
Sauvegarde automatique CREATES

Date: {datetime.now().strftime('%d/%m/%Y %H:%M')}
Fichier: {os.path.basename(backup_path)}
Taille: {file_size:.2f} Mo

Ce message a été envoyé automatiquement par le système de sauvegarde CREATES.
Ne pas répondre à cet email.

--
GIE CREATES
Système de comptabilité
        """
        msg.attach(MIMEText(body, 'plain', 'utf-8'))

        # Pièce jointe
        with open(backup_path, 'rb') as f:
            part = MIMEBase('application', 'octet-stream')
            part.set_payload(f.read())
            encoders.encode_base64(part)
            part.add_header('Content-Disposition',
                          f'attachment; filename={os.path.basename(backup_path)}')
            msg.attach(part)

        # Envoi
        server = smtplib.SMTP(config.smtp_server, config.smtp_port, timeout=30)
        server.starttls()
        server.login(config.smtp_user, config.smtp_password)
        server.send_message(msg)
        server.quit()

        return True, "Email envoyé avec succès"
    except smtplib.SMTPAuthenticationError:
        return False, "Erreur d'authentification SMTP. Vérifiez les identifiants."
    except smtplib.SMTPException as e:
        return False, f"Erreur SMTP: {str(e)}"
    except Exception as e:
        return False, f"Erreur: {str(e)}"


@app.route('/admin/backups')
@login_required
@role_required(['directeur'])
def liste_backups():
    """Page de gestion des sauvegardes"""
    backups = list_backups()
    db_path = get_db_path()
    db_size = os.path.getsize(db_path) if db_path and os.path.exists(db_path) else 0
    email_config = ConfigBackup.query.filter_by(type_destination='email').first()

    return render_template('admin/backups.html',
                           backups=backups,
                           db_size=db_size,
                           email_config=email_config)


@app.route('/admin/backups/creer', methods=['POST'])
@login_required
@role_required(['directeur'])
def creer_backup():
    """Créer une nouvelle sauvegarde manuelle"""
    backup, error = create_backup('manual')

    if error:
        flash(f'Erreur lors de la sauvegarde: {error}', 'danger')
    else:
        log_audit('backup', None, 'CREATE', new_values={'filename': backup['filename']})
        flash(f'Sauvegarde créée: {backup["filename"]}', 'success')

    return redirect(url_for('liste_backups'))


@app.route('/admin/backups/telecharger/<path:filename>')
@login_required
@role_required(['directeur'])
def telecharger_backup(filename):
    """Télécharger un fichier de sauvegarde"""
    # Sécurité: vérifier que le fichier est dans un dossier autorisé
    for folder in [BACKUP_MANUAL_FOLDER, BACKUP_DAILY_FOLDER, BACKUP_WEEKLY_FOLDER]:
        filepath = os.path.join(folder, os.path.basename(filename))
        if os.path.exists(filepath):
            log_audit('backup', None, 'DOWNLOAD', new_values={'filename': filename})
            return send_file(filepath, as_attachment=True,
                           download_name=f'creates_{filename}')

    flash('Fichier de sauvegarde non trouvé', 'danger')
    return redirect(url_for('liste_backups'))


@app.route('/admin/backups/supprimer/<path:filename>', methods=['POST'])
@login_required
@role_required(['directeur'])
def supprimer_backup(filename):
    """Supprimer un fichier de sauvegarde"""
    for folder in [BACKUP_MANUAL_FOLDER, BACKUP_DAILY_FOLDER, BACKUP_WEEKLY_FOLDER]:
        filepath = os.path.join(folder, os.path.basename(filename))
        if os.path.exists(filepath):
            os.remove(filepath)
            log_audit('backup', None, 'DELETE', new_values={'filename': filename})
            flash(f'Sauvegarde supprimée: {filename}', 'success')
            return redirect(url_for('liste_backups'))

    flash('Fichier de sauvegarde non trouvé', 'danger')
    return redirect(url_for('liste_backups'))


@app.route('/admin/backups/restaurer/<path:filename>', methods=['POST'])
@login_required
@role_required(['directeur'])
def restaurer_backup(filename):
    """Restaurer depuis une sauvegarde"""
    db_path = get_db_path()
    if not db_path:
        flash('Restauration uniquement disponible pour SQLite', 'danger')
        return redirect(url_for('liste_backups'))

    # Trouver le fichier backup
    backup_path = None
    for folder in [BACKUP_MANUAL_FOLDER, BACKUP_DAILY_FOLDER, BACKUP_WEEKLY_FOLDER]:
        filepath = os.path.join(folder, os.path.basename(filename))
        if os.path.exists(filepath):
            backup_path = filepath
            break

    if not backup_path:
        flash('Fichier de sauvegarde non trouvé', 'danger')
        return redirect(url_for('liste_backups'))

    try:
        # Créer une sauvegarde avant restauration
        pre_restore_backup, _ = create_backup('manual')

        # Restaurer
        shutil.copy2(backup_path, db_path)

        log_audit('backup', None, 'RESTORE', new_values={
            'restored_from': filename,
            'pre_restore_backup': pre_restore_backup['filename'] if pre_restore_backup else None
        })

        flash(f'Base de données restaurée depuis {filename}. '
              f'Sauvegarde pré-restauration: {pre_restore_backup["filename"] if pre_restore_backup else "N/A"}',
              'success')
    except Exception as e:
        flash(f'Erreur lors de la restauration: {str(e)}', 'danger')

    return redirect(url_for('liste_backups'))


@app.route('/admin/backups/config', methods=['GET', 'POST'])
@login_required
@role_required(['directeur'])
def config_backup():
    """Configurer les destinations de backup (email)"""
    config = ConfigBackup.query.filter_by(type_destination='email').first()

    if request.method == 'POST':
        if not config:
            config = ConfigBackup(type_destination='email')
            db.session.add(config)

        config.smtp_server = request.form.get('smtp_server', '').strip()
        config.smtp_port = int(request.form.get('smtp_port', 587))
        config.smtp_user = request.form.get('smtp_user', '').strip()

        # Ne mettre à jour le mot de passe que s'il est fourni
        new_password = request.form.get('smtp_password', '')
        if new_password:
            config.smtp_password = new_password

        config.email_destinataire = request.form.get('email_destinataire', '').strip()
        config.actif = request.form.get('actif') == 'on'

        db.session.commit()
        log_audit('config_backup', config.id, 'UPDATE', new_values={
            'smtp_server': config.smtp_server,
            'smtp_user': config.smtp_user,
            'actif': config.actif
        })
        flash('Configuration email sauvegardée', 'success')
        return redirect(url_for('liste_backups'))

    return render_template('admin/backup_config.html', config=config)


@app.route('/admin/backups/test-email', methods=['POST'])
@login_required
@role_required(['directeur'])
def test_email_backup():
    """Tester l'envoi d'email de backup"""
    config = ConfigBackup.query.filter_by(type_destination='email').first()

    if not config or not config.smtp_server:
        flash('Veuillez d\'abord configurer les paramètres SMTP', 'warning')
        return redirect(url_for('config_backup'))

    # Créer un petit fichier test
    test_file = os.path.join(BACKUP_MANUAL_FOLDER, 'test_email.txt')
    try:
        with open(test_file, 'w', encoding='utf-8') as f:
            f.write(f'Test CREATES backup - {datetime.now().strftime("%d/%m/%Y %H:%M:%S")}\n')
            f.write('Ce fichier est un test de la configuration email.\n')
            f.write('Si vous recevez cet email, la configuration est correcte.')

        # Activer temporairement la config pour le test
        was_active = config.actif
        config.actif = True

        success, message = envoyer_backup_email(test_file)

        # Restaurer l'état précédent
        config.actif = was_active

        if success:
            flash('Email de test envoyé avec succès! Vérifiez votre boîte de réception.', 'success')
            log_audit('config_backup', config.id, 'TEST_EMAIL', new_values={'success': True})
        else:
            flash(f'Erreur lors de l\'envoi: {message}', 'danger')
            log_audit('config_backup', config.id, 'TEST_EMAIL', new_values={'success': False, 'error': message})
    finally:
        # Supprimer le fichier test
        if os.path.exists(test_file):
            os.remove(test_file)

    return redirect(url_for('config_backup'))


@app.route('/admin/backups/<path:filename>/envoyer-email', methods=['POST'])
@login_required
@role_required(['directeur'])
def envoyer_backup_par_email(filename):
    """Envoyer un backup spécifique par email"""
    for folder in [BACKUP_MANUAL_FOLDER, BACKUP_DAILY_FOLDER, BACKUP_WEEKLY_FOLDER]:
        filepath = os.path.join(folder, os.path.basename(filename))
        if os.path.exists(filepath):
            success, message = envoyer_backup_email(filepath)
            if success:
                log_audit('backup', None, 'EMAIL_SENT', new_values={'filename': filename})
                flash(f'Backup "{filename}" envoyé par email', 'success')
            else:
                flash(f'Erreur envoi email: {message}', 'danger')
            return redirect(url_for('liste_backups'))

    flash('Fichier de sauvegarde non trouvé', 'danger')
    return redirect(url_for('liste_backups'))


@app.route('/admin/backups/export-excel')
@login_required
@role_required(['directeur', 'comptable'])
def export_donnees_excel():
    """Exporter toutes les données comptables en Excel"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        flash('Module openpyxl non installé', 'danger')
        return redirect(url_for('liste_backups'))

    wb = Workbook()

    # Style
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='7D8B6A', end_color='7D8B6A', fill_type='solid')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    def style_header(ws):
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border

    # === Feuille 1: Écritures ===
    ws1 = wb.active
    ws1.title = "Ecritures"
    ws1.append(['Numero', 'Date', 'Journal', 'Libelle', 'Reference', 'Total Debit', 'Total Credit', 'Valide', 'Exercice'])

    ecritures = PieceComptable.query.order_by(PieceComptable.date_piece.desc()).all()
    for e in ecritures:
        ws1.append([
            e.numero,
            e.date_piece.strftime('%d/%m/%Y') if e.date_piece else '',
            e.journal.code if e.journal else '',
            e.libelle,
            e.reference or '',
            float(e.total_debit),
            float(e.total_credit),
            'Oui' if e.valide else 'Non',
            e.exercice.annee if e.exercice else ''
        ])
    style_header(ws1)

    # === Feuille 2: Lignes d'écritures ===
    ws2 = wb.create_sheet("Lignes Ecritures")
    ws2.append(['Piece', 'Date', 'Compte', 'Intitule Compte', 'Libelle', 'Projet', 'Debit', 'Credit'])

    lignes = LigneEcriture.query.join(PieceComptable).order_by(PieceComptable.date_piece.desc()).all()
    for l in lignes:
        ws2.append([
            l.piece.numero if l.piece else '',
            l.piece.date_piece.strftime('%d/%m/%Y') if l.piece and l.piece.date_piece else '',
            l.compte.numero if l.compte else '',
            l.compte.intitule if l.compte else '',
            l.libelle or '',
            l.projet.code if l.projet else '',
            float(l.debit) if l.debit else 0,
            float(l.credit) if l.credit else 0
        ])
    style_header(ws2)

    # === Feuille 3: Projets ===
    ws3 = wb.create_sheet("Projets")
    ws3.append(['Code', 'Nom', 'Bailleur', 'Date Debut', 'Date Fin', 'Budget Total', 'Statut'])

    projets = Projet.query.all()
    for p in projets:
        ws3.append([
            p.code,
            p.nom,
            p.bailleur.nom if p.bailleur else '',
            p.date_debut.strftime('%d/%m/%Y') if p.date_debut else '',
            p.date_fin.strftime('%d/%m/%Y') if p.date_fin else '',
            float(p.budget_total) if p.budget_total else 0,
            p.statut
        ])
    style_header(ws3)

    # === Feuille 4: Lignes budgétaires ===
    ws4 = wb.create_sheet("Lignes Budget")
    ws4.append(['Projet', 'Code', 'Intitule', 'Categorie', 'Montant Prevu'])

    lignes_budget = LigneBudget.query.all()
    for lb in lignes_budget:
        ws4.append([
            lb.projet.code if lb.projet else '',
            lb.code,
            lb.intitule,
            lb.categorie.nom if lb.categorie else '',
            float(lb.montant_prevu) if lb.montant_prevu else 0
        ])
    style_header(ws4)

    # === Feuille 5: Bailleurs ===
    ws5 = wb.create_sheet("Bailleurs")
    ws5.append(['Code', 'Nom', 'Pays', 'Contact', 'Email'])

    bailleurs = Bailleur.query.all()
    for b in bailleurs:
        ws5.append([b.code, b.nom, b.pays or '', b.contact or '', b.email or ''])
    style_header(ws5)

    # === Feuille 6: Fournisseurs ===
    ws6 = wb.create_sheet("Fournisseurs")
    ws6.append(['Code', 'Nom', 'Categorie', 'NINEA', 'Telephone', 'Email', 'Actif'])

    fournisseurs = Fournisseur.query.all()
    for f in fournisseurs:
        ws6.append([
            f.code, f.nom, f.categorie or '', f.ninea or '',
            f.telephone or '', f.email or '', 'Oui' if f.actif else 'Non'
        ])
    style_header(ws6)

    # === Feuille 7: Plan comptable ===
    ws7 = wb.create_sheet("Plan Comptable")
    ws7.append(['Numero', 'Intitule', 'Classe', 'Type'])

    comptes = CompteComptable.query.order_by(CompteComptable.numero).all()
    for c in comptes:
        ws7.append([c.numero, c.intitule, c.classe, c.type_compte or ''])
    style_header(ws7)

    # === Feuille 8: Balance ===
    ws8 = wb.create_sheet("Balance")
    ws8.append(['Compte', 'Intitule', 'Total Debit', 'Total Credit', 'Solde Debiteur', 'Solde Crediteur'])

    for c in comptes:
        total_debit = db.session.query(db.func.sum(LigneEcriture.debit)).filter(
            LigneEcriture.compte_id == c.id).scalar() or 0
        total_credit = db.session.query(db.func.sum(LigneEcriture.credit)).filter(
            LigneEcriture.compte_id == c.id).scalar() or 0
        solde = float(total_debit) - float(total_credit)
        if total_debit or total_credit:
            ws8.append([
                c.numero, c.intitule,
                float(total_debit), float(total_credit),
                solde if solde > 0 else 0,
                -solde if solde < 0 else 0
            ])
    style_header(ws8)

    # Sauvegarder dans un buffer
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    log_audit('export', None, 'EXPORT_EXCEL', new_values={'type': 'full_backup'})

    filename = f'CREATES_Export_Complet_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    return send_file(output, as_attachment=True, download_name=filename,
                    mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


# =============================================================================
# ROUTES - RECHERCHE GLOBALE
# =============================================================================

@app.route('/recherche')
@login_required
def recherche_globale():
    """Recherche globale dans l'application"""
    q = request.args.get('q', '').strip()

    if not q or len(q) < 2:
        return render_template('recherche/resultats.html', q=q, resultats=None)

    resultats = {
        'ecritures': [],
        'projets': [],
        'bailleurs': [],
        'fournisseurs': [],
        'comptes': []
    }

    # Recherche dans les écritures
    ecritures = PieceComptable.query.filter(
        db.or_(
            PieceComptable.numero.ilike(f'%{q}%'),
            PieceComptable.libelle.ilike(f'%{q}%'),
            PieceComptable.reference.ilike(f'%{q}%')
        )
    ).order_by(PieceComptable.date_piece.desc()).limit(10).all()
    resultats['ecritures'] = ecritures

    # Recherche dans les projets
    projets = Projet.query.filter(
        db.or_(
            Projet.code.ilike(f'%{q}%'),
            Projet.nom.ilike(f'%{q}%'),
            Projet.description.ilike(f'%{q}%')
        )
    ).limit(10).all()
    resultats['projets'] = projets

    # Recherche dans les bailleurs
    bailleurs = Bailleur.query.filter(
        db.or_(
            Bailleur.code.ilike(f'%{q}%'),
            Bailleur.nom.ilike(f'%{q}%')
        )
    ).limit(10).all()
    resultats['bailleurs'] = bailleurs

    # Recherche dans les fournisseurs
    fournisseurs = Fournisseur.query.filter(
        db.or_(
            Fournisseur.code.ilike(f'%{q}%'),
            Fournisseur.nom.ilike(f'%{q}%'),
            Fournisseur.ninea.ilike(f'%{q}%')
        )
    ).limit(10).all()
    resultats['fournisseurs'] = fournisseurs

    # Recherche dans les comptes
    comptes = CompteComptable.query.filter(
        db.or_(
            CompteComptable.numero.ilike(f'%{q}%'),
            CompteComptable.intitule.ilike(f'%{q}%')
        )
    ).order_by(CompteComptable.numero).limit(10).all()
    resultats['comptes'] = comptes

    # Compter le total
    total = sum(len(v) for v in resultats.values())

    return render_template('recherche/resultats.html', q=q, resultats=resultats, total=total)


@app.route('/api/recherche')
@login_required
def api_recherche():
    """API de recherche rapide pour autocompletion"""
    q = request.args.get('q', '').strip()

    if not q or len(q) < 2:
        return jsonify([])

    suggestions = []

    # Écritures (max 3)
    ecritures = PieceComptable.query.filter(
        db.or_(
            PieceComptable.numero.ilike(f'%{q}%'),
            PieceComptable.libelle.ilike(f'%{q}%')
        )
    ).limit(3).all()
    for e in ecritures:
        suggestions.append({
            'type': 'ecriture',
            'label': f'{e.numero} - {e.libelle[:30]}',
            'url': url_for('detail_ecriture', id=e.id)
        })

    # Projets (max 3)
    projets = Projet.query.filter(
        db.or_(
            Projet.code.ilike(f'%{q}%'),
            Projet.nom.ilike(f'%{q}%')
        )
    ).limit(3).all()
    for p in projets:
        suggestions.append({
            'type': 'projet',
            'label': f'{p.code} - {p.nom[:30]}',
            'url': url_for('detail_projet', id=p.id)
        })

    # Fournisseurs (max 2)
    fournisseurs = Fournisseur.query.filter(
        db.or_(
            Fournisseur.code.ilike(f'%{q}%'),
            Fournisseur.nom.ilike(f'%{q}%')
        )
    ).limit(2).all()
    for f in fournisseurs:
        suggestions.append({
            'type': 'fournisseur',
            'label': f'{f.code} - {f.nom[:30]}',
            'url': url_for('detail_fournisseur', id=f.id)
        })

    return jsonify(suggestions)


# =============================================================================
# ROUTES - AIDE
# =============================================================================

@app.route('/aide')
@login_required
def aide_index():
    """Page d'accueil de l'aide"""
    return render_template('aide/index.html')


@app.route('/aide/guide')
@login_required
def guide_utilisateur():
    """Guide utilisateur avec tutoriels"""
    return render_template('aide/guide.html')


@app.route('/aide/manuel')
@login_required
def manuel_gestion():
    """Manuel de gestion CREATES"""
    return render_template('aide/manuel.html')


@app.route('/aide/faq')
@login_required
def faq():
    """Questions fréquentes"""
    return render_template('aide/faq.html')


@app.route('/aide/manuel/pdf')
@login_required
def telecharger_manuel():
    """Télécharger le Manuel de Gestion en PDF"""
    return send_from_directory('static/docs', 'Manuel_Gestion_CREATES.pdf', as_attachment=True)


# =============================================================================
# INITIALISATION BASE DE DONNEES
# =============================================================================

def init_db():
    """Initialiser la base de données avec les données de base"""
    db.create_all()

    # Vérifier si déjà initialisé
    if Devise.query.first():
        return

    # Devises
    devises = [
        Devise(code='XOF', nom='Franc CFA', symbole='FCFA', taux_base=1),
        Devise(code='USD', nom='Dollar US', symbole='$', taux_base=600),
        Devise(code='EUR', nom='Euro', symbole='€', taux_base=655),
        Devise(code='CHF', nom='Franc Suisse', symbole='CHF', taux_base=700),
    ]
    db.session.add_all(devises)

    # Exercice comptable
    exercice = ExerciceComptable(
        annee=2025,
        date_debut=date(2025, 1, 1),
        date_fin=date(2025, 12, 31)
    )
    db.session.add(exercice)

    # Plan comptable SYSCOHADA pour ONG - Conforme aux normes
    comptes = [
        # Classe 1 - Capitaux propres (compte 19 supprimé - non standard SYSCOHADA)
        CompteComptable(numero='10', intitule='Capital', classe=1, type_compte='passif'),
        CompteComptable(numero='101', intitule='Capital social', classe=1, type_compte='passif'),
        CompteComptable(numero='11', intitule='Réserves', classe=1, type_compte='passif'),
        CompteComptable(numero='12', intitule='Report à nouveau', classe=1, type_compte='passif'),
        CompteComptable(numero='13', intitule='Résultat de l\'exercice', classe=1, type_compte='passif'),
        CompteComptable(numero='14', intitule='Subventions d\'investissement', classe=1, type_compte='passif'),
        CompteComptable(numero='15', intitule='Provisions réglementées', classe=1, type_compte='passif'),
        CompteComptable(numero='16', intitule='Emprunts et dettes', classe=1, type_compte='passif'),
        CompteComptable(numero='17', intitule='Dettes de crédit-bail', classe=1, type_compte='passif'),
        CompteComptable(numero='18', intitule='Dettes liées à des participations', classe=1, type_compte='passif'),

        # Classe 2 - Immobilisations
        CompteComptable(numero='21', intitule='Immobilisations incorporelles', classe=2, type_compte='actif'),
        CompteComptable(numero='211', intitule='Frais de développement', classe=2, type_compte='actif'),
        CompteComptable(numero='212', intitule='Brevets, licences', classe=2, type_compte='actif'),
        CompteComptable(numero='22', intitule='Terrains', classe=2, type_compte='actif'),
        CompteComptable(numero='23', intitule='Bâtiments', classe=2, type_compte='actif'),
        CompteComptable(numero='24', intitule='Matériel et outillage', classe=2, type_compte='actif'),
        CompteComptable(numero='241', intitule='Matériel industriel', classe=2, type_compte='actif'),
        CompteComptable(numero='244', intitule='Matériel informatique', classe=2, type_compte='actif'),
        CompteComptable(numero='245', intitule='Matériel de transport', classe=2, type_compte='actif'),
        CompteComptable(numero='246', intitule='Mobilier de bureau', classe=2, type_compte='actif'),
        CompteComptable(numero='28', intitule='Amortissements', classe=2, type_compte='actif'),
        CompteComptable(numero='281', intitule='Amort. immobilisations incorporelles', classe=2, type_compte='actif'),
        CompteComptable(numero='284', intitule='Amort. matériel', classe=2, type_compte='actif'),

        # Classe 3 - Stocks
        CompteComptable(numero='31', intitule='Stocks de matières premières', classe=3, type_compte='actif'),
        CompteComptable(numero='32', intitule='Stocks fournitures', classe=3, type_compte='actif'),
        CompteComptable(numero='38', intitule='Stocks en cours de route', classe=3, type_compte='actif'),
        CompteComptable(numero='39', intitule='Dépréciations des stocks', classe=3, type_compte='actif'),

        # Classe 4 - Tiers
        CompteComptable(numero='40', intitule='Fournisseurs', classe=4, type_compte='passif'),
        CompteComptable(numero='401', intitule='Fournisseurs locaux', classe=4, type_compte='passif'),
        CompteComptable(numero='402', intitule='Fournisseurs étrangers', classe=4, type_compte='passif'),
        CompteComptable(numero='41', intitule='Clients et bailleurs', classe=4, type_compte='actif'),
        CompteComptable(numero='411', intitule='Bailleurs de fonds', classe=4, type_compte='actif'),
        # Sous-comptes par bailleur
        CompteComptable(numero='4111', intitule='Bailleur - Nitidae', classe=4, type_compte='actif'),
        CompteComptable(numero='4112', intitule='Bailleur - GIUB', classe=4, type_compte='actif'),
        CompteComptable(numero='4113', intitule='Bailleur - AFD', classe=4, type_compte='actif'),
        CompteComptable(numero='4114', intitule='Bailleur - Union Européenne', classe=4, type_compte='actif'),
        CompteComptable(numero='4119', intitule='Autres bailleurs', classe=4, type_compte='actif'),
        CompteComptable(numero='42', intitule='Personnel', classe=4, type_compte='passif'),
        CompteComptable(numero='421', intitule='Personnel - Rémunérations dues', classe=4, type_compte='passif'),
        CompteComptable(numero='422', intitule='Personnel - Avances et acomptes', classe=4, type_compte='actif'),
        CompteComptable(numero='43', intitule='Organismes sociaux', classe=4, type_compte='passif'),
        CompteComptable(numero='431', intitule='Sécurité sociale (CSS)', classe=4, type_compte='passif'),
        CompteComptable(numero='432', intitule='Caisse de retraite (IPRES)', classe=4, type_compte='passif'),
        CompteComptable(numero='433', intitule='Mutuelles santé (IPM)', classe=4, type_compte='passif'),
        CompteComptable(numero='44', intitule='État et collectivités', classe=4, type_compte='passif'),
        CompteComptable(numero='441', intitule='État - Impôt sur les bénéfices', classe=4, type_compte='passif'),
        CompteComptable(numero='442', intitule='État - TVA collectée', classe=4, type_compte='passif'),
        CompteComptable(numero='443', intitule='État - TVA déductible', classe=4, type_compte='actif'),
        CompteComptable(numero='444', intitule='État - Retenues à la source (BRS)', classe=4, type_compte='passif'),
        CompteComptable(numero='445', intitule='État - IRVM/IRCM', classe=4, type_compte='passif'),
        CompteComptable(numero='446', intitule='État - Patente et CFCE', classe=4, type_compte='passif'),
        CompteComptable(numero='47', intitule='Comptes transitoires', classe=4, type_compte='actif'),
        CompteComptable(numero='471', intitule='Débiteurs divers', classe=4, type_compte='actif'),
        CompteComptable(numero='472', intitule='Créditeurs divers', classe=4, type_compte='passif'),
        CompteComptable(numero='48', intitule='Charges/Produits constatés d\'avance', classe=4, type_compte='passif'),

        # Classe 5 - Trésorerie
        CompteComptable(numero='52', intitule='Banques', classe=5, type_compte='actif'),
        CompteComptable(numero='521', intitule='Banque compte principal', classe=5, type_compte='actif'),
        # Sous-comptes par devise
        CompteComptable(numero='5211', intitule='Banque XOF', classe=5, type_compte='actif'),
        CompteComptable(numero='5212', intitule='Banque USD', classe=5, type_compte='actif'),
        CompteComptable(numero='5213', intitule='Banque CHF', classe=5, type_compte='actif'),
        CompteComptable(numero='5214', intitule='Banque EUR', classe=5, type_compte='actif'),
        CompteComptable(numero='522', intitule='Banque compte projet', classe=5, type_compte='actif'),
        CompteComptable(numero='53', intitule='Établissements financiers', classe=5, type_compte='actif'),
        CompteComptable(numero='57', intitule='Caisse', classe=5, type_compte='actif'),
        CompteComptable(numero='571', intitule='Caisse siège', classe=5, type_compte='actif'),
        CompteComptable(numero='572', intitule='Caisse terrain', classe=5, type_compte='actif'),
        CompteComptable(numero='58', intitule='Virements internes', classe=5, type_compte='actif'),

        # Classe 6 - Charges
        CompteComptable(numero='60', intitule='Achats', classe=6, type_compte='charge'),
        CompteComptable(numero='601', intitule='Achats fournitures bureau', classe=6, type_compte='charge'),
        CompteComptable(numero='602', intitule='Achats fournitures terrain', classe=6, type_compte='charge'),
        CompteComptable(numero='603', intitule='Achats consommables', classe=6, type_compte='charge'),
        CompteComptable(numero='604', intitule='Achats matières premières', classe=6, type_compte='charge'),
        CompteComptable(numero='605', intitule='Achats équipements', classe=6, type_compte='charge'),
        CompteComptable(numero='61', intitule='Transports', classe=6, type_compte='charge'),
        CompteComptable(numero='611', intitule='Transport personnel', classe=6, type_compte='charge'),
        CompteComptable(numero='612', intitule='Transport matériel', classe=6, type_compte='charge'),
        CompteComptable(numero='613', intitule='Transport aérien', classe=6, type_compte='charge'),
        CompteComptable(numero='62', intitule='Services extérieurs', classe=6, type_compte='charge'),
        CompteComptable(numero='621', intitule='Locations immobilières', classe=6, type_compte='charge'),
        CompteComptable(numero='622', intitule='Locations matériel/véhicules', classe=6, type_compte='charge'),
        CompteComptable(numero='623', intitule='Entretien et réparations', classe=6, type_compte='charge'),
        CompteComptable(numero='624', intitule='Honoraires et consultants', classe=6, type_compte='charge'),
        CompteComptable(numero='625', intitule='Déplacements et missions', classe=6, type_compte='charge'),
        CompteComptable(numero='6251', intitule='Frais de déplacement local', classe=6, type_compte='charge'),
        CompteComptable(numero='6252', intitule='Frais de mission international', classe=6, type_compte='charge'),
        CompteComptable(numero='6253', intitule='Hébergement', classe=6, type_compte='charge'),
        CompteComptable(numero='6254', intitule='Per diem', classe=6, type_compte='charge'),
        CompteComptable(numero='626', intitule='Télécommunications', classe=6, type_compte='charge'),
        CompteComptable(numero='6261', intitule='Téléphone et internet', classe=6, type_compte='charge'),
        CompteComptable(numero='6262', intitule='Courrier et affranchissement', classe=6, type_compte='charge'),
        CompteComptable(numero='627', intitule='Services bancaires', classe=6, type_compte='charge'),
        CompteComptable(numero='628', intitule='Assurances', classe=6, type_compte='charge'),
        CompteComptable(numero='63', intitule='Autres services', classe=6, type_compte='charge'),
        CompteComptable(numero='631', intitule='Formation', classe=6, type_compte='charge'),
        CompteComptable(numero='632', intitule='Ateliers et réunions', classe=6, type_compte='charge'),
        CompteComptable(numero='633', intitule='Communication et publication', classe=6, type_compte='charge'),
        CompteComptable(numero='634', intitule='Études et recherches', classe=6, type_compte='charge'),
        CompteComptable(numero='635', intitule='Sous-traitance', classe=6, type_compte='charge'),
        # Classe 64 - Impôts et taxes (détaillé pour le Sénégal)
        CompteComptable(numero='64', intitule='Impôts et taxes', classe=6, type_compte='charge'),
        CompteComptable(numero='641', intitule='Patente', classe=6, type_compte='charge'),
        CompteComptable(numero='642', intitule='CFCE (Contribution Foncière)', classe=6, type_compte='charge'),
        CompteComptable(numero='643', intitule='Taxes sur véhicules', classe=6, type_compte='charge'),
        CompteComptable(numero='644', intitule='TVA non récupérable', classe=6, type_compte='charge'),
        CompteComptable(numero='645', intitule='Droits d\'enregistrement', classe=6, type_compte='charge'),
        CompteComptable(numero='646', intitule='Droits de douane', classe=6, type_compte='charge'),
        CompteComptable(numero='647', intitule='Autres impôts et taxes', classe=6, type_compte='charge'),
        CompteComptable(numero='65', intitule='Autres charges', classe=6, type_compte='charge'),
        CompteComptable(numero='651', intitule='Pertes sur créances', classe=6, type_compte='charge'),
        CompteComptable(numero='652', intitule='Pénalités et amendes', classe=6, type_compte='charge'),
        CompteComptable(numero='66', intitule='Charges de personnel', classe=6, type_compte='charge'),
        CompteComptable(numero='661', intitule='Salaires bruts', classe=6, type_compte='charge'),
        CompteComptable(numero='6611', intitule='Salaires personnel permanent', classe=6, type_compte='charge'),
        CompteComptable(numero='6612', intitule='Salaires personnel projet', classe=6, type_compte='charge'),
        CompteComptable(numero='662', intitule='Indemnités et primes', classe=6, type_compte='charge'),
        CompteComptable(numero='6621', intitule='Indemnité de logement', classe=6, type_compte='charge'),
        CompteComptable(numero='6622', intitule='Indemnité de transport', classe=6, type_compte='charge'),
        CompteComptable(numero='6623', intitule='Prime de rendement', classe=6, type_compte='charge'),
        CompteComptable(numero='663', intitule='Charges sociales patronales', classe=6, type_compte='charge'),
        CompteComptable(numero='6631', intitule='Cotisations CSS (employeur)', classe=6, type_compte='charge'),
        CompteComptable(numero='6632', intitule='Cotisations IPRES (employeur)', classe=6, type_compte='charge'),
        CompteComptable(numero='6633', intitule='Cotisations IPM (employeur)', classe=6, type_compte='charge'),
        CompteComptable(numero='664', intitule='Charges sociales salariales', classe=6, type_compte='charge'),
        CompteComptable(numero='67', intitule='Charges financières', classe=6, type_compte='charge'),
        CompteComptable(numero='671', intitule='Intérêts des emprunts', classe=6, type_compte='charge'),
        CompteComptable(numero='672', intitule='Pertes de change', classe=6, type_compte='charge'),
        CompteComptable(numero='68', intitule='Dotations amortissements et provisions', classe=6, type_compte='charge'),
        CompteComptable(numero='681', intitule='Dotations aux amortissements', classe=6, type_compte='charge'),
        CompteComptable(numero='682', intitule='Dotations aux provisions', classe=6, type_compte='charge'),
        CompteComptable(numero='69', intitule='Charges exceptionnelles', classe=6, type_compte='charge'),

        # Classe 7 - Produits
        CompteComptable(numero='70', intitule='Ventes et prestations', classe=7, type_compte='produit'),
        CompteComptable(numero='701', intitule='Ventes de services', classe=7, type_compte='produit'),
        CompteComptable(numero='702', intitule='Prestations de conseil', classe=7, type_compte='produit'),
        CompteComptable(numero='74', intitule='Subventions d\'exploitation', classe=7, type_compte='produit'),
        CompteComptable(numero='741', intitule='Subventions projets', classe=7, type_compte='produit'),
        # Sous-comptes par projet
        CompteComptable(numero='7411', intitule='Subvention projet LED', classe=7, type_compte='produit'),
        CompteComptable(numero='7412', intitule='Subvention projet SOR4D', classe=7, type_compte='produit'),
        CompteComptable(numero='7413', intitule='Subvention projet AMSANA', classe=7, type_compte='produit'),
        CompteComptable(numero='7419', intitule='Subventions autres projets', classe=7, type_compte='produit'),
        CompteComptable(numero='742', intitule='Subventions fonctionnement', classe=7, type_compte='produit'),
        CompteComptable(numero='75', intitule='Autres produits', classe=7, type_compte='produit'),
        CompteComptable(numero='751', intitule='Produits accessoires', classe=7, type_compte='produit'),
        CompteComptable(numero='76', intitule='Produits financiers', classe=7, type_compte='produit'),
        CompteComptable(numero='761', intitule='Intérêts bancaires', classe=7, type_compte='produit'),
        CompteComptable(numero='762', intitule='Gains de change', classe=7, type_compte='produit'),
        CompteComptable(numero='77', intitule='Produits exceptionnels', classe=7, type_compte='produit'),
        CompteComptable(numero='78', intitule='Reprises amortissements et provisions', classe=7, type_compte='produit'),
        CompteComptable(numero='79', intitule='Transferts de charges', classe=7, type_compte='produit'),
    ]
    db.session.add_all(comptes)

    # Journaux comptables
    journaux = [
        Journal(code='AC', nom='Journal des Achats', type_journal='achat'),
        Journal(code='BQ', nom='Journal de Banque', type_journal='banque'),
        Journal(code='CA', nom='Journal de Caisse', type_journal='caisse'),
        Journal(code='OD', nom='Opérations Diverses', type_journal='od'),
        Journal(code='SAL', nom='Journal des Salaires', type_journal='od'),
    ]
    db.session.add_all(journaux)

    # Catégories budgétaires (basées sur vos budgets)
    categories = [
        CategorieBudget(code='LABOR', nom='Personnel / Salaires', ordre=1),
        CategorieBudget(code='TRAVEL', nom='Voyages et Déplacements', ordre=2),
        CategorieBudget(code='SUPPLIES', nom='Fournitures et Équipements', ordre=3),
        CategorieBudget(code='PROGRAM', nom='Coûts Programmes / Activités', ordre=4),
        CategorieBudget(code='ADMIN', nom='Frais Administratifs', ordre=5),
        CategorieBudget(code='OVERHEAD', nom='Frais Généraux / Indirect', ordre=6),
        CategorieBudget(code='AUDIT', nom='Audit et Évaluation', ordre=7),
    ]
    db.session.add_all(categories)

    # Créer un utilisateur administrateur par défaut
    admin = Utilisateur(
        email='admin@creates.sn',
        nom='Administrateur',
        prenom='CREATES',
        password_hash=generate_password_hash('admin123'),  # Mot de passe à changer!
        role='directeur',
        actif=True,
        created_by='system'
    )
    db.session.add(admin)

    db.session.commit()
    print("Base de données initialisée avec succès!")
    print("Utilisateur admin créé: admin@creates.sn / admin123 (à changer!)")


# =============================================================================
# MAIN
# =============================================================================

if __name__ == '__main__':
    with app.app_context():
        init_db()
    # SECURITY: Debug mode disabled in production
    debug_mode = os.environ.get('FLASK_DEBUG', 'false').lower() == 'true'
    port = int(os.environ.get('PORT', 5001))
    app.run(debug=debug_mode, port=port)
